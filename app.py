import pandas as pd
import datetime
import streamlit as st
st.set_page_config(page_title="Course Scheduler", layout="wide")
import matplotlib.pyplot as plt
from ortools.sat.python import cp_model
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import threading
import time
import os

# Set up logging at the beginning of your app.py
logging.basicConfig(level=logging.ERROR)


# Import template generator
from template_generator import create_excel_template

class CourseScheduler:
    def __init__(self):
        self.course_run_data = None
        self.consultant_data = None
        self.priority_data = None
        self.annual_leaves = None
        self.affinity_matrix_data = None
        self.public_holidays_data = None
        self.weekly_calendar = None
        self.week_to_month_map = {}
        self.weekly_working_days = {}
        self.fleximatrix = {}  # Dictionary mapping (course, language) to list of qualified trainers
        self.course_champions = {}  # Dictionary to store champions for each course

    def load_data_from_excel(self, uploaded_file):
        """Load all required data from an Excel file with multiple sheets"""
        try:
            # Read the Excel file with multiple sheets
            excel_file = pd.ExcelFile(uploaded_file)

            # Load each required sheet
            self.course_run_data = pd.read_excel(excel_file, sheet_name='CourseData')
            self.consultant_data = pd.read_excel(excel_file, sheet_name='TrainerData')
            self.priority_data = pd.read_excel(excel_file, sheet_name='PriorityData')
            self.annual_leaves = pd.read_excel(excel_file, sheet_name='AnnualLeaves')
            self.affinity_matrix_data = pd.read_excel(excel_file, sheet_name='AffinityMatrix')
            self.public_holidays_data = pd.read_excel(excel_file, sheet_name='PublicHolidays')

            # Parse date columns
            date_columns = ['Start_Date', 'End_Date']
            for col in date_columns:
                if col in self.annual_leaves.columns:
                    self.annual_leaves[col] = pd.to_datetime(self.annual_leaves[col]).dt.date

            if 'Start Date' in self.public_holidays_data.columns:
                self.public_holidays_data['Start Date'] = pd.to_datetime(
                    self.public_holidays_data['Start Date']).dt.date

            if 'End Date' in self.public_holidays_data.columns:
                self.public_holidays_data['End Date'] = pd.to_datetime(self.public_holidays_data['End Date']).dt.date

            # Load fleximatrix - which trainers can teach which courses (wide format)
            flexi_sheet = pd.read_excel(excel_file, sheet_name='Fleximatrix')

            # Process the wide-format Fleximatrix
            self.fleximatrix, self.course_champions = self.process_wide_fleximatrix(flexi_sheet)

            # Load adjusted monthly demand data
            self.monthly_demand = pd.read_excel(excel_file, sheet_name='MonthlyDemand')

            # Load course week restrictions (if exists)
            self.week_restrictions = {}
            try:
                # Check if the sheet exists
                restrictions_sheet = pd.read_excel(excel_file, sheet_name='WeekRestrictions')

                # Process each restriction
                for _, row in restrictions_sheet.iterrows():
                    course = row['Course']
                    week_type = row['Week Type']  # 'First', 'Second', 'Third', 'Fourth', 'Last'
                    restricted = row['Restricted']  # Boolean: True = cannot run, False = can run

                    if course not in self.week_restrictions:
                        self.week_restrictions[course] = {}

                    self.week_restrictions[course][week_type] = restricted

                print(f"Loaded {len(restrictions_sheet)} week restrictions for courses")
            except Exception as e:
                print(f"No week restrictions found or error loading them: {e}")
                self.week_restrictions = {}

            return True

        except Exception as e:
            st.error(f"Error loading data: {e}")
            return False

    def process_wide_fleximatrix(self, flexi_sheet):
        """Process a wide-format Fleximatrix where consultants are columns"""
        fleximatrix = {}
        course_champions = {}

        # Get list of consultants (all columns except CourseName, CategoryName, Language, and Champion)
        consultants = [col for col in flexi_sheet.columns
                       if col not in ['CourseName', 'CategoryName', 'Language', 'Champion']]

        # Process each row (course+language combination)
        for _, row in flexi_sheet.iterrows():
            course = row['CourseName']
            language = row['Language']
            champion = row['Champion']

            # Initialize empty list for this course+language
            fleximatrix[(course, language)] = []

            # Add the champion
            if champion and str(champion).strip():
                course_champions[(course, language)] = champion

            # Add all qualified consultants (marked with "U")
            for consultant in consultants:
                if row[consultant] == "U":
                    fleximatrix[(course, language)].append(consultant)

        return fleximatrix, course_champions


    def initialize_calendar(self, year, weekend_selection):
        """Initialize the weekly calendar based on year and weekend selection"""

        # Function to Generate Work Week Start Dates
        def get_week_start_dates(year, weekend_selection):
            first_day = datetime.date(year, 1, 1)
            if weekend_selection == "FS":  # Friday-Saturday
                week_start = 6  # Sunday
            else:  # "Saturday-Sunday"
                week_start = 0  # Monday

            weeks = []
            current_date = first_day

            # Find the first day that matches our week start
            while current_date.weekday() != week_start:
                current_date += datetime.timedelta(days=1)

            # Generate all week start dates for the year
            while current_date.year == year:
                weeks.append(current_date)
                current_date += datetime.timedelta(days=7)  # Move to next week

            return weeks

        # Generate Work Week Start Dates
        self.weekly_calendar = get_week_start_dates(year, weekend_selection)

        # Ensure correct mapping of weeks to months
        self.week_to_month_map = {}
        self.week_position_in_month = {}  # Track position of week within its month

        # First, map weeks to months
        for i, week_start in enumerate(self.weekly_calendar):
            month = week_start.month  # Extract the month from the start date
            self.week_to_month_map[i + 1] = month  # Map week number to month

        # Then, calculate position of each week within its month
        for week_num, month in self.week_to_month_map.items():
            # Get all weeks in this month
            month_weeks = [w for w, m in self.week_to_month_map.items() if m == month]
            month_weeks.sort()

            # Find position (1-indexed)
            position = month_weeks.index(week_num) + 1

            # Also mark if it's the last week of the month
            is_last = (position == len(month_weeks))

            # Store as a dictionary with position info
            self.week_position_in_month[week_num] = {
                'position': position,
                'total_weeks': len(month_weeks),
                'is_first': (position == 1),
                'is_second': (position == 2),
                'is_third': (position == 3),
                'is_fourth': (position == 4),
                'is_last': is_last
            }

        # Calculate working days for each week accounting for public holidays
        self.calculate_working_days()

        return True


    def calculate_working_days(self):
        """Calculate working days for each week accounting for public holidays"""
        self.weekly_working_days = {}

        for i, week_start in enumerate(self.weekly_calendar):
            week_num = i + 1
            working_days = 5  # Assume a full workweek

            # Check if this week is completely inside a long holiday
            fully_inside_long_holiday = False

            for _, row in self.public_holidays_data.iterrows():
                holiday_length = (row["End Date"] - row["Start Date"]).days + 1

                # If it's a long holiday (25+ days) and the week is fully inside it
                if holiday_length >= 25:
                    if week_start >= row["Start Date"] and (week_start + datetime.timedelta(days=4)) <= row["End Date"]:
                        fully_inside_long_holiday = True
                        break

            # If fully inside a long holiday, block the week completely
            if fully_inside_long_holiday:
                working_days = 0
            else:
                # Adjust only affected days for regular holidays
                for _, row in self.public_holidays_data.iterrows():
                    holiday_days_in_week = sum(
                        1 for d in range(5)  # Check only workdays
                        if row["Start Date"] <= (week_start + datetime.timedelta(days=d)) <= row["End Date"]
                        and (week_start + datetime.timedelta(days=d)).weekday() not in [5, 6]  # Exclude weekends
                    )
                    working_days -= holiday_days_in_week
                    working_days = max(0, working_days)  # Ensure no negative values

            self.weekly_working_days[week_num] = working_days


    def is_trainer_available(self, trainer_name, week_num):
        """Check if trainer is available during the given week (not on leave)"""
        week_start = self.weekly_calendar[week_num - 1]
        week_end = week_start + datetime.timedelta(days=4)  # Assuming 5-day courses

        # Check against all leave periods
        for _, leave in self.annual_leaves[self.annual_leaves["Name"] == trainer_name].iterrows():
            # If leave period overlaps with course week, trainer is unavailable
            if (leave["Start_Date"] <= week_end and leave["End_Date"] >= week_start):
                return False
        return True


    def get_trainer_priority(self, trainer_name):
        """Get priority level for a trainer based on their title"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        priority = self.priority_data.loc[self.priority_data["Title"] == title, "Priority"].iloc[0]
        return priority


    def is_freelancer(self, trainer_name):
        """Check if a trainer is a freelancer"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        return title == "Freelancer"

    def run_optimization(self, monthly_weight=5, 
                         affinity_weight=2,
                         utilization_target=70, solver_time_minutes=5,
                         num_workers=8, min_course_spacing=2,
                         solution_strategy="BALANCED",
                         enforce_monthly_distribution=False,
                         max_affinity_constraints=50,
                         prioritize_all_courses=False,
                         accelerated_mode=False,
                         enforce_champions=True,
                         progress_callback=None):

        # Helper function to log and update progress
        def log_progress(message, percent=None):
            print(message)
            if progress_callback:
                # Always default to the most recent percent if none is provided
                if percent is None:
                    # Don't update progress bar if percent is None
                    progress_callback(None, message)
                else:
                    progress_callback(percent, message)
        
        # Apply accelerated mode optimizations if enabled
        if accelerated_mode:
            log_progress("Running in accelerated mode - applying speed optimizations...", 0.01)
            
            # Check model size to determine level of acceleration needed
            total_courses = sum(self.course_run_data["Runs"])
            
            if total_courses > 500:
                log_progress(f"Extremely large problem detected ({total_courses} courses) - applying maximum optimizations", 0.02)
                # Maximum optimizations for extremely large problems
                max_affinity_constraints = 0  # Disable affinity constraints completely
                solution_strategy = "FIND_FEASIBLE_FAST"
                num_workers = min(32, os.cpu_count() or 8)  # Use all available CPU cores if possible
                monthly_weight = 1  # Minimal monthly distribution weight
                # Drastically shorten time limit
                solver_time_minutes = min(solver_time_minutes, 5)
                
                # Apply sampling for extremely large problems - only optimize a subset of courses
                if total_courses > 700:
                    log_progress("Applying course sampling for enormous problem - will optimize in batches", 0.03)
                    # We'll need to sample courses by sorting by priority and taking most important ones
                    self.course_run_data = self._sample_courses_for_optimization(self.course_run_data, max_courses=500)
                    log_progress(f"Reduced to {sum(self.course_run_data['Runs'])} highest priority courses", 0.04)
                
                log_progress("Maximum optimizations applied - sacrificing constraint quality for speed", 0.05)
            elif total_courses > 300:
                log_progress(f"Large problem detected ({total_courses} courses) - applying aggressive optimizations", 0.02)
                # Very aggressive settings for large problems
                max_affinity_constraints = max(3, max_affinity_constraints // 8)
                solution_strategy = "FIND_FEASIBLE_FAST"
                num_workers = min(16, num_workers * 2)
                monthly_weight = monthly_weight // 4
                min_course_spacing = 1  # Minimum possible spacing
                # Shorten time limit to avoid excessive solving
                solver_time_minutes = min(solver_time_minutes, 10)
                log_progress("Aggressive optimizations applied for large problem", 0.03)
            else:
                # Standard optimizations for smaller problems
                max_affinity_constraints = max(5, max_affinity_constraints // 4)
                solution_strategy = "FIND_FEASIBLE_FAST"
                num_workers = min(16, num_workers * 2)
                monthly_weight = monthly_weight // 2
                min_course_spacing = max(1, min_course_spacing - 1)

        # Get total F2F runs
        log_progress("Starting optimization process...", 0.01)
        total_f2f_runs = sum(self.course_run_data["Runs"])
        log_progress(f"Total F2F runs: {total_f2f_runs}", 0.03)

        # Create adjusted monthly demand dictionary from the demand dataframe
        log_progress("Calculating monthly demand distribution...", 0.05)
        # Convert percentages to actual course counts
        adjusted_f2f_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage  # Normalize to ensure percentages sum to 1
            # Calculate how many courses should be in this month (round to nearest integer)
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Adjust rounding errors to ensure total matches
        total_allocated = sum(adjusted_f2f_demand.values())
        if total_allocated != total_f2f_runs:
            # Find the month with the highest demand and adjust
            max_month = max(adjusted_f2f_demand.items(), key=lambda x: x[1])[0]
            adjusted_f2f_demand[max_month] += (total_f2f_runs - total_allocated)

        log_progress("Monthly demand calculation completed", 0.10)
        for month, demand in sorted(adjusted_f2f_demand.items()):
            log_progress(f"Month {month}: {demand} courses")

        # Initialize model and variables
        log_progress("Initializing optimization model...", 0.12)
        model = cp_model.CpModel()
        schedule = {}
        trainer_assignments = {}
        max_weeks = len(self.weekly_calendar)

        log_progress("Creating variables for courses and trainers...", 0.15)
        
        # All penalty lists
        month_deviation_penalties = []
        affinity_penalties = []
        trainer_utilization_penalties = []
        champion_assignment_penalties = []
        unscheduled_course_penalties = []  # New list for unscheduled course penalties

        # Create the schedule variables for each course and run
        for _, row in self.course_run_data.iterrows():
            try:
                course, delivery_type, language, runs, duration = row["Course Name"], row["Delivery Type"], row["Language"], \
                    row["Runs"], row["Duration"]

                for i in range(runs):
                    # Create a variable for the start week
                    start_week = model.NewIntVar(1, max_weeks, f"start_week_{course}_{i}")
                    schedule[(course, delivery_type, language, i)] = start_week

                    # If champions are prioritized and there is a champion for this course, 
                    # prepare to add constraints after checking week availability
                    use_champion = False
                    champion_idx = -1
                    
                    if enforce_champions and (course, language) in self.course_champions:
                        champion = self.course_champions.get((course, language))
                        if champion in self.fleximatrix.get((course, language), []):
                            # Get the index of the champion in the qualified trainers list
                            champion_idx = self.fleximatrix.get((course, language)).index(champion)
                            use_champion = True
                            log_progress(f"Will prioritize champion {champion} for {course} ({language}) when available")
                    
                    # Create normal trainer assignment variable (we'll add champion constraints later)
                    trainer_var = model.NewIntVar(0, len(self.fleximatrix.get((course, language), [])), f"trainer_{course}_{i}")
                    trainer_assignments[(course, delivery_type, language, i)] = trainer_var
                    
                    # Only schedule in weeks with enough working days AND available trainers
                    valid_weeks = []
                    for w, days in self.weekly_working_days.items():
                        if days >= duration:
                            # WEEK RESTRICTION CHECK: Skip if this course can't run in this week position
                            if course in self.week_restrictions:
                                # Get week position info
                                week_info = self.week_position_in_month.get(w, {})

                                # Check each restriction type
                                skip_week = False

                                if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                                    skip_week = True
                                elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                                    skip_week = True
                                elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                                    skip_week = True
                                elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                                    skip_week = True
                                elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                                    skip_week = True

                                if skip_week:
                                    # Skip this week for this course due to restriction
                                    continue

                            # Check if at least one qualified trainer is available this week
                            for trainer in self.fleximatrix.get((course, language), []):
                                if self.is_trainer_available(trainer, w):
                                    valid_weeks.append(w)
                                    break

                    if valid_weeks:
                        model.AddAllowedAssignments([start_week], [[w] for w in valid_weeks])
                        
                        # Now add champion constraints for each valid week if needed
                        if use_champion and champion_idx >= 0:
                            champion = self.fleximatrix.get((course, language))[champion_idx]
                            
                            # For each valid week, check if the champion is available
                            for week in valid_weeks:
                                if self.is_trainer_available(champion, week):
                                    # Create a boolean variable that is true if this course is scheduled in this week
                                    is_in_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                                    model.Add(start_week == week).OnlyEnforceIf(is_in_week)
                                    model.Add(start_week != week).OnlyEnforceIf(is_in_week.Not())
                                    
                                    # If the course is scheduled in this week, the trainer must be the champion
                                    model.Add(trainer_var == champion_idx).OnlyEnforceIf(is_in_week)
                                    
                                    log_progress(f"Added constraint: If {course} run {i+1} is in week {week}, champion {champion} must teach it", None)
                                    
                    else:
                        print(f"Warning: No valid weeks for {course} run {i + 1}")
            except Exception as e:
                print(f"Error processing course {course}: {e}")
                continue

        log_progress("Adding monthly distribution constraints...", 0.30)
        # CONSTRAINT 1: Track which month each course is assigned to and enforce monthly distribution
        print("Adding fixed monthly distribution constraints")

        # For each month, explicitly track which courses are scheduled in it
        # In your run_optimization method, after you've calculated target_demand:

        # For each month, track and enforce
        for month in range(1, 13):
            target_demand = adjusted_f2f_demand.get(month, 0)

            # Create a list of all Boolean variables indicating if a course is scheduled in this month
            courses_in_month = []

            # Get all weeks that belong to this month
            month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]

            if not month_weeks:  # Skip months with no weeks
                continue

            # For each course run, create a Boolean variable indicating if it's in this month
            for (course, delivery_type, language, i), week_var in schedule.items():
                is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")

                # Add constraints: is_in_month is True if and only if week_var is one of month's weeks
                week_choices = []
                for week in month_weeks:
                    is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                    week_choices.append(is_in_this_week)

                model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())

                courses_in_month.append(is_in_month)

            # CORE CONSTRAINT: Monthly distribution
            if courses_in_month and target_demand > 0:  # Only if we have variables for this month
                if enforce_monthly_distribution:
                    # Hard constraint
                    model.Add(sum(courses_in_month) == target_demand)
                    print(f"  Month {month}: Enforcing exactly {target_demand} courses (hard constraint)")
                else:
                    # Soft constraint with penalties
                    month_deviation = model.NewIntVar(0, total_f2f_runs, f"month_{month}_deviation")

                    # Set it equal to the absolute difference between actual and target
                    actual_courses = sum(courses_in_month)

                    # We need to model absolute value: |actual - target|
                    # First, create variable for whether actual >= target
                    is_over_target = model.NewBoolVar(f"month_{month}_over_target")
                    model.Add(actual_courses >= target_demand).OnlyEnforceIf(is_over_target)
                    model.Add(actual_courses < target_demand).OnlyEnforceIf(is_over_target.Not())

                    # If actual >= target, deviation = actual - target
                    over_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_over_dev")
                    model.Add(over_dev == actual_courses - target_demand).OnlyEnforceIf(is_over_target)
                    model.Add(over_dev == 0).OnlyEnforceIf(is_over_target.Not())

                    # If actual < target, deviation = target - actual
                    under_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_under_dev")
                    model.Add(under_dev == target_demand - actual_courses).OnlyEnforceIf(is_over_target.Not())
                    model.Add(under_dev == 0).OnlyEnforceIf(is_over_target)

                    # Total deviation is sum of over and under deviations
                    model.Add(month_deviation == over_dev + under_dev)

                    # Add penalties for deviation (using the monthly weight parameter)
                    for _ in range(monthly_weight):
                        month_deviation_penalties.append(month_deviation)

                    print(f"  Month {month}: Target of {target_demand} courses (soft constraint with penalty)")

        # CONSTRAINT 2: Minimum spacing between runs of same course
        print("Adding spacing between runs of the same course")

        for course_name in set(self.course_run_data["Course Name"]):
            course_runs = []
            for (course, delivery_type, language, i), var in schedule.items():
                if course == course_name:
                    course_runs.append((i, var))

            # FIX: Sort by the first element of the tuple (i) instead of trying to compare tuples
            # This was the source of the error - we need to sort by run index only
            course_runs.sort(key=lambda x: x[0])

            for i in range(len(course_runs) - 1):
                run_num1, var1 = course_runs[i]
                run_num2, var2 = course_runs[i + 1]

                # Hard constraint for minimum spacing (using the parameter)
                model.Add(var2 >= var1 + min_course_spacing)

        # CONSTRAINT 3: Add constraints for course affinities
        log_progress("Adding affinity constraints for course pairs...", 0.50)
        print(f"Adding affinity constraints for course pairs")

        # Define maximum affinity constraints to add (parameter)
        affinity_count = 0
        # Add constraints for course affinities (soft constraints/penalties)
        for _, row in self.affinity_matrix_data.iterrows():
            if affinity_count >= max_affinity_constraints:
                break

            c1, c2, gap_weeks = row["Course 1"], row["Course 2"], row["Gap Weeks"]

            c1_runs = []
            c2_runs = []

            for (course, _, _, i), var in schedule.items():
                if course == c1:
                    c1_runs.append((i, var))
                elif course == c2:
                    c2_runs.append((i, var))

            # Skip if either course doesn't have any runs
            if not c1_runs or not c2_runs:
                continue

            # Sort by run index
            c1_runs.sort(key=lambda x: x[0])
            c2_runs.sort(key=lambda x: x[0])

            affinity_count += 1

            # Only check first run of each course to reduce constraints
            run1, var1 = c1_runs[0]
            run2, var2 = c2_runs[0]

            # Create variables to check if either course is in Q4
            is_c1_q4 = model.NewBoolVar(f"{c1}_{run1}_in_q4")
            is_c2_q4 = model.NewBoolVar(f"{c2}_{run2}_in_q4")

            # Get Q4 weeks
            q4_weeks = [w for w, m in self.week_to_month_map.items() if m in [10, 11, 12]]

            # Set up Q4 detection for course 1
            c1_q4_choices = []
            for week in q4_weeks:
                is_in_this_week = model.NewBoolVar(f"{c1}_{run1}_in_week_{week}")
                model.Add(var1 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var1 != week).OnlyEnforceIf(is_in_this_week.Not())
                c1_q4_choices.append(is_in_this_week)

            model.AddBoolOr(c1_q4_choices).OnlyEnforceIf(is_c1_q4)
            model.AddBoolAnd([choice.Not() for choice in c1_q4_choices]).OnlyEnforceIf(is_c1_q4.Not())

            # Set up Q4 detection for course 2
            c2_q4_choices = []
            for week in q4_weeks:
                is_in_this_week = model.NewBoolVar(f"{c2}_{run2}_in_week_{week}")
                model.Add(var2 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var2 != week).OnlyEnforceIf(is_in_this_week.Not())
                c2_q4_choices.append(is_in_this_week)

            model.AddBoolOr(c2_q4_choices).OnlyEnforceIf(is_c2_q4)
            model.AddBoolAnd([choice.Not() for choice in c2_q4_choices]).OnlyEnforceIf(is_c2_q4.Not())

            # Either course is in Q4
            either_in_q4 = model.NewBoolVar(f"either_{c1}_{c2}_in_q4")
            model.AddBoolOr([is_c1_q4, is_c2_q4]).OnlyEnforceIf(either_in_q4)
            model.AddBoolAnd([is_c1_q4.Not(), is_c2_q4.Not()]).OnlyEnforceIf(either_in_q4.Not())

            # Soft affinity constraint with reduced gap for Q4
            too_close = model.NewBoolVar(f"affinity_too_close_{c1}_{c2}_{run1}_{run2}")

            # Regular gap weeks for non-Q4
            far_enough_after_normal = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_normal")
            far_enough_before_normal = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_normal")

            # Reduced gap weeks (50% reduction) for Q4
            reduced_gap = max(1, gap_weeks // 2)  # Ensure minimum 1 week gap
            far_enough_after_q4 = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_q4")
            far_enough_before_q4 = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_q4")

            # Normal gap constraints
            model.Add(var2 >= var1 + gap_weeks).OnlyEnforceIf(far_enough_after_normal)
            model.Add(var2 < var1 + gap_weeks).OnlyEnforceIf(far_enough_after_normal.Not())

            model.Add(var1 >= var2 + gap_weeks).OnlyEnforceIf(far_enough_before_normal)
            model.Add(var1 < var2 + gap_weeks).OnlyEnforceIf(far_enough_before_normal.Not())

            # Q4 gap constraints (reduced)
            model.Add(var2 >= var1 + reduced_gap).OnlyEnforceIf(far_enough_after_q4)
            model.Add(var2 < var1 + reduced_gap).OnlyEnforceIf(far_enough_after_q4.Not())

            model.Add(var1 >= var2 + reduced_gap).OnlyEnforceIf(far_enough_before_q4)
            model.Add(var1 < var2 + reduced_gap).OnlyEnforceIf(far_enough_before_q4.Not())

            # Logic for too_close based on regular or Q4 gaps
            # Not too close if:
            # (NOT in Q4 AND (far enough apart in either direction))
            # OR
            # (In Q4 AND (far enough apart with reduced gap in either direction))
            far_enough_normal = model.NewBoolVar(f"far_enough_{c1}_{c2}_{run1}_{run2}_normal")
            model.AddBoolOr([far_enough_after_normal, far_enough_before_normal]).OnlyEnforceIf(far_enough_normal)
            model.AddBoolAnd([far_enough_after_normal.Not(), far_enough_before_normal.Not()]).OnlyEnforceIf(
                far_enough_normal.Not())

            far_enough_q4 = model.NewBoolVar(f"far_enough_{c1}_{c2}_{run1}_{run2}_q4")
            model.AddBoolOr([far_enough_after_q4, far_enough_before_q4]).OnlyEnforceIf(far_enough_q4)
            model.AddBoolAnd([far_enough_after_q4.Not(), far_enough_before_q4.Not()]).OnlyEnforceIf(far_enough_q4.Not())

            # Not too close if appropriate gap is maintained based on Q4 status
            not_too_close_normal = model.NewBoolVar(f"not_too_close_normal_{c1}_{c2}_{run1}_{run2}")
            model.AddBoolAnd([either_in_q4.Not(), far_enough_normal]).OnlyEnforceIf(not_too_close_normal)
            model.AddBoolOr([either_in_q4, far_enough_normal.Not()]).OnlyEnforceIf(not_too_close_normal.Not())

            not_too_close_q4 = model.NewBoolVar(f"not_too_close_q4_{c1}_{c2}_{run1}_{run2}")
            model.AddBoolAnd([either_in_q4, far_enough_q4]).OnlyEnforceIf(not_too_close_q4)
            model.AddBoolOr([either_in_q4.Not(), far_enough_q4.Not()]).OnlyEnforceIf(not_too_close_q4.Not())

            # Final too_close logic
            model.AddBoolOr([not_too_close_normal, not_too_close_q4]).OnlyEnforceIf(too_close.Not())
            model.AddBoolAnd([not_too_close_normal.Not(), not_too_close_q4.Not()]).OnlyEnforceIf(too_close)

            # Add penalty for being too close
            for _ in range(affinity_weight):
                affinity_penalties.append(too_close)

            print(f"  Added affinity constraint: {c1} and {c2} should be {gap_weeks} weeks apart (reduced to {reduced_gap} in Q4)")
            
            # Log each affinity constraint with shortened course names for readability
            c1_short = c1[:40] + "..." if len(c1) > 40 else c1
            c2_short = c2[:40] + "..." if len(c2) > 40 else c2
            log_message = f"Added affinity constraint: {c1_short} and {c2_short} should be {gap_weeks} weeks apart (reduced to {reduced_gap} in Q4)"
            log_progress(log_message)

        # CONSTRAINT 4: Trainer-specific constraints
        log_progress("Adding trainer assignment constraints", 0.55)
        log_progress("Adding constraint: trainer can only teach one course per week", 0.56)

        # 4.1: Trainer availability - only assign trainers who are available during the scheduled week
        for (course, delivery_type, language, i), week_var in schedule.items():
            trainer_var = trainer_assignments.get((course, delivery_type, language, i))
            if trainer_var is None:
                continue

            qualified_trainers = self.fleximatrix.get((course, language), [])

            # For each week, set which trainers are available
            for week in range(1, max_weeks + 1):
                is_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                model.Add(week_var == week).OnlyEnforceIf(is_this_week)
                model.Add(week_var != week).OnlyEnforceIf(is_this_week.Not())

                # For each trainer, check availability
                for t_idx, trainer in enumerate(qualified_trainers):
                    if not self.is_trainer_available(trainer, week):
                        # If trainer is not available this week, cannot select this trainer
                        is_this_trainer = model.NewBoolVar(f"{course}_{i}_trainer_{t_idx}_week_{week}")
                        model.Add(trainer_var == t_idx).OnlyEnforceIf(is_this_trainer)
                        model.Add(trainer_var != t_idx).OnlyEnforceIf(is_this_trainer.Not())

                        # Cannot have both is_this_week and is_this_trainer be true
                        model.AddBoolOr([is_this_week.Not(), is_this_trainer.Not()])

        # If prioritize_all_courses is true, add that message
        if prioritize_all_courses:
            log_progress("Adding priority for scheduling all courses", 0.58)

        # 4.2: Workload limits - track and limit total days per trainer
        trainer_workload = {name: [] for name in self.consultant_data["Name"]}

        # Track course assignments for each trainer
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            qualified_trainers = self.fleximatrix.get((course, language), [])

            for t_idx, trainer in enumerate(qualified_trainers):
                is_assigned = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}")
                model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned)
                model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned.Not())

                # Accumulate workload
                trainer_workload[trainer].append((is_assigned, duration))

                # 4.3: Champion priority - add penalty for not using the champion for their courses
                champion = self.course_champions.get((course, language))
                if champion == trainer:
                    not_using_champion = model.NewBoolVar(f"not_using_champion_{course}_{i}")
                    model.Add(trainer_var != t_idx).OnlyEnforceIf(not_using_champion)
                    model.Add(trainer_var == t_idx).OnlyEnforceIf(not_using_champion.Not())
                    champion_assignment_penalties.append(not_using_champion)

        # Add constraints for maximum workload
        for trainer, workload_items in trainer_workload.items():
            if not workload_items:
                continue

            max_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Max_Days"].iloc[0]

            # Calculate total workload
            total_workload = model.NewIntVar(0, max_days, f"total_workload_{trainer}")
            weighted_sum = []
            for is_assigned, duration in workload_items:
                weighted_sum.append((is_assigned, duration))

            if weighted_sum:
                weighted_terms = []
                for is_assigned, duration in weighted_sum:
                    term = model.NewIntVar(0, duration, f"term_{id(is_assigned)}_{duration}")
                    model.Add(term == duration).OnlyEnforceIf(is_assigned)
                    model.Add(term == 0).OnlyEnforceIf(is_assigned.Not())
                    weighted_terms.append(term)
                model.Add(total_workload == sum(weighted_terms))
                # Enforce maximum workload
                model.Add(total_workload <= max_days)
                
                # Add hard constraint for minimum utilization if specified
                min_days = max(1, int(max_days * utilization_target / 100))  # Convert percentage to days
                if min_days > 0:
                    # Commenting out the hard constraint as requested
                    # model.Add(total_workload >= min_days)
                    log_progress(f"Skipping minimum utilization constraint for Trainer {trainer} (would be {min_days} days)")

        # NEW CODE: Add constraint to prevent trainers from teaching multiple courses in same week
        # Add this right after the trainer workload constraints
        print("Adding constraint: trainer can only teach one course per week")

        # For each trainer
        for trainer_name in self.consultant_data["Name"]:
            # For each week
            for week in range(1, max_weeks + 1):
                # Find all course runs that this trainer might teach in this week
                week_assignments = []

                for (course, delivery_type, language, i), week_var in schedule.items():
                    trainer_var = trainer_assignments.get((course, delivery_type, language, i))
                    if trainer_var is None:
                        continue

                    # Get qualified trainers for this course
                    qualified_trainers = self.fleximatrix.get((course, language), [])
                    if trainer_name not in qualified_trainers:
                        continue

                    # Get index of this trainer
                    t_idx = qualified_trainers.index(trainer_name)

                    # Create variable for "this course is taught by this trainer in this week"
                    is_assigned = model.NewBoolVar(f"{course}_{i}_{trainer_name}_week_{week}")

                    # is_assigned is true if and only if:
                    # 1. The course is scheduled in this week, AND
                    # 2. This trainer is assigned to it
                    is_this_week = model.NewBoolVar(f"{course}_{i}_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_this_week.Not())

                    is_this_trainer = model.NewBoolVar(f"{course}_{i}_trainer_{t_idx}")
                    model.Add(trainer_var == t_idx).OnlyEnforceIf(is_this_trainer)
                    model.Add(trainer_var != t_idx).OnlyEnforceIf(is_this_trainer.Not())

                    # Both conditions must be true for is_assigned to be true
                    model.AddBoolAnd([is_this_week, is_this_trainer]).OnlyEnforceIf(is_assigned)
                    model.AddBoolOr([is_this_week.Not(), is_this_trainer.Not()]).OnlyEnforceIf(is_assigned.Not())

                    week_assignments.append(is_assigned)

                # If there are multiple potential assignments for this trainer in this week
                if len(week_assignments) > 1:
                    # Add constraint: at most one assignment per trainer per week
                    model.Add(sum(week_assignments) <= 1)
        # If we're prioritizing all courses being scheduled, add variables to track this
        # Here's the fixed code section for the "prioritize_all_courses" functionality
        # This replaces the problem section in the run_optimization method

        if prioritize_all_courses:
            print("Adding priority for scheduling all courses")
            for (course, delivery_type, language, i), week_var in schedule.items():
                # Create a boolean variable that tracks if this course is scheduled
                is_scheduled = model.NewBoolVar(f"{course}_{i}_is_scheduled")

                # Create constraints: is_scheduled is True iff week_var is assigned to a valid week
                # First get all possible week values
                all_valid_weeks = []
                for w, days in self.weekly_working_days.items():
                    duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[
                        0]
                    if days >= duration:
                        # Check week restrictions
                        if course in self.week_restrictions:
                            week_info = self.week_position_in_month.get(w, {})
                            skip_week = False

                            if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                                skip_week = True
                            elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                                skip_week = True
                            elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                                skip_week = True
                            elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                                skip_week = True
                            elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                                skip_week = True

                            if skip_week:
                                continue

                        # Check if any trainer is available
                        qualified_trainers = self.fleximatrix.get((course, language), [])
                        for trainer in qualified_trainers:
                            if self.is_trainer_available(trainer, w):
                                all_valid_weeks.append(w)
                                break

                if all_valid_weeks:
                    # Course is scheduled if week_var equals any valid week
                    week_matches = []
                    for w in all_valid_weeks:
                        week_match = model.NewBoolVar(f"{course}_{i}_week_{w}")
                        model.Add(week_var == w).OnlyEnforceIf(week_match)
                        model.Add(week_var != w).OnlyEnforceIf(week_match.Not())
                        week_matches.append(week_match)

                    # If week_var matches any valid week, the course is scheduled
                    model.AddBoolOr(week_matches).OnlyEnforceIf(is_scheduled)
                    model.AddBoolAnd([match.Not() for match in week_matches]).OnlyEnforceIf(is_scheduled.Not())

                    # Create the not_scheduled variable and add it to penalties
                    not_scheduled = model.NewBoolVar(f"{course}_{i}_not_scheduled")

                    # This is the corrected line that replaces AddBoolNot
                    model.Add(is_scheduled == False).OnlyEnforceIf(not_scheduled)
                    model.Add(not_scheduled == False).OnlyEnforceIf(is_scheduled)

                    # Add to penalty list (with high weight)
                    unscheduled_course_penalties.append(not_scheduled)

        # 4. Modify the objective function to include unscheduled course penalties
        # Find the objective function (around line 835) and modify it:

        # Calculate combined penalty weights

        if prioritize_all_courses:
            # If prioritizing all courses, use a very high weight for unscheduled courses
            # and adjust other weights to allow more flexibility
            unscheduled_weight = 20  # Very high weight to make this the top priority
            monthly_weight = monthly_weight // 2 if enforce_monthly_distribution else monthly_weight
            min_course_spacing = max(1, min_course_spacing - 1)  # Reduce spacing requirements

            # Combined objective function with course scheduling priority
            model.Minimize(
                unscheduled_weight * sum(unscheduled_course_penalties) +
                monthly_weight * sum(month_deviation_penalties) +
                affinity_weight * sum(affinity_penalties)
            )
        else:
            # Original objective function without champion and utilization penalties
            model.Minimize(
                monthly_weight * sum(month_deviation_penalties) +
                affinity_weight * sum(affinity_penalties)
            )

        # Initialize solver with customized parameters
        solver = cp_model.CpSolver()
        
        try:
            # Base parameters that should work in all versions
            solver.parameters.max_time_in_seconds = solver_time_minutes * 60  # Convert to seconds
            solver.parameters.num_search_workers = num_workers
            
            # Try to set additional parameters that may only exist in newer versions
            try:
                solver.parameters.log_search_progress = True
            except:
                log_progress("Log search progress parameter not supported in this version", None)
        except Exception as e:
            log_progress(f"Warning: Some solver parameters could not be set: {str(e)}", None)
        
        # Create a simple solution counter callback instead of timeout
        class SolutionCounter(cp_model.CpSolverSolutionCallback):
            def __init__(self):
                cp_model.CpSolverSolutionCallback.__init__(self)
                self._solution_count = 0
                self._start_time = datetime.datetime.now()
                
            def on_solution_callback(self):
                try:
                    self._solution_count += 1
                    current_time = datetime.datetime.now()
                    elapsed_seconds = (current_time - self._start_time).total_seconds()
                    
                    if self._solution_count == 1:
                        log_progress(f"First solution found in {elapsed_seconds:.1f} seconds", 0.75)
                    elif self._solution_count % 10 == 0:  # Log every 10 solutions
                        log_progress(f"Found {self._solution_count} solutions in {elapsed_seconds:.1f} seconds", 0.80)
                except Exception as e:
                    # Catch any errors in the callback to prevent crashes
                    pass
                
            def solution_count(self):
                return self._solution_count
        
        log_progress(f"Configuring solver with {solver_time_minutes} minute time limit...", 0.60)
        
        # Create and use solution counter
        solution_counter = SolutionCounter()
        
        # Set solution strategy
        if solution_strategy == "MAXIMIZE_QUALITY":
            solver.parameters.optimize_with_max_hs = True
        elif solution_strategy == "FIND_FEASIBLE_FAST":
            solver.parameters.search_branching = cp_model.FIXED_SEARCH
            solver.parameters.optimize_with_core = False

        # Solve the model
        log_progress("Starting solver...", 0.65)
        
        # Create a separate thread to update status periodically
        def status_updater():
            start_time = datetime.datetime.now()
            update_interval = 10  # seconds
            while not solver_done:
                # Sleep for a bit
                time.sleep(update_interval)
                
                # Calculate elapsed time
                elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
                minutes = int(elapsed_time // 60)
                seconds = int(elapsed_time % 60)
                
                # Update status
                log_progress(f"Optimizer still running ({minutes}m {seconds}s elapsed, max {solver_time_minutes}m)...", 
                           0.70 + min(0.15, elapsed_time / (solver_time_minutes * 60) * 0.15))
                
                # Check if we're approaching the time limit
                if elapsed_time >= solver_time_minutes * 60 * 0.9:  # 90% of time limit
                    log_progress("Approaching time limit, preparing to return best solution found", 0.85)
                
                # Adjust update interval as time progresses
                update_interval = min(30, update_interval + 5)  # gradually increase interval up to 30 seconds
        
        # Flag to signal when solver is done
        solver_done = False
        
        # Start the status update thread
        status_thread = threading.Thread(target=status_updater)
        status_thread.daemon = True  # Thread will exit when main thread exits
        status_thread.start()
        
        try:
            # For extremely large problems, try a simpler approach first
            if accelerated_mode and sum(self.course_run_data["Runs"]) > 500:
                log_progress("Using simplified solver approach for extremely large problem", 0.70)
                # Don't use callback for very large problems - it can cause crashes
                status = solver.Solve(model)
            else:
                # Use callback for normal-sized problems
                status = solver.Solve(model, solution_counter)
        except Exception as e:
            log_progress(f"Error during solving: {str(e)}", 0.85)
            # Try again without the callback as a fallback
            log_progress("Retrying without custom callback...", 0.86)
            status = solver.Solve(model)
        finally:
            # Signal the status thread to stop
            solver_done = True
            # Wait for the thread to actually stop
            time.sleep(0.5)
            
        log_progress("Solver completed", 0.90)

        # Print status information
        print(f"Solver status: {solver.StatusName(status)}")
        print(f"Objective value: {solver.ObjectiveValue()}")
        print(f"Wall time: {solver.WallTime():.2f} seconds")

        print("Final Weekly Course Schedule with Trainers:")
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            schedule_results = []
            unscheduled_courses = []  # Track unscheduled courses
            total_courses = 0
            scheduled_courses = 0

            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    total_courses += 1
                    assigned_week = solver.Value(week_var)
                    if assigned_week > 0:  # Course was scheduled
                        scheduled_courses += 1
                        start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")

                        # Get trainer assignment if it exists
                        if (course, delivery_type, language, i) in trainer_assignments:
                            trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                            trainer_idx = solver.Value(trainer_var)
                            
                            # Check if this course has qualified trainers in the fleximatrix
                            if (course, language) in self.fleximatrix and len(self.fleximatrix[(course, language)]) > 0:
                                if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                                    trainer = self.fleximatrix[(course, language)][trainer_idx]
                                    is_champion = "✓" if self.course_champions.get((course, language)) == trainer else " "
                                else:
                                    # Handle out of range index
                                    trainer = "Unknown (Index Error)"
                                    is_champion = " "
                            else:
                                # Handle missing course in fleximatrix
                                trainer = "No Qualified Trainers"
                                is_champion = " "

                            schedule_results.append({
                                "Week": assigned_week,
                                "Start Date": start_date,
                                "Course": course,
                                "Delivery Type": delivery_type,
                                "Language": language,
                                "Run": i + 1,
                                "Trainer": trainer,
                                "Champion": is_champion
                            })

                    else:  # Course wasn't scheduled
                        unscheduled_courses.append({
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1
                        })
                except Exception as e:
                    print(f"Error processing result for {course} (run {i+1}): {e}")
                    # Add to unscheduled due to error
                    unscheduled_courses.append({
                        "Course": course,
                        "Delivery Type": delivery_type,
                        "Language": language,
                        "Run": i + 1,
                        "Error": str(e)
                    })

            # Convert to DataFrame
            schedule_df = pd.DataFrame(schedule_results)

            # Sort by week first, then by course name
            schedule_df = schedule_df.sort_values(by=["Week", "Course"])

            return status, schedule_df, solver, schedule, trainer_assignments, unscheduled_courses
        else:
            return status, None, solver, schedule, trainer_assignments, []

    def run_incremental_optimization(self, solver_time_minutes=5):
        """Run optimization by incrementally adding constraints to diagnose issues"""
        print("Starting incremental constraint optimization...")

        # Track feasibility at each step
        diagnostics = []

        # Initialize model and variables
        model = cp_model.CpModel()
        schedule = {}
        trainer_assignments = {}
        max_weeks = len(self.weekly_calendar)

        # Create the schedule variables for each course and run
        print("Step 1: Basic scheduling variables")
        for _, row in self.course_run_data.iterrows():
            course, delivery_type, language, runs, duration = row["Course Name"], row["Delivery Type"], row["Language"], \
                row["Runs"], row["Duration"]

            for i in range(runs):
                # Create a variable for the start week
                start_week = model.NewIntVar(1, max_weeks, f"start_week_{course}_{i}")
                schedule[(course, delivery_type, language, i)] = start_week

                # Create trainer assignment variable
                qualified_trainers = self.fleximatrix.get((course, language), [])
                if not qualified_trainers:
                    print(f"Warning: No qualified trainers for {course} ({language})")
                    continue

                trainer_var = model.NewIntVar(0, len(qualified_trainers) - 1, f"trainer_{course}_{i}")
                trainer_assignments[(course, delivery_type, language, i)] = trainer_var

        # Check feasibility with just variables
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 30  # Short timeout for each step
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Basic variables only",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        # STEP 2: Add minimum spacing between runs of same course
        print("Step 2: Adding course spacing constraints")
        for course_name in set(self.course_run_data["Course Name"]):
            course_runs = []
            for (course, delivery_type, language, i), var in schedule.items():
                if course == course_name:
                    course_runs.append((i, var))

            # Sort by run index
            course_runs.sort(key=lambda x: x[0])

            for i in range(len(course_runs) - 1):
                run_num1, var1 = course_runs[i]
                run_num2, var2 = course_runs[i + 1]

                # Hard constraint for minimum spacing
                model.Add(var2 >= var1 + 2)  # 2-week minimum spacing

        # Check feasibility with spacing constraints
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Course spacing constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible with just course spacing constraints!")
            return "INFEASIBLE", diagnostics, None

        # STEP 3: Add trainer availability constraints
        print("Step 3: Adding trainer availability constraints")
        for (course, delivery_type, language, i), week_var in schedule.items():
            trainer_var = trainer_assignments.get((course, delivery_type, language, i))
            if trainer_var is None:
                continue

            qualified_trainers = self.fleximatrix.get((course, language), [])

            # For each week, set which trainers are available
            for week in range(1, max_weeks + 1):
                is_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                model.Add(week_var == week).OnlyEnforceIf(is_this_week)
                model.Add(week_var != week).OnlyEnforceIf(is_this_week.Not())

                # For each trainer, check availability
                for t_idx, trainer in enumerate(qualified_trainers):
                    if not self.is_trainer_available(trainer, week):
                        # If trainer is not available this week, cannot select this trainer
                        is_this_trainer = model.NewBoolVar(f"{course}_{i}_trainer_{t_idx}_week_{week}")
                        model.Add(trainer_var == t_idx).OnlyEnforceIf(is_this_trainer)
                        model.Add(trainer_var != t_idx).OnlyEnforceIf(is_this_trainer.Not())

                        # Cannot have both is_this_week and is_this_trainer be true
                        model.AddBoolOr([is_this_week.Not(), is_this_trainer.Not()])

        # Check feasibility with trainer availability
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Trainer availability constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding trainer availability constraints!")
            return "INFEASIBLE", diagnostics, None

        # STEP 4: Add week restrictions
        print("Step 4: Adding week restriction constraints")
        for course in self.week_restrictions:
            for (c, delivery_type, language, i), week_var in schedule.items():
                if c != course:
                    continue

                # For each week, check if it's restricted
                for week in range(1, max_weeks + 1):
                    week_info = self.week_position_in_month.get(week, {})
                    skip_week = False

                    if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                        skip_week = True
                    elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                        skip_week = True
                    elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                        skip_week = True
                    elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                        skip_week = True
                    elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                        skip_week = True

                    if skip_week:
                        model.Add(week_var != week)

        # Check feasibility with week restrictions
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Week restriction constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding week restrictions!")
            return "INFEASIBLE", diagnostics, None

        # STEP 5: Add monthly distribution constraints
        print("Step 5: Adding monthly distribution constraints")

        # Get total F2F runs
        total_f2f_runs = sum(self.course_run_data["Runs"])

        # Create adjusted monthly demand dictionary
        adjusted_f2f_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Adjust rounding errors
        total_allocated = sum(adjusted_f2f_demand.values())
        if total_allocated != total_f2f_runs:
            max_month = max(adjusted_f2f_demand.items(), key=lambda x: x[1])[0]
            adjusted_f2f_demand[max_month] += (total_f2f_runs - total_allocated)

        # For each month, track and enforce
        for month in range(1, 13):
            target_demand = adjusted_f2f_demand.get(month, 0)
            if target_demand == 0:
                continue

            courses_in_month = []
            # Get all weeks that belong to this month
            month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]

            if not month_weeks:
                continue

            # For each course run, create a Boolean indicating if it's in this month
            for (course, delivery_type, language, i), week_var in schedule.items():
                is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")

                # Add constraints: is_in_month is True iff week_var is in month's weeks
                week_choices = []
                for week in month_weeks:
                    is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                    week_choices.append(is_in_this_week)

                model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())

                courses_in_month.append(is_in_month)

            # Enforce the target demand for this month
            if courses_in_month and target_demand > 0:
                model.Add(sum(courses_in_month) == target_demand)

        # Check feasibility with monthly distribution
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Monthly distribution constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding monthly distribution!")
            return "INFEASIBLE", diagnostics, None

        # STEP 6: Final full-featured model with all constraints and objective
        print("Step 6: Building full model with objective function")

        # Add penalties for objective function
        affinity_penalties = []
        trainer_utilization_penalties = []
        champion_assignment_penalties = []

        # Add affinity constraints (soft)
        for _, row in self.affinity_matrix_data.iterrows():
            c1, c2, gap_weeks = row["Course 1"], row["Course 2"], row["Gap Weeks"]

            c1_runs = []
            c2_runs = []

            for (course, _, _, i), var in schedule.items():
                if course == c1:
                    c1_runs.append((i, var))
                elif course == c2:
                    c2_runs.append((i, var))

            if not c1_runs or not c2_runs:
                continue

            # Sort by run index
            c1_runs.sort(key=lambda x: x[0])
            c2_runs.sort(key=lambda x: x[0])

            # Only check first run of each course
            run1, var1 = c1_runs[0]
            run2, var2 = c2_runs[0]

            # Create variables to check if either course is in Q4
            is_c1_q4 = model.NewBoolVar(f"{c1}_{run1}_in_q4")
            is_c2_q4 = model.NewBoolVar(f"{c2}_{run2}_in_q4")

            # Get Q4 weeks
            q4_weeks = [w for w, m in self.week_to_month_map.items() if m in [10, 11, 12]]

            # Set up Q4 detection for course 1
            c1_q4_choices = []
            for week in q4_weeks:
                is_in_this_week = model.NewBoolVar(f"{c1}_{run1}_in_week_{week}")
                model.Add(var1 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var1 != week).OnlyEnforceIf(is_in_this_week.Not())
                c1_q4_choices.append(is_in_this_week)
            model.AddBoolOr(c1_q4_choices).OnlyEnforceIf(is_c1_q4)
            model.AddBoolAnd([choice.Not() for choice in c1_q4_choices]).OnlyEnforceIf(is_c1_q4.Not())

            # Set up Q4 detection for course 2
            c2_q4_choices = []
            for week in q4_weeks:
                is_in_this_week = model.NewBoolVar(f"{c2}_{run2}_in_week_{week}")
                model.Add(var2 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var2 != week).OnlyEnforceIf(is_in_this_week.Not())
                c2_q4_choices.append(is_in_this_week)
            model.AddBoolOr(c2_q4_choices).OnlyEnforceIf(is_c2_q4)
            model.AddBoolAnd([choice.Not() for choice in c2_q4_choices]).OnlyEnforceIf(is_c2_q4.Not())

            # Either course in Q4 means reduced gap
            either_in_q4 = model.NewBoolVar(f"{c1}_{c2}_either_in_q4")
            model.AddBoolOr([is_c1_q4, is_c2_q4]).OnlyEnforceIf(either_in_q4)
            model.AddBoolAnd([is_c1_q4.Not(), is_c2_q4.Not()]).OnlyEnforceIf(either_in_q4.Not())

            # Soft affinity constraint with reduced gap for Q4
            too_close = model.NewBoolVar(f"affinity_too_close_{c1}_{c2}_{run1}_{run2}")

            # Regular gap weeks for non-Q4
            far_enough_after_normal = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_normal")
            far_enough_before_normal = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_normal")

            # Reduced gap weeks (50% reduction) for Q4
            reduced_gap = max(1, gap_weeks // 2)  # Ensure minimum 1 week gap
            far_enough_after_q4 = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_q4")
            far_enough_before_q4 = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_q4")

            # Normal gap constraints
            model.Add(var2 >= var1 + gap_weeks).OnlyEnforceIf([far_enough_after_normal, either_in_q4.Not()])
            model.Add(var2 <= var1 - gap_weeks).OnlyEnforceIf([far_enough_before_normal, either_in_q4.Not()])

            # Reduced gap constraints for Q4
            model.Add(var2 >= var1 + reduced_gap).OnlyEnforceIf([far_enough_after_q4, either_in_q4])
            model.Add(var2 <= var1 - reduced_gap).OnlyEnforceIf([far_enough_before_q4, either_in_q4])

            # Combine normal and Q4 constraints
            model.AddBoolOr([far_enough_after_normal, far_enough_before_normal]).OnlyEnforceIf([too_close.Not(), either_in_q4.Not()])
            model.AddBoolOr([far_enough_after_q4, far_enough_before_q4]).OnlyEnforceIf([too_close.Not(), either_in_q4])

            # Add violation constraints
            model.Add(var2 < var1 + reduced_gap).OnlyEnforceIf([too_close, either_in_q4])
            model.Add(var2 > var1 - reduced_gap).OnlyEnforceIf([too_close, either_in_q4])
            model.Add(var2 < var1 + gap_weeks).OnlyEnforceIf([too_close, either_in_q4.Not()])
            model.Add(var2 > var1 - gap_weeks).OnlyEnforceIf([too_close, either_in_q4.Not()])

            affinity_penalties.append(too_close)
            print(f"  Added affinity constraint: {c1} and {c2} should be {gap_weeks} weeks apart (reduced to {reduced_gap} in Q4)")

        # Add champion assignments (soft)
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            qualified_trainers = self.fleximatrix.get((course, language), [])

            for t_idx, trainer in enumerate(qualified_trainers):
                # Check if this is a champion course
                champion = self.course_champions.get((course, language))
                if champion == trainer:
                    not_using_champion = model.NewBoolVar(f"not_using_champion_{course}_{i}")
                    model.Add(trainer_var != t_idx).OnlyEnforceIf(not_using_champion)
                    model.Add(trainer_var == t_idx).OnlyEnforceIf(not_using_champion.Not())
                    champion_assignment_penalties.append(not_using_champion)

        # Add trainer utilization (soft)
        trainer_workload = {name: [] for name in self.consultant_data["Name"]}

        # Track course assignments for each trainer
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            qualified_trainers = self.fleximatrix.get((course, language), [])

            for t_idx, trainer in enumerate(qualified_trainers):
                is_assigned = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}")
                model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned)
                model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned.Not())

                # Accumulate workload
                trainer_workload[trainer].append((is_assigned, duration))

        # Calculate and constrain workload
        for trainer, workload_items in trainer_workload.items():
            if not workload_items:
                continue

            max_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Max_Days"].iloc[0]

            # Calculate total workload
            total_workload = model.NewIntVar(0, max_days, f"total_workload_{trainer}")
            weighted_sum = []
            for is_assigned, duration in workload_items:
                weighted_sum.append((is_assigned, duration))

            if weighted_sum:
                weighted_terms = []
                for is_assigned, duration in weighted_sum:
                    term = model.NewIntVar(0, duration, f"term_{id(is_assigned)}_{duration}")
                    model.Add(term == duration).OnlyEnforceIf(is_assigned)
                    model.Add(term == 0).OnlyEnforceIf(is_assigned.Not())
                    weighted_terms.append(term)
                model.Add(total_workload == sum(weighted_terms))
                # Enforce maximum workload
                model.Add(total_workload <= max_days)
                
                # Add hard constraint for minimum utilization if specified
                min_days = max(1, int(max_days * utilization_target / 100))  # Convert percentage to days
                if min_days > 0:
                    # Commenting out the hard constraint as requested
                    # model.Add(total_workload >= min_days)
                    log_progress(f"Skipping minimum utilization constraint for Trainer {trainer} (would be {min_days} days)")

        # Objective function
        model.Minimize(
            2 * sum(affinity_penalties) +
            4 * sum(champion_assignment_penalties) +
            3 * sum(trainer_utilization_penalties)
        )

        # Solve with final model
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = solver_time_minutes * 60
        status = solver.Solve(model)

        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Full model with objective",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        # Create and return results if feasible
        if feasible:
            schedule_results = []
            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    assigned_week = solver.Value(week_var)
                    start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")

                    # Get trainer assignment if it exists
                    if (course, delivery_type, language, i) in trainer_assignments:
                        trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                        trainer_idx = solver.Value(trainer_var)

                        # Check if this course has qualified trainers in the fleximatrix
                        if (course, language) in self.fleximatrix and len(self.fleximatrix[(course, language)]) > 0:
                            if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                                trainer = self.fleximatrix[(course, language)][trainer_idx]
                                is_champion = "✓" if self.course_champions.get((course, language)) == trainer else " "
                            else:
                                # Handle out of range index
                                trainer = "Unknown (Index Error)"
                                is_champion = " "
                        else:
                            # Handle missing course in fleximatrix
                            trainer = "No Qualified Trainers"
                            is_champion = " "

                        schedule_results.append({
                            "Week": assigned_week,
                            "Start Date": start_date,
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1,
                            "Trainer": trainer,
                            "Champion": is_champion
                        })

                except Exception as e:
                    print(f"Error processing result for {course} (run {i+1}): {e}")
                    # Add to unscheduled due to error
                    unscheduled_courses.append({
                        "Course": course,
                        "Delivery Type": delivery_type,
                        "Language": language,
                        "Run": i + 1,
                        "Error": str(e)
                    })

            # Convert to DataFrame
            schedule_df = pd.DataFrame(schedule_results)

            # Sort by week first, then by course name
            schedule_df = schedule_df.sort_values(by=["Week", "Course"])

            return "FEASIBLE", diagnostics, schedule_df
        else:
            return "INFEASIBLE", diagnostics, None

    # Here's debugging code to add to the "plot_weekly_course_bar_chart" method in your CourseScheduler class
    # Find this method in your app.py and replace it with this version:

    def plot_weekly_course_bar_chart(self, schedule, solver):
        """Creates a bar chart showing number of courses per week."""
        try:
            max_weeks = len(self.weekly_calendar)

            # Print debugging information
            print(f"DEBUG: max_weeks = {max_weeks}")
            print(f"DEBUG: schedule length = {len(schedule)}")

            # Count courses per week
            weekly_counts = {week: 0 for week in range(1, max_weeks + 1)}

            for (course, _, _, _), week_var in schedule.items():
                try:
                    assigned_week = solver.Value(week_var)
                    if 1 <= assigned_week <= max_weeks:  # Add range check
                        weekly_counts[assigned_week] += 1
                    else:
                        print(f"DEBUG: Week value {assigned_week} is out of range (1-{max_weeks})")
                except Exception as e:
                    print(f"DEBUG: Error getting week value for {course}: {e}")
                    continue

            # Convert to dataframe for plotting
            df = pd.DataFrame({
                'Week': list(weekly_counts.keys()),
                'Courses': list(weekly_counts.values())
            })

            # Create a bar chart
            plt.figure(figsize=(12, 6))
            bars = plt.bar(df['Week'], df['Courses'], width=0.8)

            # Add week dates as second x-axis
            ax1 = plt.gca()
            ax2 = ax1.twiny()

            # Select a subset of week numbers for readability
            step = max(1, max_weeks // 12)  # Show about 12 date labels
            selected_weeks = list(range(1, max_weeks + 1, step))

            # Get date strings for selected weeks
            date_labels = []
            for w in selected_weeks:
                if 0 <= w - 1 < len(self.weekly_calendar):  # Add range check
                    date_labels.append(self.weekly_calendar[w - 1].strftime("%b %d"))
                else:
                    date_labels.append(f"Week {w}")

            # Set positions and labels
            ax2.set_xticks([w for w in selected_weeks])
            ax2.set_xticklabels(date_labels, rotation=45)

            # Formatting
            plt.title("Number of Courses per Week", fontsize=16)
            ax1.set_xlabel("Week Number", fontsize=14)
            ax1.set_ylabel("Number of Courses", fontsize=14)
            ax2.set_xlabel("Week Start Date", fontsize=14)

            # Set y limit to make small differences more visible
            max_courses = max(weekly_counts.values()) if weekly_counts else 0
            ax1.set_ylim(0, max(4, max_courses + 1))

            # Highlight weeks with no courses
            for week, count in weekly_counts.items():
                if count == 0:
                    plt.axvspan(week - 0.4, week + 0.4, color='lightgray', alpha=0.3)

            # Add grid for readability
            plt.grid(axis='y', linestyle='--', alpha=0.7)

            plt.tight_layout()

            # Return the figure
            return plt.gcf()
        except Exception as e:
            print(f"DEBUG: Error in plot_weekly_course_bar_chart: {e}")
            # Return a simple error figure
            plt.figure(figsize=(8, 6))
            plt.text(0.5, 0.5, f"Error generating chart:\n{str(e)}",
                     horizontalalignment='center', verticalalignment='center')
            return plt.gcf()

    # And here's the fixed trainer workload chart function:

    def plot_trainer_workload_chart(self, schedule, trainer_assignments, solver):
        """Creates a bar chart showing trainer workload and utilization"""
        try:
            # Calculate days assigned to each trainer
            trainer_days = {name: 0 for name in self.consultant_data["Name"]}

            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    # Get the course duration
                    duration_series = self.course_run_data.loc[
                        self.course_run_data["Course Name"] == course, "Duration"]
                    if len(duration_series) == 0:
                        print(f"DEBUG: Course {course} not found in course_run_data")
                        continue

                    duration = duration_series.iloc[0]

                    # Check if this course+run has a trainer assignment
                    if (course, delivery_type, language, i) not in trainer_assignments:
                        print(f"DEBUG: No trainer assignment for {course} run {i + 1}")
                        continue

                    # Get trainer assignment
                    trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                    trainer_idx = solver.Value(trainer_var)

                    # Check if the course+language has qualified trainers
                    if (course, language) not in self.fleximatrix:
                        print(f"DEBUG: No fleximatrix entry for {course} ({language})")
                        continue

                    qualified_trainers = self.fleximatrix[(course, language)]

                    # Check if trainer index is valid
                    if 0 <= trainer_idx < len(qualified_trainers):
                        trainer = qualified_trainers[trainer_idx]
                        trainer_days[trainer] += duration
                    else:
                        print(
                            f"DEBUG: Invalid trainer index {trainer_idx} for {course} (max {len(qualified_trainers) - 1})")
                except Exception as e:
                    print(f"DEBUG: Error processing course {course} run {i + 1}: {e}")
                    continue

            # Prepare data for plotting - only include trainers with assignments
            trainers = []
            assigned_days = []
            max_days_list = []
            colors = []

            for name, days in sorted(trainer_days.items(), key=lambda x: x[1], reverse=True):
                if days > 0:  # Only include trainers with assignments
                    trainers.append(name)
                    assigned_days.append(days)

                    # Get max days from consultant data
                    max_days_series = self.consultant_data.loc[self.consultant_data["Name"] == name, "Max_Days"]

                    if len(max_days_series) > 0:
                        max_days = max_days_series.iloc[0]
                        max_days_list.append(max_days)

                        # Get title for color
                        title_series = self.consultant_data.loc[self.consultant_data["Name"] == name, "Title"]
                        title = title_series.iloc[0] if len(title_series) > 0 else "Unknown"
                        colors.append("tab:orange" if title == "Freelancer" else "tab:blue")
                    else:
                        print(f"DEBUG: Trainer {name} not found in consultant_data")
                        max_days_list.append(100)  # Default value
                        colors.append("tab:gray")  # Default color

            # Check if we have data to plot
            if not trainers:
                plt.figure(figsize=(10, 6))
                plt.text(0.5, 0.5, "No trainer assignments to display",
                         horizontalalignment='center', verticalalignment='center', fontsize=14)
                return plt.gcf()

            # Create figure and axis
            fig, ax = plt.subplots(figsize=(12, 6))

            # Plot assigned days
            bars = ax.bar(trainers, assigned_days, label="Assigned Days", color=colors)

            # Plot max days as a line
            ax.plot(trainers, max_days_list, 'rx-', label="Maximum Days", linewidth=2)

            # Add percentage labels
            for i, (bar, max_days) in enumerate(zip(bars, max_days_list)):
                height = bar.get_height()
                percentage = (height / max_days * 100) if max_days > 0 else 0
                ax.text(bar.get_x() + bar.get_width() / 2., height + 5,
                        f"{percentage:.1f}%", ha='center', va='bottom', fontsize=9)

            # Improve readability
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()

            # Add labels and title
            plt.ylabel("Training Days")
            plt.title("Trainer Workload and Utilization")
            plt.legend()
            plt.grid(axis='y', linestyle='--', alpha=0.7)

            # Return the figure
            return plt.gcf()
        except Exception as e:
            print(f"DEBUG: Error in plot_trainer_workload_chart: {e}")
            # Return a simple error figure
            plt.figure(figsize=(8, 6))
            plt.text(0.5, 0.5, f"Error generating chart:\n{str(e)}",
                     horizontalalignment='center', verticalalignment='center')
            return plt.gcf()

    def generate_monthly_validation(self, schedule, solver):
        """Generates a monthly validation report for the schedule"""
        # Get adjusted F2F demand from the monthly demand dataframe
        adjusted_f2f_demand = {}

        # Create the dictionary using the same logic as in run_optimization
        total_f2f_runs = sum(self.course_run_data["Runs"])
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Count courses by month
        monthly_counts = {month: 0 for month in range(1, 13)}

        for (course, delivery_type, language, i), week_var in schedule.items():
            assigned_week = solver.Value(week_var)
            month = self.week_to_month_map[assigned_week]
            monthly_counts[month] += 1

        # Create validation dataframe
        validation_data = []

        total_target = 0
        total_actual = 0
        all_match = True

        for month in range(1, 13):
            target = adjusted_f2f_demand.get(month, 0)
            actual = monthly_counts[month]
            diff = actual - target

            status = "MATCH" if diff == 0 else "OVER" if diff > 0 else "UNDER"
            if diff != 0:
                all_match = False

            validation_data.append({
                "Month": str(month),  # Convert month to string
                "Target": target,
                "Actual": actual,
                "Difference": diff,
                "Status": status
            })

            total_target += target
            total_actual += actual

        # Add total row
        validation_data.append({
            "Month": "TOTAL",
            "Target": total_target,
            "Actual": total_actual,
            "Difference": total_actual - total_target,
            "Status": "MATCH" if all_match else "MISMATCH"
        })

        # Convert to DataFrame and ensure proper column types
        df = pd.DataFrame(validation_data)
        df['Target'] = df['Target'].astype(int)
        df['Actual'] = df['Actual'].astype(int)
        df['Difference'] = df['Difference'].astype(int)
        
        return df

    def generate_trainer_utilization_report(self, schedule, trainer_assignments, solver):
        """Generates a report on trainer utilization"""
        # Calculate days assigned to each trainer
        trainer_days = {name: 0 for name in self.consultant_data["Name"]}
        trainer_courses = {name: 0 for name in self.consultant_data["Name"]}
        champion_courses = {name: 0 for name in self.consultant_data["Name"]}

        for (course, delivery_type, language, i), week_var in schedule.items():
            try:
                duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]

                # Get trainer assignment
                if (course, delivery_type, language, i) not in trainer_assignments:
                    continue
                
                trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                trainer_idx = solver.Value(trainer_var)

                # Check if this course has an entry in the fleximatrix
                if (course, language) not in self.fleximatrix:
                    continue
                
                if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                    trainer = self.fleximatrix[(course, language)][trainer_idx]
                    trainer_days[trainer] += duration
                    trainer_courses[trainer] += 1

                    # Check if this is a champion course
                    if self.course_champions.get((course, language)) == trainer:
                        champion_courses[trainer] += 1
            except Exception as e:
                print(f"Error in utilization report for {course} (run {i+1}): {e}")
                continue

        # Create trainer utilization report dataframe
        utilization_data = []

        for _, row in self.consultant_data.iterrows():
            name = row["Name"]
            title = row["Title"]
            max_days = row["Max_Days"]
            assigned_days = trainer_days[name]
            utilization = (assigned_days / max_days * 100) if max_days > 0 else 0
            courses = trainer_courses[name]
            champion = champion_courses[name]

            utilization_data.append({
                "Name": name,
                "Title": title,
                "Max Days": max_days,
                "Assigned Days": assigned_days,
                "Utilization %": round(utilization, 1),
                "Courses": courses,
                "Champion Courses": champion
            })

        # Add total row
        total_days = sum(trainer_days.values())
        max_days_sum = self.consultant_data["Max_Days"].sum()
        overall_utilization = (total_days / max_days_sum * 100) if max_days_sum > 0 else 0

        utilization_data.append({
            "Name": "TOTAL",
            "Title": "",
            "Max Days": max_days_sum,
            "Assigned Days": total_days,
            "Utilization %": round(overall_utilization, 1),
            "Courses": sum(trainer_courses.values()),
            "Champion Courses": sum(champion_courses.values())
        })

        return pd.DataFrame(utilization_data)

    def generate_excel_report(self, schedule_df, monthly_validation_df, trainer_utilization_df):
        """Generate an Excel report with multiple sheets containing all results"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if schedule_df is not None:
                schedule_df.to_excel(writer, sheet_name='Schedule', index=False)
            else:
                # Create an empty DataFrame if None
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Schedule', index=False)

            if monthly_validation_df is not None:
                monthly_validation_df.to_excel(writer, sheet_name='Monthly Validation', index=False)
            else:
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Monthly Validation',
                                                                     index=False)

            if trainer_utilization_df is not None:
                trainer_utilization_df.to_excel(writer, sheet_name='Trainer Utilization', index=False)
            else:
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Trainer Utilization',
                                                                     index=False)
            
            # Create the new Calendar View sheet
            if schedule_df is not None and not schedule_df.empty:
                # Get all trainers from the consultant data
                trainers = self.consultant_data["Name"].tolist()
                
                # Create a new dataframe with trainers as rows
                calendar_data = []
                for trainer in trainers:
                    calendar_data.append({"Trainer": trainer})
                
                calendar_df = pd.DataFrame(calendar_data)
                
                # Add the dataframe to the Excel but we'll manually format it later
                calendar_df.to_excel(writer, sheet_name='Calendar View', index=False)
                
                # Get the worksheet and workbook objects
                worksheet = writer.sheets['Calendar View']
                workbook = writer.book
                
                # Define fill colors
                light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Assigned
                dark_green_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")   # Champion
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")         # Vacation
                purple_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")      # Holiday
                
                # Create header row with week start dates
                for col_idx, week_start_date in enumerate(self.weekly_calendar, start=2):
                    # Format date as DD-MMM
                    formatted_date = week_start_date.strftime("%d-%b")
                    week_num = col_idx - 1  # Adjusted for Excel column indexing
                    
                    # Write the date to the header
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = formatted_date
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    
                    # Format column width
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 10
                
                # Create dictionary to track trainer assignments by week
                trainer_assignments_by_week = {}
                champion_assignments_by_week = {}
                
                # Process schedule data to get assignments
                for _, row in schedule_df.iterrows():
                    week = row['Week']
                    trainer = row['Trainer']
                    is_champion = row['Champion'].strip() == "✓"
                    
                    if (trainer, week) not in trainer_assignments_by_week:
                        trainer_assignments_by_week[(trainer, week)] = 0
                    
                    trainer_assignments_by_week[(trainer, week)] += 1
                    
                    if is_champion:
                        if (trainer, week) not in champion_assignments_by_week:
                            champion_assignments_by_week[(trainer, week)] = 0
                        
                        champion_assignments_by_week[(trainer, week)] += 1
                
                # Process all trainers and weeks
                for row_idx, trainer in enumerate(trainers, start=2):
                    for col_idx, week_start_date in enumerate(self.weekly_calendar, start=2):
                        week_num = col_idx - 1  # Adjusted for Excel column indexing
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Check for public holidays
                        has_public_holiday = False
                        for _, holiday in self.public_holidays_data.iterrows():
                            week_end_date = week_start_date + datetime.timedelta(days=4)  # 5-day work week
                            if (holiday["Start Date"] <= week_end_date and holiday["End Date"] >= week_start_date):
                                has_public_holiday = True
                                break
                        
                        # Check for trainer vacation
                        is_on_vacation = False
                        for _, leave in self.annual_leaves[self.annual_leaves["Name"] == trainer].iterrows():
                            week_end_date = week_start_date + datetime.timedelta(days=4)  # 5-day work week
                            if (leave["Start_Date"] <= week_end_date and leave["End_Date"] >= week_start_date):
                                is_on_vacation = True
                                break
                        
                        # Check for trainer assignment
                        is_assigned = (trainer, week_num) in trainer_assignments_by_week
                        is_champion_assignment = (trainer, week_num) in champion_assignments_by_week
                        
                        # Apply appropriate fill and set value
                        if is_on_vacation:
                            cell.fill = red_fill
                            cell.value = "Vacation"
                        elif has_public_holiday:
                            cell.fill = purple_fill
                            cell.value = "Holiday"
                        elif is_champion_assignment:
                            cell.fill = dark_green_fill
                            cell.value = "Champion"
                            # White text for dark background
                            cell.font = Font(color="FFFFFF")
                        elif is_assigned:
                            cell.fill = light_green_fill
                            cell.value = "Assigned"
                        else:
                            cell.value = ""
                
                # Add a legend
                legend_row = len(trainers) + 4
                
                worksheet.cell(row=legend_row, column=1).value = "Legend:"
                worksheet.cell(row=legend_row, column=1).font = Font(bold=True)
                
                # Trainer assigned
                worksheet.cell(row=legend_row + 1, column=1).value = "Light Green:"
                worksheet.cell(row=legend_row + 1, column=2).value = "Trainer Assigned"
                worksheet.cell(row=legend_row + 1, column=2).fill = light_green_fill
                
                # Champion assigned
                worksheet.cell(row=legend_row + 2, column=1).value = "Dark Green:"
                worksheet.cell(row=legend_row + 2, column=2).value = "Champion Assigned"
                worksheet.cell(row=legend_row + 2, column=2).fill = dark_green_fill
                worksheet.cell(row=legend_row + 2, column=2).font = Font(color="FFFFFF")
                
                # Vacation
                worksheet.cell(row=legend_row + 3, column=1).value = "Red:"
                worksheet.cell(row=legend_row + 3, column=2).value = "Trainer on Vacation"
                worksheet.cell(row=legend_row + 3, column=2).fill = red_fill
                
                # Public holiday
                worksheet.cell(row=legend_row + 4, column=1).value = "Purple:"
                worksheet.cell(row=legend_row + 4, column=2).value = "Public Holiday"
                worksheet.cell(row=legend_row + 4, column=2).fill = purple_fill
                
                # Adjust column width for trainer names
                worksheet.column_dimensions['A'].width = 25
            else:
                # Create an empty calendar view if no schedule data
                pd.DataFrame(columns=["No schedule data available"]).to_excel(writer, sheet_name='Calendar View', index=False)

        output.seek(0)
        return output


    def analyze_constraints_visually(self):
        """Generate visual analysis of constraints to help identify sources of infeasibility"""
        import matplotlib.pyplot as plt
        import seaborn as sns
        import numpy as np
        from matplotlib.colors import LinearSegmentedColormap

        # 1. Analyze monthly distribution vs. available weeks
        months = range(1, 13)
        total_f2f_runs = sum(self.course_run_data["Runs"])

        # Calculate monthly demand
        monthly_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()
        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            monthly_demand[month] = demand

        # Calculate available weeks per month
        monthly_weeks = {}
        for week, month in self.week_to_month_map.items():
            if month not in monthly_weeks:
                monthly_weeks[month] = 0
            monthly_weeks[month] += 1

        # Calculate available working days per month
        monthly_working_days = {}
        for week, month in self.week_to_month_map.items():
            if month not in monthly_working_days:
                monthly_working_days[month] = 0
            monthly_working_days[month] += self.weekly_working_days.get(week, 0)

        # Create dataframe for plotting
        monthly_data = []
        for month in range(1, 13):
            monthly_data.append({
                'Month': month,
                'Demand': monthly_demand.get(month, 0),
                'Weeks': monthly_weeks.get(month, 0),
                'Working Days': monthly_working_days.get(month, 0)
            })

        monthly_df = pd.DataFrame(monthly_data)

        # Calculate two metrics:
        # 1. Courses per week: How many courses need to be scheduled per available week
        # 2. Demand pressure: Percentage of working days required for courses
        monthly_df['Courses per Week'] = monthly_df['Demand'] / monthly_df['Weeks'].replace(0, np.nan)

        # Assuming each course takes 5 days on average
        avg_course_days = self.course_run_data["Duration"].mean()
        monthly_df['Demand Pressure (%)'] = (monthly_df['Demand'] * avg_course_days * 100) / monthly_df[
            'Working Days'].replace(0, np.nan)

        # Create figures
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))

        # Plot 1: Monthly Demand vs. Available Weeks
        ax1 = axes[0, 0]
        monthly_df.plot(x='Month', y=['Demand', 'Weeks'], kind='bar', ax=ax1)
        ax1.set_title('Monthly Course Demand vs. Available Weeks')
        ax1.set_xlabel('Month')
        ax1.set_ylabel('Count')
        ax1.axhline(y=monthly_df['Demand'].mean(), color='r', linestyle='--', label='Avg. Demand')
        ax1.legend()

        # Plot 2: Courses per Week
        ax2 = axes[0, 1]
        monthly_df.plot(x='Month', y='Courses per Week', kind='bar', ax=ax2, color='orange')
        ax2.set_title('Courses per Available Week')
        ax2.set_xlabel('Month')
        ax2.set_ylabel('Courses per Week')
        ax2.axhline(y=1, color='g', linestyle='--', label='1 Course/Week')
        ax2.axhline(y=2, color='y', linestyle='--', label='2 Courses/Week')
        ax2.axhline(y=3, color='r', linestyle='--', label='3 Courses/Week')
        ax2.legend()

        # Plot 3: Demand Pressure
        ax3 = axes[1, 0]
        bars = monthly_df.plot(x='Month', y='Demand Pressure (%)', kind='bar', ax=ax3, color='purple')
        ax3.set_title('Monthly Demand Pressure (% of Working Days Required)')
        ax3.set_xlabel('Month')
        ax3.set_ylabel('Percentage of Working Days')
        ax3.axhline(y=100, color='r', linestyle='--', label='100% (Full Capacity)')
        ax3.axhline(y=80, color='y', linestyle='--', label='80% (Threshold)')

        # Color code bars by severity
        for i, p in enumerate(bars.patches):
            pressure = monthly_df['Demand Pressure (%)'].iloc[i]
            if np.isnan(pressure):
                p.set_color('gray')
            elif pressure > 100:
                p.set_color('red')
            elif pressure > 80:
                p.set_color('orange')
            elif pressure > 60:
                p.set_color('yellow')
            else:
                p.set_color('green')

        ax3.legend()

        # Plot 4: Week Restrictions Analysis
        ax4 = axes[1, 1]

        # Count how many courses are restricted in each week position
        restrictions = {'First': 0, 'Second': 0, 'Third': 0, 'Fourth': 0, 'Last': 0}

        for course, course_restrictions in self.week_restrictions.items():
            for position, is_restricted in course_restrictions.items():
                if is_restricted:
                    restrictions[position] += 1

        # Calculate percentage of courses restricted
        total_courses = len(set(self.course_run_data["Course Name"]))
        restriction_percentages = {pos: (count * 100 / total_courses) if total_courses > 0 else 0
                                   for pos, count in restrictions.items()}

        # Create restriction dataframe
        restriction_df = pd.DataFrame({
            'Week Position': list(restriction_percentages.keys()),
            'Percentage Restricted': list(restriction_percentages.values())
        })

        # Plot restriction percentages
        restriction_df.plot(x='Week Position', y='Percentage Restricted', kind='bar', ax=ax4, color='teal')
        ax4.set_title('Percentage of Courses with Week Restrictions')
        ax4.set_xlabel('Week Position in Month')
        ax4.set_ylabel('Percentage of Courses')
        ax4.axhline(y=50, color='y', linestyle='--', label='50% Threshold')
        ax4.legend()

        # Adjust layout
        plt.tight_layout()

        # Create another figure for the heatmap of week restrictions
        plt.figure(figsize=(12, 10))

        # Create a matrix of week restrictions for heatmap
        courses = list(set(self.course_run_data["Course Name"]))
        weeks = list(range(1, len(self.weekly_calendar) + 1))

        # Initialize matrix with 1s (allowed)
        heatmap_data = np.ones((len(courses), len(weeks)))

        # Fill in restricted weeks with 0s
        # Modify this part too:
        # Fill in restricted weeks with 0s
        for i, course in enumerate(courses):
            for j, week in enumerate(weeks):
                # Make sure we don't exceed the valid week range
                week_num = week + 1
                if week_num <= len(self.weekly_calendar):  # Add this check
                    # Check if this week is restricted due to working days
                    duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[
                        0]
                    if self.weekly_working_days.get(week_num, 0) < duration:
                        heatmap_data[i, j] = 0

                    # Check week position restrictions
                    if course in self.week_restrictions:
                        week_info = self.week_position_in_month.get(week_num, {})

                        if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                            heatmap_data[i, j] = 0

        # Create heatmap
        course_week_heatmap = plt.figure(figsize=(18, 8))
        sns.heatmap(heatmap_data, cmap=['red', 'green'], cbar=False,
                    xticklabels=[f"W{w + 1}" for w in range(len(weeks))],
                    yticklabels=courses)
        plt.title('Course-Week Availability Matrix (Green = Available, Red = Restricted)')
        plt.xlabel('Week Number')
        plt.ylabel('Course')
        plt.tight_layout()

        # Analysis of trainer qualifications
        plt.figure(figsize=(10, 8))

        # Count how many qualified trainers per course/language
        course_lang_trainers = {}
        for (course, language), trainers in self.fleximatrix.items():
            course_lang_trainers[(course, language)] = len(trainers)

        # Convert to DataFrame
        trainers_df = pd.DataFrame([
            {'Course': course, 'Language': lang, 'Qualified Trainers': count}
            for (course, lang), count in course_lang_trainers.items()
        ])

        # Get runs per course/language
        runs_df = self.course_run_data.groupby(['Course Name', 'Language'])['Runs'].sum().reset_index()
        runs_df.columns = ['Course', 'Language', 'Runs']

        # Merge trainers and runs
        analysis_df = pd.merge(trainers_df, runs_df, on=['Course', 'Language'])

        # Add trainers per run ratio
        analysis_df['Trainers per Run'] = analysis_df['Qualified Trainers'] / analysis_df['Runs']

        # Sort by trainers per run (ascending)
        analysis_df = analysis_df.sort_values('Trainers per Run')

        # Plot trainers per run
        trainer_ratio_fig = plt.figure(figsize=(12, 6))
        trainer_ratio_plot = sns.barplot(x='Course', y='Trainers per Run', data=analysis_df)
        plt.title('Qualified Trainers per Course Run')
        plt.xticks(rotation=90)
        plt.axhline(y=1, color='r', linestyle='--', label='1:1 Ratio')
        plt.axhline(y=2, color='y', linestyle='--', label='2:1 Ratio')
        plt.tight_layout()

        # Trainer availability heatmap
        trainer_avail_fig = plt.figure(figsize=(18, 10))

        # Create a matrix of trainer availability
        trainers = list(self.consultant_data["Name"])

        # Initialize matrix with 1s (available)
        availability_data = np.ones((len(trainers), len(weeks)))

        # Fill in unavailable weeks with 0s
        for i, trainer in enumerate(trainers):
            for j, week in enumerate(weeks):
                # Make sure we don't exceed the length of weekly_calendar
                week_num = week + 1
                if week_num <= len(self.weekly_calendar):  # Add this check
                    if not self.is_trainer_available(trainer, week_num):
                        availability_data[i, j] = 0


        # Create heatmap
        sns.heatmap(availability_data, cmap=['red', 'green'], cbar=False,
                    xticklabels=[f"W{w + 1}" for w in range(len(weeks))],
                    yticklabels=trainers)
        plt.title('Trainer Availability Matrix (Green = Available, Red = Unavailable)')
        plt.xlabel('Week Number')
        plt.ylabel('Trainer')
        plt.tight_layout()

        # Calculate Critical Metrics
        total_runs = sum(self.course_run_data["Runs"])
        total_available_weeks = sum(w > 0 for w in self.weekly_working_days.values())
        total_available_days = sum(self.weekly_working_days.values())
        total_course_days = sum(self.course_run_data["Runs"] * self.course_run_data["Duration"])

        # Create a summary stats figure
        summary_fig = plt.figure(figsize=(10, 6))

        # Format as a table
        cell_text = [
            ["Total Course Runs", f"{total_runs}"],
            ["Available Weeks", f"{total_available_weeks}"],
            ["Available Working Days", f"{total_available_days}"],
            ["Required Course Days", f"{total_course_days}"],
            ["Overall Capacity Utilization",
             f"{(total_course_days / total_available_days * 100) if total_available_days > 0 else 0:.1f}%"],
            ["Runs per Available Week", f"{(total_runs / total_available_weeks) if total_available_weeks > 0 else 0:.2f}"]
        ]

        # Add color coding
        cell_colors = [['lightgray', 'lightgray'] for _ in range(len(cell_text))]

        # Highlight potential issues
        capacity_pct = (total_course_days / total_available_days * 100) if total_available_days > 0 else 0
        if capacity_pct > 90:
            cell_colors[4][1] = 'lightcoral'
        elif capacity_pct > 75:
            cell_colors[4][1] = 'lightyellow'
        else:
            cell_colors[4][1] = 'lightgreen'

        runs_per_week = (total_runs / total_available_weeks) if total_available_weeks > 0 else 0
        if runs_per_week > 1.5:
            cell_colors[5][1] = 'lightcoral'
        elif runs_per_week > 1:
            cell_colors[5][1] = 'lightyellow'
        else:
            cell_colors[5][1] = 'lightgreen'

        # Create table
        plt.axis('off')
        summary_table = plt.table(cellText=cell_text, cellColours=cell_colors,
                                  colLabels=["Metric", "Value"],
                                  loc='center', cellLoc='center')

        # Format table
        summary_table.auto_set_font_size(False)
        summary_table.set_fontsize(12)
        summary_table.scale(1, 2)

        plt.title('Overall Scheduling Feasibility Metrics', fontsize=16)
        plt.tight_layout()

        # Return all figures for Streamlit
        return {
            'monthly_analysis': fig,
            'course_week_heatmap': course_week_heatmap,
            'trainer_ratio_plot': trainer_ratio_fig,
            'trainer_avail_heatmap': trainer_avail_fig,
            'summary_metrics': summary_fig
        }

    # Updated function that doesn't use the unsupported 'pattern' property


    # Add this method to the CourseScheduler class
    def analyze_unscheduled_course(self, course, language, run):
        """Analyze why a specific course couldn't be scheduled"""
        analysis = {}

        # 1. Check trainer qualification
        qualified_trainers = self.fleximatrix.get((course, language), [])
        analysis["qualified_trainers"] = len(qualified_trainers)

        if qualified_trainers:
            # 2. Check course duration
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            analysis["duration"] = duration

            # 3. Check valid weeks (accounting for working days)
            valid_weeks = []
            for w, days in self.weekly_working_days.items():
                if days >= duration:
                    valid_weeks.append(w)
            analysis["weeks_with_enough_days"] = len(valid_weeks)

            # 4. Check week restrictions
            if course in self.week_restrictions:
                restricted_positions = []
                for position, is_restricted in self.week_restrictions[course].items():
                    if is_restricted:
                        restricted_positions.append(position)
                analysis["restricted_week_positions"] = restricted_positions

                # Calculate impact of restrictions
                restricted_weeks = 0
                for w in valid_weeks:
                    week_info = self.week_position_in_month.get(w, {})
                    if (week_info.get('is_first') and 'First' in restricted_positions) or \
                            (week_info.get('is_second') and 'Second' in restricted_positions) or \
                            (week_info.get('is_third') and 'Third' in restricted_positions) or \
                            (week_info.get('is_fourth') and 'Fourth' in restricted_positions) or \
                            (week_info.get('is_last') and 'Last' in restricted_positions):
                        restricted_weeks += 1

                analysis["weeks_restricted"] = restricted_weeks
                analysis["weeks_available_after_restrictions"] = len(valid_weeks) - restricted_weeks
            else:
                analysis["restricted_week_positions"] = []
                analysis["weeks_restricted"] = 0
                analysis["weeks_available_after_restrictions"] = len(valid_weeks)

            # 5. Check trainer availability in valid weeks
            available_weeks = {}
            for trainer in qualified_trainers:
                available_weeks[trainer] = []
                for w in valid_weeks:
                    # Skip weeks with restrictions
                    if course in self.week_restrictions:
                        week_info = self.week_position_in_month.get(w, {})
                        skip_week = False
                        for position in restricted_positions:
                            if position == 'First' and week_info.get('is_first'):
                                skip_week = True
                            elif position == 'Second' and week_info.get('is_second'):
                                skip_week = True
                            elif position == 'Third' and week_info.get('is_third'):
                                skip_week = True
                            elif position == 'Fourth' and week_info.get('is_fourth'):
                                skip_week = True
                            elif position == 'Last' and week_info.get('is_last'):
                                skip_week = True
                        if skip_week:
                            continue

                    if self.is_trainer_available(trainer, w):
                        available_weeks[trainer].append(w)

            # Count how many trainers are available for each week
            weeks_with_trainers = {}
            for w in valid_weeks:
                weeks_with_trainers[w] = sum(1 for trainer, weeks in available_weeks.items() if w in weeks)

            analysis["trainer_availability"] = available_weeks
            analysis["weeks_with_trainers"] = weeks_with_trainers
            analysis["weeks_with_any_trainer"] = sum(1 for count in weeks_with_trainers.values() if count > 0)

        return analysis

    def _sample_courses_for_optimization(self, course_df, max_courses=500):
        """Sample the most important courses for optimization when dealing with extremely large problems"""
        # Make a copy to avoid modifying the original dataframe
        df = course_df.copy()
        
        # Calculate the total runs before sampling
        total_runs_before = df["Runs"].sum()
        
        # Count total runs so far
        total_runs = 0
        
        # Dictionary to track which courses to keep
        courses_to_keep = {}
        
        # First pass: Sort courses by importance criteria
        # Priority 1: Required courses (you can define what makes a course "required")
        # Priority 2: Courses with champions assigned
        # Priority 3: Courses with the most qualified trainers
        
        # Create a scoring system for course importance
        scored_courses = []
        for _, row in df.iterrows():
            course = row["Course Name"]
            language = row["Language"]
            runs = row["Runs"]
            
            # Skip courses we've already decided on
            if (course, language) in courses_to_keep:
                continue
                
            # Calculate a score for this course (higher is more important)
            score = 0
            
            # Check if it has a champion (courses with champions get priority)
            if (course, language) in self.course_champions:
                score += 100
                
            # Check how many qualified trainers it has (more trainers = more flexibility)
            trainer_count = len(self.fleximatrix.get((course, language), []))
            score += min(trainer_count * 5, 50)  # Cap at 50 points
            
            # Add the course to our scoring list
            scored_courses.append({
                "course": course,
                "language": language,
                "runs": runs,
                "score": score,
                "row_idx": _
            })
        
        # Sort courses by score (highest first)
        scored_courses.sort(key=lambda x: x["score"], reverse=True)
        
        # Find the row indices to keep
        indices_to_keep = []
        
        # Take courses until we reach max_courses
        for course_info in scored_courses:
            if total_runs + course_info["runs"] <= max_courses:
                indices_to_keep.append(course_info["row_idx"])
                total_runs += course_info["runs"]
            else:
                # If including all runs would exceed our limit, 
                # reduce the number of runs for this course
                remaining = max_courses - total_runs
                if remaining > 0:
                    # Create a modified row with reduced runs
                    reduced_row = df.loc[course_info["row_idx"]].copy()
                    reduced_row["Runs"] = remaining
                    # Add this row to a list of modified rows to append later
                    df.at[course_info["row_idx"], "Runs"] = remaining
                    indices_to_keep.append(course_info["row_idx"])
                    total_runs += remaining
                break
        
        # Filter the dataframe
        sampled_df = df.loc[indices_to_keep].copy().reset_index(drop=True)
        
        # Add a note about how many courses were dropped
        print(f"Sampling: Reduced from {total_runs_before} to {sampled_df['Runs'].sum()} total course runs")
        
        return sampled_df


# Create Streamlit application
def main():
    try:
        # App title with better styling
        st.title("📚 Course Scheduler App")
        st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            color: #1E88E5;
            margin-bottom: 1rem;
        }
        .section-header {
            background-color: #f0f2f6;
            padding: 10px;
            border-radius: 5px;
            margin-top: 15px;
            margin-bottom: 15px;
        }
        .success-box {
            background-color: #d4edda;
            color: #155724;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        .warning-box {
            background-color: #fff3cd;
            color: #856404;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        .error-box {
            background-color: #f8d7da;
            color: #721c24;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        </style>
        """, unsafe_allow_html=True)

        st.markdown("Optimize your course schedules and trainer assignments efficiently")

        # Initialize session state for storing the scheduler object
        if 'scheduler' not in st.session_state:
            st.session_state.scheduler = CourseScheduler()

        if 'schedule_df' not in st.session_state:
            st.session_state.schedule_df = None

        if 'validation_df' not in st.session_state:
            st.session_state.validation_df = None

        if 'utilization_df' not in st.session_state:
            st.session_state.utilization_df = None

        if 'optimization_status' not in st.session_state:
            st.session_state.optimization_status = None

        if 'unscheduled_courses' not in st.session_state:
            st.session_state.unscheduled_courses = []

        # Create tabs for a more organized workflow
        tab1, tab2, tab3, tab4 = st.tabs([
            "📤 Data Input & Setup",
            "⚙️ Optimization Settings",
            "📊 Results",
            "🔍 Debug & Troubleshoot"
        ])

        with tab1:
            st.markdown('<div class="section-header"><h2>1. Upload & Setup Data</h2></div>', unsafe_allow_html=True)

            col1, col2 = st.columns([1, 1])

            with col1:
                # Template generator section
                with st.expander("Not sure about the format? Download our template first", expanded=False):
                    st.write("Use this template as a starting point for your data:")
                    if st.button("Generate Template", key="generate_template_button"):
                        template = create_excel_template()
                        st.download_button(
                            "Download Excel Template",
                            data=template,
                            file_name="course_scheduler_template.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_template_btn"
                        )

                # Data upload section
                uploaded_file = st.file_uploader("Upload Excel file with course and trainer data", type=["xlsx"])

                if uploaded_file is not None:
                    if st.button("Load Data", key="load_data_btn"):
                        with st.spinner("Loading data from Excel..."):
                            success = st.session_state.scheduler.load_data_from_excel(uploaded_file)
                            if success:
                                st.markdown('<div class="success-box">✅ Data loaded successfully!</div>',
                                            unsafe_allow_html=True)
                            else:
                                st.markdown(
                                    '<div class="error-box">❌ Failed to load data. Please check your Excel file format.</div>',
                                    unsafe_allow_html=True)

            with col2:
                # Data format guide
                st.markdown("### Required Excel Sheets")
                with st.expander("View format requirements", expanded=False):
                    st.markdown("""
                    Your Excel file should contain these sheets:
    
                    1. **CourseData**: Course information
                       - Columns: Course Name, Delivery Type, Language, Runs, Duration
    
                    2. **TrainerData**: Trainer information
                       - Columns: Name, Title, Max_Days
    
                    3. **PriorityData**: Trainer priority levels
                       - Columns: Title, Priority
    
                    4. **AnnualLeaves**: Trainer leave periods
                       - Columns: Name, Start_Date, End_Date
    
                    5. **AffinityMatrix**: Course timing relationships
                       - Columns: Course 1, Course 2, Gap Weeks
    
                    6. **PublicHolidays**: Holiday periods
                       - Columns: Start Date, End Date
    
                    7. **Fleximatrix**: Trainer qualifications
                       - "U" marks qualified trainers
    
                    8. **WeekRestrictions**: Course timing constraints
                       - Columns: Course, Week Type, Restricted, Notes
    
                    9. **MonthlyDemand**: Distribution targets
                       - Columns: Month, Percentage
                    """)

                # Calendar setup (only shown after data is loaded)
                if st.session_state.scheduler.course_run_data is not None:
                    st.markdown("### Calendar Setup")
                    year = st.number_input("Select scheduling year", min_value=2020, max_value=2030, value=2025)
                    weekend_options = {"FS": "Friday-Saturday", "SS": "Saturday-Sunday"}
                    weekend_selection = st.selectbox("Select weekend configuration",
                                                     options=list(weekend_options.keys()),
                                                     format_func=lambda x: weekend_options[x])

                    if st.button("Initialize Calendar", key="init_calendar_btn"):
                        with st.spinner("Generating calendar..."):
                            success = st.session_state.scheduler.initialize_calendar(year, weekend_selection)
                            if success:
                                st.markdown(
                                    f'<div class="success-box">✅ Calendar created with {len(st.session_state.scheduler.weekly_calendar)} weeks</div>',
                                    unsafe_allow_html=True)
                            else:
                                st.markdown('<div class="error-box">❌ Failed to create calendar</div>',
                                            unsafe_allow_html=True)

        # In the "Optimization Settings" tab
        with tab2:
            if st.session_state.scheduler.course_run_data is None:
                st.info("Please load your data in the 'Data Input & Setup' tab first")
            else:
                st.markdown('<div class="section-header"><h2>2. Optimization Parameters</h2></div>',
                            unsafe_allow_html=True)

                # Create a container with custom styling for the sliders
                st.markdown("""
                <style>
                .slider-container {
                    background-color: #f8f9fa;
                    border-radius: 10px;
                    padding: 20px;
                    border: 1px solid #e9ecef;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
                    margin-bottom: 20px;
                }
                .slider-header {
                    font-size: 1.2rem;
                    font-weight: 600;
                    margin-bottom: 15px;
                    color: #1E88E5;
                    border-bottom: 1px solid #e9ecef;
                    padding-bottom: 10px;
                }
                </style>
                """, unsafe_allow_html=True)

                # Create a two-column layout with sliders in the left column
                col1, col2 = st.columns([1, 1])

                with col1:
                    # Priority Weight Sliders in a nice frame
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Priority Weights</div>', unsafe_allow_html=True)
                    monthly_weight = st.slider("Monthly Distribution Priority", 1, 10, 5,
                                               help="Higher values enforce monthly targets more strictly")
                    affinity_weight = st.slider("Course Affinity Priority", 1, 10, 2,
                                                help="Higher values enforce gaps between related courses")
                    st.markdown('</div>', unsafe_allow_html=True)

                    # Constraint Management sliders in a nice frame
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Constraint Settings</div>', unsafe_allow_html=True)
                    min_course_spacing = st.slider(
                        "Minimum Weeks Between Course Runs",
                        min_value=1,
                        max_value=8,
                        value=2,
                        help="Lower values allow courses to be scheduled closer together"
                    )
                    utilization_target = st.slider(
                        "Target Utilization Percentage",
                        min_value=50,
                        max_value=100,
                        value=70,
                        help="Target percentage of max workdays for trainers"
                    )
                    # Add a slider to limit affinity constraints
                    max_affinity = st.slider(
                        "Maximum Affinity Constraints",
                        min_value=10,
                        max_value=700,
                        value=200,
                        help="Limit the number of course affinity constraints to reduce memory usage"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                with col2:
                    # Checkboxes and other options
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Optimization Strategy</div>', unsafe_allow_html=True)
                    enforce_monthly = st.checkbox(
                        "Enforce Monthly Distribution as Hard Constraint",
                        value=False,
                        help="Uncheck to allow flexibility in monthly distribution (recommended for feasibility)"
                    )
                    enforce_champions = st.checkbox(
                        "Prioritize Champion Trainers",
                        value=True,
                        help="Uncheck to allow any qualified trainer without champion priority"
                    )

                    prioritize_all_courses = st.checkbox(
                        "Prioritize scheduling all courses",
                        value=False,
                        help="When enabled, the optimizer will prioritize scheduling all courses, potentially at the expense of other goals"
                    )
                    
                    accelerated_mode = st.checkbox(
                        "Accelerated Mode",
                        value=False,
                        help="Enable faster optimization with simplified model (use if regular optimization is too slow)"
                    )

                    solver_time = st.slider(
                        "Solver Time Limit (minutes)",
                        min_value=1,
                        max_value=60,
                        value=5,
                        help="Maximum time the optimizer will run before returning best solution found"
                    )
                    num_workers = st.slider(
                        "Number of Search Workers",
                        min_value=1,
                        max_value=16,
                        value=8,
                        help="More workers can find solutions faster but use more CPU"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)




                    # Troubleshooting options
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Troubleshooting</div>', unsafe_allow_html=True)
                    if st.button("Disable All Week Restrictions", key="disable_restrictions_btn"):
                        st.session_state.scheduler.week_restrictions = {}
                        st.markdown(
                            '<div class="success-box">✅ All week restrictions have been disabled for this optimization run</div>',
                            unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                # Run the optimization button at the bottom
                if st.session_state.scheduler.weekly_calendar is not None:
                    st.markdown("### Run Optimization")
                    # This is just the modified section for the "Run Optimization" button
                    # in the Optimization Settings tab (tab2)

                    # Find this section in app.py around line 1800-1850
                    if st.button("Optimize Schedule", key="optimize_schedule_btn"):
                        try:
                            # Create containers for progress bar and output
                            progress_container = st.container()
                            output_container = st.container()
                            
                            # Initialize progress bar and status text
                            progress_bar = progress_container.progress(0)
                            
                            # Create a scrollable text area for logs instead of just a single line
                            with output_container:
                                st.markdown("### Optimization Progress")
                                log_area = st.empty()
                                log_messages = ["Initializing optimization..."]
                            
                            # Function to update progress and output - now keeps a history of messages
                            def update_progress(percent, message):
                                if percent is not None:
                                    progress_bar.progress(percent)
                                if message and message.strip():
                                    log_messages.append(message)
                                    # Display the last 15 messages (or all if fewer) to keep the display manageable
                                    displayed_msgs = log_messages[-15:]
                                    log_area.markdown('\n\n'.join(displayed_msgs))
                            
                            with st.spinner(f"Running optimization (maximum time: {solver_time} minutes)..."):
                                status, schedule_df, solver, schedule, trainer_assignments, unscheduled_courses = st.session_state.scheduler.run_optimization(
                                    monthly_weight=monthly_weight,
                                    affinity_weight=affinity_weight,
                                    utilization_target=utilization_target,
                                    solver_time_minutes=solver_time,
                                    num_workers=num_workers,
                                    min_course_spacing=min_course_spacing,
                                    solution_strategy="BALANCED", # Default to balanced strategy
                                    enforce_monthly_distribution=enforce_monthly,
                                    max_affinity_constraints=max_affinity,
                                    prioritize_all_courses=prioritize_all_courses,
                                    accelerated_mode=accelerated_mode,
                                    progress_callback=update_progress,  # Pass the progress callback
                                    enforce_champions=enforce_champions  # Pass the enforce_champions parameter
                                )

                                # Store the optimization status in session state
                                st.session_state.optimization_status = status
                                st.session_state.unscheduled_courses = unscheduled_courses if unscheduled_courses else []

                                # Check if the optimization was successful
                                if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                    # Store results in session state
                                    st.session_state.schedule_df = schedule_df

                                    # Generate the validation and utilization reports
                                    st.session_state.validation_df = st.session_state.scheduler.generate_monthly_validation(
                                        schedule, solver)
                                    st.session_state.utilization_df = st.session_state.scheduler.generate_trainer_utilization_report(
                                        schedule, trainer_assignments, solver)

                                    # Display success message with unscheduled course info
                                    if st.session_state.unscheduled_courses:
                                        st.warning(
                                            f"Optimization completed with {len(st.session_state.unscheduled_courses)} unscheduled courses. See Results tab for details.")
                                    else:
                                        st.success("Optimization completed successfully! All courses were scheduled.")
                                else:
                                    st.error(
                                        f"Optimization failed with status: {solver.StatusName(status)}. Try adjusting your parameters or check the Debug tab.")
                        except Exception as e:
                            st.error(f"An error occurred: {e}")
                            logging.error(f"Optimization error: {str(e)}")

        with tab3:
            if st.session_state.schedule_df is not None:
                st.markdown('<div class="section-header"><h2>4. Optimization Results</h2></div>', unsafe_allow_html=True)

                # Display tabs for different views
                result_tab1, result_tab2, result_tab3, result_tab4 = st.tabs(
                    ["📅 Schedule", "📊 Monthly Validation", "👥 Trainer Utilization", "📈 Visualizations"])

                with result_tab1:
                    st.subheader("Course Schedule")
                    st.dataframe(st.session_state.schedule_df, use_container_width=True)

                with result_tab2:
                    st.subheader("Monthly Validation")
                    st.dataframe(st.session_state.validation_df, use_container_width=True)

                with result_tab3:
                    st.subheader("Trainer Utilization")
                    st.dataframe(st.session_state.utilization_df, use_container_width=True)

                with result_tab4:
                    st.subheader("Visualizations")

                    # We need to rerun the plotting functions
                    # Add this code to the Results tab (tab3) section of your app.py
                    # Replace the existing visualization button code with this modified version

                    # Replace the visualization button code in the Results tab section with this version:

                    # Update the Generate Visualizations button to include unscheduled_courses in the returned values
                    # Find this in the result_tab4 (Visualizations) tab

                    if st.button("Generate Visualizations", key="generate_viz_btn"):
                        with st.spinner("Generating visualizations..."):
                            try:
                                # Only try to generate new optimization if we don't have a completed one
                                if not hasattr(st.session_state, 'schedule_df') or st.session_state.schedule_df is None:
                                    st.warning(
                                        "No optimization results found. Running a quick optimization to generate visualizations...")
                                    # Run a quick new optimization just to get solver state and schedule objects
                                    status, schedule_df, solver, schedule, trainer_assignments, _ = st.session_state.scheduler.run_optimization(
                                        monthly_weight=5,
                                        affinity_weight=2,
                                        utilization_target=70, solver_time_minutes=1,
                                        num_workers=8, min_course_spacing=2,
                                        solution_strategy="FIND_FEASIBLE_FAST",
                                        enforce_monthly_distribution=False,
                                        prioritize_all_courses=False,
                                        accelerated_mode=True,
                                        enforce_champions=True
                                    )

                                    if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                        st.error(
                                            f"Could not find a feasible solution for visualization. Status: {solver.StatusName(status)}")
                                        st.info(
                                            "Try running a full optimization with the 'Optimize Schedule' button first.")
                                        return
                                else:
                                    # Run a very short optimization to get the solver state and variables
                                    # but use the existing solution as a starting point
                                    st.info("Using existing optimization results for visualization...")
                                    status, _, solver, schedule, trainer_assignments, _ = st.session_state.scheduler.run_optimization(
                                        monthly_weight=5, 
                                        affinity_weight=2,
                                        utilization_target=70, solver_time_minutes=0.1,  # Very short time
                                        num_workers=8, min_course_spacing=2,
                                        solution_strategy="FIND_FEASIBLE_FAST",
                                        enforce_monthly_distribution=False,
                                        prioritize_all_courses=False,
                                        accelerated_mode=True,
                                        enforce_champions=True
                                    )

                                # Rest of the visualization code stays the same...

                                # Safety check that schedule and solver are valid
                                if schedule and solver and len(schedule) > 0:
                                    # Generate and display charts
                                    col1, col2 = st.columns(2)

                                    with col1:
                                        st.subheader("Weekly Course Distribution")
                                        try:
                                            fig1 = st.session_state.scheduler.plot_weekly_course_bar_chart(schedule,
                                                                                                           solver)
                                            st.pyplot(fig1)
                                        except Exception as e:
                                            st.error(f"Error generating weekly chart: {str(e)}")
                                            st.info("The visualization couldn't be generated due to data issues.")

                                    with col2:
                                        st.subheader("Trainer Workload")
                                        try:
                                            fig2 = st.session_state.scheduler.plot_trainer_workload_chart(
                                                schedule, trainer_assignments, solver)
                                            st.pyplot(fig2)
                                        except Exception as e:
                                            st.error(f"Error generating trainer chart: {str(e)}")
                                            st.info("The visualization couldn't be generated due to data issues.")
                                else:
                                    st.error("Couldn't generate valid solver state for visualization.")
                                    st.info(
                                        "Please try running a complete optimization first with the 'Optimize Schedule' button.")

                            except Exception as e:
                                st.error(f"Error generating visualizations: {str(e)}")
                                st.info(
                                    "Try running a complete optimization first with the 'Optimize Schedule' button.")
                        # Add this to your app.py file in the Results tab section,
                        # under the result_tab4 (Visualizations) tab:

                # Export Results
                st.markdown("### Export Results")
                st.write("Download your complete optimization results:")

                if st.button("Generate Excel Report", key="generate_excel_btn"):
                    with st.spinner("Generating Excel report..."):
                        output = st.session_state.scheduler.generate_excel_report(
                            st.session_state.schedule_df,
                            st.session_state.validation_df,
                            st.session_state.utilization_df
                        )

                        st.download_button(
                            label="Download Excel Report",
                            data=output,
                            file_name="course_schedule_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_btn"
                        )
            else:
                if st.session_state.optimization_status == cp_model.INFEASIBLE:
                    st.error(
                        "The last optimization run was INFEASIBLE. Please go to the 'Debug & Analyze' tab to diagnose and fix constraints before trying again.")
                elif st.session_state.optimization_status is not None:
                    st.error(
                        f"The last optimization run failed with status: {st.session_state.optimization_status}. Please adjust your parameters and try again.")
                else:
                    st.info("Run the optimization in the 'Optimization Settings' tab to see results here.")

                # Add this code to the Results tab section (tab3)
                # After loading and displaying the schedule dataframe

                # Instead, always access unscheduled_courses through session state:
                if 'unscheduled_courses' in st.session_state and st.session_state.unscheduled_courses:
                    st.warning(f"{len(st.session_state.unscheduled_courses)} courses could not be scheduled")
                    with st.expander("View Unscheduled Courses"):
                        unscheduled_df = pd.DataFrame(st.session_state.unscheduled_courses)
                        st.dataframe(unscheduled_df, use_container_width=True)


                # In the results tab, after displaying the schedule dataframe:
                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    st.warning(f"{len(st.session_state.unscheduled_courses)} courses could not be scheduled")
                    with st.expander("View Unscheduled Courses"):
                        unscheduled_df = pd.DataFrame(st.session_state.unscheduled_courses)
                        st.dataframe(unscheduled_df, use_container_width=True)

                        st.info("""
                        **Why couldn't these courses be scheduled?**
                        - Insufficient qualified trainers
                        - Trainer availability constraints
                        - Week restrictions limiting valid weeks
                        - Spacing requirements between course runs

                        Try enabling "Prioritize scheduling all courses" in the Optimization Settings tab or relaxing some constraints.
                        """)


        with tab4:
            if st.session_state.scheduler.course_run_data is None:
                st.info("Please load your data in the 'Data Input & Setup' tab first")
            elif st.session_state.scheduler.weekly_calendar is None:
                st.info("Please initialize the calendar before debugging")
            else:
                st.markdown('<div class="section-header"><h2>3. Debug & Analysis Tools</h2></div>', unsafe_allow_html=True)
                st.markdown("Use these tools to diagnose issues and optimize your schedule:")

                # Create two columns for side-by-side analysis tools
                debug_col1, debug_col2 = st.columns(2)

                with debug_col1:
                    st.markdown("### Incremental Constraint Diagnosis")
                    st.markdown("Find exactly which constraint is causing infeasibility:")

                    if st.button("Diagnose with Incremental Constraints", key="incremental_constraints_btn"):
                        with st.spinner("Running incremental constraint analysis..."):
                            status, diagnostics, schedule_df = st.session_state.scheduler.run_incremental_optimization()

                            # Display diagnostics in a table
                            st.subheader("Constraint Diagnosis Results")
                            diagnosis_df = pd.DataFrame(diagnostics)
                            st.dataframe(diagnosis_df, use_container_width=True)

                            # Interpretation
                            st.subheader("Interpretation")

                            # Find the first step that became infeasible
                            infeasible_step = None
                            for step in diagnostics:
                                if not step["feasible"]:
                                    infeasible_step = step["step"]
                                    break

                            if infeasible_step:
                                st.markdown(
                                    f'<div class="error-box">The model becomes infeasible when adding: {infeasible_step}</div>',
                                    unsafe_allow_html=True)

                                if infeasible_step == "Course spacing constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Try reducing the minimum spacing between course runs using the slider in the Optimization Settings tab.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Trainer availability constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> There might be insufficient trainer availability. Check your trainer leave periods or add more qualified trainers.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Week restriction constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Week restrictions are too limiting. Try disabling some week restrictions using the button in Advanced Settings.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Monthly distribution constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Monthly distribution requirements cannot be satisfied. Uncheck "Enforce Monthly Distribution as Hard Constraint" option.</div>',
                                        unsafe_allow_html=True)
                            else:
                                st.markdown(
                                    '<div class="success-box">✅ The model is feasible with all constraints! If you\'re still having issues with the full optimization, it might be due to complex interactions between constraints.</div>',
                                    unsafe_allow_html=True)

                            # If we got a feasible schedule, show it
                            if schedule_df is not None:
                                st.subheader("Feasible Schedule Found")
                                st.dataframe(schedule_df, use_container_width=True)

                with debug_col2:
                    st.markdown("### Visual Constraint Analysis")
                    st.markdown("Visualize constraints and potential bottlenecks:")

                    if st.button("Visual Constraint Analysis", key="visual_analysis_btn"):
                        with st.spinner("Generating visual analysis of constraints..."):
                            figures = st.session_state.scheduler.analyze_constraints_visually()

                            st.subheader("Summary Metrics")
                            st.pyplot(figures['summary_metrics'])

                            with st.expander("Monthly Demand Analysis", expanded=False):
                                st.pyplot(figures['monthly_analysis'])

                            with st.expander("Course-Week Availability", expanded=False):
                                st.pyplot(figures['course_week_heatmap'])

                            with st.expander("Trainer Analysis", expanded=False):
                                st.pyplot(figures['trainer_ratio_plot'])
                                st.pyplot(figures['trainer_avail_heatmap'])

                            st.info(
                                "Red areas in the heatmaps indicate restrictions or unavailability that could be causing infeasibility.")

            # Replace the function definition:
            def display_scheduling_problem_analysis():
                st.subheader("Analyze Unscheduled Course")

                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    # Create dropdown to select unscheduled course
                    course_options = [f"{c['Course']} (Run {c['Run']}, {c['Language']})" for c in
                                      st.session_state.unscheduled_courses]
                    selected_course = st.selectbox("Select course to analyze", course_options)

                    if selected_course:
                        # Extract course details from selection
                        idx = course_options.index(selected_course)
                        course_info = st.session_state.unscheduled_courses[idx]

                        # Run analysis
                        analysis = st.session_state.scheduler.analyze_unscheduled_course(
                            course_info['Course'], course_info['Language'], course_info['Run'])

                        # Display results
                        col1, col2 = st.columns(2)

                        with col1:
                            st.markdown("### Course Information")
                            st.markdown(f"**Course**: {course_info['Course']}")
                            st.markdown(f"**Language**: {course_info['Language']}")
                            st.markdown(f"**Run**: {course_info['Run']}")
                            st.markdown(f"**Qualified Trainers**: {analysis['qualified_trainers']}")
                            st.markdown(f"**Course Duration**: {analysis['duration']} days")

                            if analysis['qualified_trainers'] == 0:
                                st.error("❌ No qualified trainers for this course")
                                st.markdown("**Solution**: Add more qualified trainers in the Fleximatrix")

                        with col2:
                            st.markdown("### Week Availability")
                            st.markdown(f"**Weeks with enough workdays**: {analysis['weeks_with_enough_days']}")

                            if 'restricted_week_positions' in analysis:
                                restrictions = analysis['restricted_week_positions']
                                if restrictions:
                                    st.markdown(f"**Restricted week positions**: {', '.join(restrictions)}")
                                    st.markdown(f"**Weeks eliminated by restrictions**: {analysis['weeks_restricted']}")

                            st.markdown(f"**Weeks with any trainer available**: {analysis['weeks_with_any_trainer']}")

                            if analysis['weeks_with_any_trainer'] == 0:
                                st.error("❌ No weeks with both enough workdays and available trainers")
                                st.markdown("**Solution**: Reduce leave overlaps or add more qualified trainers")

                        # Rest of the function...
                        # Draw a timeline of week availability
                        if analysis['qualified_trainers'] > 0:
                            # ... (rest of the function code)
                            pass  # Keep the rest of the function as is
                else:
                    st.info(
                        "No unscheduled courses to analyze. Run an optimization first to identify scheduling problems.")

            # And then modify how it's used in tab4:
            # Instead of:
            # with st.expander("Course Scheduling Problem Analysis", expanded=False):
            #     display_scheduling_problem_analysis()

            # Replace with inline code:
            with st.expander("Course Scheduling Problem Analysis", expanded=False):
                st.subheader("Analyze Unscheduled Course")

                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    # Create dropdown to select unscheduled course
                    course_options = [f"{c['Course']} (Run {c['Run']}, {c['Language']})" for c in
                                      st.session_state.unscheduled_courses]
                    selected_course = st.selectbox("Select course to analyze", course_options)

                    if selected_course:
                        # Extract course details from selection
                        idx = course_options.index(selected_course)
                        course_info = st.session_state.unscheduled_courses[idx]

                        # Run analysis
                        analysis = st.session_state.scheduler.analyze_unscheduled_course(
                            course_info['Course'], course_info['Language'], course_info['Run'])

                        # Display results
                        col1, col2 = st.columns(2)

                        with col1:
                            st.markdown("### Course Information")
                            st.markdown(f"**Course**: {course_info['Course']}")
                            st.markdown(f"**Language**: {course_info['Language']}")
                            st.markdown(f"**Run**: {course_info['Run']}")
                            st.markdown(f"**Qualified Trainers**: {analysis['qualified_trainers']}")
                            st.markdown(f"**Course Duration**: {analysis['duration']} days")

                            if analysis['qualified_trainers'] == 0:
                                st.error("❌ No qualified trainers for this course")
                                st.markdown("**Solution**: Add more qualified trainers in the Fleximatrix")

                        with col2:
                            st.markdown("### Week Availability")
                            st.markdown(f"**Weeks with enough workdays**: {analysis['weeks_with_enough_days']}")

                            if 'restricted_week_positions' in analysis:
                                restrictions = analysis['restricted_week_positions']
                                if restrictions:
                                    st.markdown(f"**Restricted week positions**: {', '.join(restrictions)}")
                                    st.markdown(f"**Weeks eliminated by restrictions**: {analysis['weeks_restricted']}")

                            st.markdown(f"**Weeks with any trainer available**: {analysis['weeks_with_any_trainer']}")

                            if analysis['weeks_with_any_trainer'] == 0:
                                st.error("❌ No weeks with both enough workdays and available trainers")
                                st.markdown("**Solution**: Reduce leave overlaps or add more qualified trainers")

                        # Draw a timeline of week availability
                        if analysis['qualified_trainers'] > 0:
                            st.markdown("### Week-by-Week Availability")

                            # Create dataframe for visualization
                            weeks = sorted(analysis.get('weeks_with_trainers', {}).keys())
                            if weeks:
                                week_data = []

                                for week in weeks:
                                    trainers_available = analysis['weeks_with_trainers'].get(week, 0)
                                    month = st.session_state.scheduler.week_to_month_map.get(week, 0)
                                    week_info = st.session_state.scheduler.week_position_in_month.get(week, {})

                                    # Check if this week is restricted
                                    is_restricted = False
                                    if course_info['Course'] in st.session_state.scheduler.week_restrictions:
                                        restrictions = st.session_state.scheduler.week_restrictions[
                                            course_info['Course']]
                                        if (week_info.get('is_first') and restrictions.get('First', False)) or \
                                                (week_info.get('is_second') and restrictions.get('Second', False)) or \
                                                (week_info.get('is_third') and restrictions.get('Third', False)) or \
                                                (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or \
                                                (week_info.get('is_last') and restrictions.get('Last', False)):
                                            is_restricted = True

                                    working_days = st.session_state.scheduler.weekly_working_days.get(week, 0)

                                    week_data.append({
                                        'Week': week,
                                        'Month': month,
                                        'Trainers Available': trainers_available,
                                        'Working Days': working_days,
                                        'Is Restricted': is_restricted
                                    })

                                week_df = pd.DataFrame(week_data)

                                # Color columns based on conditions
                                def color_trainers(val):
                                    if val == 0:
                                        color = 'red'
                                    elif val < 3:
                                        color = 'orange'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_days(val, duration=analysis['duration']):
                                    if val < duration:
                                        color = 'red'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_restricted(val):
                                    return 'background-color: red' if val else 'background-color: white'

                                # Apply styling
                                styled_df = week_df.style.applymap(color_trainers, subset=['Trainers Available']) \
                                    .applymap(lambda x: color_days(x, analysis['duration']), subset=['Working Days']) \
                                    .applymap(color_restricted, subset=['Is Restricted'])

                                st.dataframe(styled_df)

                                # Identify the problem
                                problems = []
                                if analysis['weeks_with_enough_days'] == 0:
                                    problems.append("❌ No weeks have enough working days for this course duration")
                                if analysis.get('weeks_restricted') == analysis['weeks_with_enough_days']:
                                    problems.append("❌ All potential weeks are restricted by week position constraints")
                                if analysis['weeks_with_any_trainer'] == 0:
                                    problems.append("❌ No trainers are available in valid weeks")

                                st.markdown("### Summary")

                                if problems:
                                    for problem in problems:
                                        st.error(problem)
                                else:
                                    st.warning(
                                        "⚠️ This course may be unscheduled due to complex constraint interactions")

                                st.markdown("### Recommended Actions")
                                actions = []

                                if analysis['qualified_trainers'] < 2:
                                    actions.append("✅ Add more qualified trainers for this course")
                                if analysis['duration'] > 3 and analysis['weeks_with_enough_days'] < 10:
                                    actions.append(
                                        "✅ Review public holidays that may be limiting available working days")
                                if analysis.get('restricted_week_positions'):
                                    actions.append(
                                        "✅ Relax week position restrictions (use 'Disable All Week Restrictions' button)")
                                if analysis['weeks_with_any_trainer'] < 5:
                                    actions.append("✅ Review trainer leave patterns to reduce overlapping leaves")

                                # Always suggest prioritizing courses
                                actions.append(
                                    "✅ Enable 'Prioritize scheduling all courses' option in Optimization Settings")

                                for action in actions:
                                    st.markdown(action)
                else:
                    st.info(
                        "No unscheduled courses to analyze. Run an optimization first to identify scheduling problems.")
        pass
    except Exception as e:
        logging.error(f"Application error: {e}")
        st.error(f"An error occurred: {e}")


if __name__ == "__main__":
    # Add this before your main() call
    st.write("App is running")
    main()

