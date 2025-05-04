import pandas as pd
import datetime
import streamlit as st
st.set_page_config(page_title="Course Scheduler", layout="wide")
import matplotlib.pyplot as plt
from ortools.sat.python import cp_model
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import threading
import time
import os
import json
import pickle
import shutil
import numpy as np
import queue
from threading import Lock

# Set up logging at the beginning of your app.py
logging.basicConfig(level=logging.ERROR)

# Add before the TimeoutCallback class
message_queue = queue.Queue()
solution_queue = queue.Queue()
update_lock = Lock()

def save_intermediate_results(schedule_df, validation_df, workload_df, unscheduled, solver_status=None):
    """Save intermediate results to files in case of connection timeout"""
    try:
        # Create temp directory if it doesn't exist
        if not os.path.exists("temp_results"):
            os.makedirs("temp_results")
        
        # Save results to files
        schedule_df.to_pickle("temp_results/schedule.pkl")
        validation_df.to_pickle("temp_results/validation.pkl")
        workload_df.to_pickle("temp_results/workload.pkl")
        
        # Save unscheduled courses and solver status
        with open("temp_results/metadata.json", "w") as f:
            json.dump({
                "unscheduled": unscheduled,
                "solver_status": solver_status,
                "timestamp": datetime.datetime.now().isoformat(),
                "last_save": datetime.datetime.now().isoformat()
            }, f)
            
    except Exception as e:
        logging.error(f"Error saving intermediate results: {e}")

def load_saved_results():
    """Load saved optimization results if they exist"""
    try:
        if not os.path.exists("temp_results"):
            return None, None, None, None, None, None
            
        # Load metadata first
        with open("temp_results/metadata.json", "r") as f:
            metadata = json.load(f)
            timestamp = datetime.datetime.fromisoformat(metadata["timestamp"])
            solver_status = metadata.get("solver_status")
            unscheduled = metadata.get("unscheduled", [])
            
        # Check if saved results are too old (more than 24 hours)
        if (datetime.datetime.now() - timestamp).total_seconds() > 86400:
            cleanup_temp_results()
            return None, None, None, None, None, None
            
        # Load DataFrames
        schedule_df = pd.read_pickle("temp_results/schedule.pkl")
        validation_df = pd.read_pickle("temp_results/validation.pkl")
        workload_df = pd.read_pickle("temp_results/workload.pkl")
            
        return schedule_df, validation_df, workload_df, unscheduled, timestamp, solver_status
        
    except Exception as e:
        logging.error(f"Error loading results: {e}")
        return None, None, None, None, None, None

def cleanup_temp_results():
    """Remove temporary result files"""
    try:
        if os.path.exists("temp_results"):
            shutil.rmtree("temp_results")
    except Exception as e:
        logging.error(f"Error cleaning up temp files: {e}")

def save_results(schedule_df, validation_df, workload_df, unscheduled):
    """Save optimization results to files"""
    try:
        # Create temp directory if it doesn't exist
        if not os.path.exists("temp_results"):
            os.makedirs("temp_results")
        
        # Save each DataFrame to a pickle file
        schedule_df.to_pickle("temp_results/schedule.pkl")
        validation_df.to_pickle("temp_results/validation.pkl")
        workload_df.to_pickle("temp_results/workload.pkl")
        
        # Save unscheduled courses
        with open("temp_results/unscheduled.pkl", "wb") as f:
            pickle.dump(unscheduled, f)
            
        # Save timestamp
        with open("temp_results/timestamp.txt", "w") as f:
            f.write(str(datetime.datetime.now()))
            
    except Exception as e:
        print(f"Error saving results: {e}")

def load_saved_results():
    """Load saved optimization results if they exist"""
    try:
        if not os.path.exists("temp_results"):
            return None, None, None, None, None
            
        # Check timestamp
        with open("temp_results/timestamp.txt", "r") as f:
            timestamp = datetime.datetime.strptime(f.read().strip(), "%Y-%m-%d %H:%M:%S.%f")
            
        # Load DataFrames
        schedule_df = pd.read_pickle("temp_results/schedule.pkl")
        validation_df = pd.read_pickle("temp_results/validation.pkl")
        workload_df = pd.read_pickle("temp_results/workload.pkl")
        
        # Load unscheduled courses
        with open("temp_results/unscheduled.pkl", "rb") as f:
            unscheduled = pickle.load(f)
            
        return schedule_df, validation_df, workload_df, unscheduled, timestamp
        
    except Exception as e:
        print(f"Error loading results: {e}")
        return None, None, None, None, None

def cleanup_temp_results():
    """Remove temporary result files"""
    try:
        if os.path.exists("temp_results"):
            shutil.rmtree("temp_results")
    except Exception as e:
        print(f"Error cleaning up temp files: {e}")

# Import template generator
from template_generator import create_excel_template

# At the top of the file, after imports
def get_penalty_weights():
    """Get penalty weights from session state or use defaults"""
    return {
        'monthly': st.session_state.get('monthly_weight', 5),
        'below_min': st.session_state.get('below_min_weight', 50),
        'above_max': st.session_state.get('above_max_weight', 10),
        'affinity': st.session_state.get('affinity_weight', 2)
    }

class CourseScheduler:
    def __init__(self):
        self.course_run_data = None
        self.consultant_data = None
        self.priority_data = None
        self.annual_leaves = None
        self.affinity_matrix_data = None
        self.public_holidays_data = None
        self.weekly_calendar = None
        self.week_to_month_map = {}
        self.weekly_working_days = {}
        self.fleximatrix = {}  # Dictionary mapping (course, language) to list of qualified trainers
        self.course_champions = {}  # Dictionary to store champions for each course
        self.high_volume_courses = {}  # Dict to store courses with >5 combined runs
        self.course_runs_by_language = {}  # Dict to store run counts by language

    def _sample_courses_for_optimization(self, course_data, max_courses=500):
        """
        Sample courses for optimization when dealing with extremely large problems.
        Prioritizes courses based on various factors.
        
        Args:
            course_data: DataFrame containing course information
            max_courses: Maximum number of courses to include in optimization
            
        Returns:
            DataFrame containing sampled courses
        """
        # Create a copy to avoid modifying original
        df = course_data.copy()
        
        # Calculate priority score for each course
        df['priority_score'] = 0
        
        # Factor 1: Number of qualified trainers (fewer trainers = higher priority)
        df['qualified_trainers'] = df.apply(
            lambda row: len(self.fleximatrix.get((row['Course Name'], row['Language']), [])), 
            axis=1
        )
        df['trainer_score'] = 1 / (df['qualified_trainers'] + 1)  # Add 1 to avoid division by zero
        
        # Factor 2: Course duration (longer courses = higher priority)
        df['duration_score'] = df['Duration'] / df['Duration'].max()
        
        # Factor 3: Number of runs (more runs = higher priority)
        df['runs_score'] = df['Runs'] / df['Runs'].max()
        
        # Factor 4: Champion courses get priority
        df['is_champion'] = df.apply(
            lambda row: 1 if (row['Course Name'], row['Language']) in self.course_champions else 0,
            axis=1
        )
        
        # Combine factors into final priority score
        df['priority_score'] = (
            df['trainer_score'] * 0.3 +  # Limited trainers is important
            df['duration_score'] * 0.2 +  # Longer courses need more planning
            df['runs_score'] * 0.3 +      # More runs need more coordination
            df['is_champion'] * 0.2        # Champion courses get some priority
        )
        
        # Sort by priority score and select top courses
        df = df.sort_values('priority_score', ascending=False)
        
        # Calculate total runs and trim until we're under max_courses
        total_runs = df['Runs'].cumsum()
        df = df[total_runs <= max_courses].copy()
        
        # Drop the temporary columns we added
        df = df.drop(['priority_score', 'qualified_trainers', 'trainer_score', 
                     'duration_score', 'runs_score', 'is_champion'], axis=1)
        
        print(f"Sampled {len(df)} courses with {df['Runs'].sum()} total runs")
        return df

    def load_data_from_excel(self, uploaded_file):
        """Load all required data from an Excel file with multiple sheets"""
        try:
            # Read the Excel file with multiple sheets
            excel_file = pd.ExcelFile(uploaded_file)

            # Load each required sheet
            self.course_run_data = pd.read_excel(excel_file, sheet_name='CourseData')
            self.consultant_data = pd.read_excel(excel_file, sheet_name='TrainerData')
            self.priority_data = pd.read_excel(excel_file, sheet_name='PriorityData')
            self.annual_leaves = pd.read_excel(excel_file, sheet_name='AnnualLeaves')
            self.affinity_matrix_data = pd.read_excel(excel_file, sheet_name='AffinityMatrix')
            self.public_holidays_data = pd.read_excel(excel_file, sheet_name='PublicHolidays')

            # Parse date columns
            date_columns = ['Start_Date', 'End_Date']
            for col in date_columns:
                if col in self.annual_leaves.columns:
                    self.annual_leaves[col] = pd.to_datetime(self.annual_leaves[col]).dt.date

            if 'Start Date' in self.public_holidays_data.columns:
                self.public_holidays_data['Start Date'] = pd.to_datetime(
                    self.public_holidays_data['Start Date']).dt.date

            if 'End Date' in self.public_holidays_data.columns:
                self.public_holidays_data['End Date'] = pd.to_datetime(self.public_holidays_data['End Date']).dt.date

            # Load fleximatrix - which trainers can teach which courses (wide format)
            flexi_sheet = pd.read_excel(excel_file, sheet_name='Fleximatrix')

            # Process the wide-format Fleximatrix
            self.fleximatrix, self.course_champions = self.process_wide_fleximatrix(flexi_sheet)

            # Load adjusted monthly demand data
            self.monthly_demand = pd.read_excel(excel_file, sheet_name='MonthlyDemand')

            # Load course week restrictions (if exists)
            self.week_restrictions = {}
            try:
                # Check if the sheet exists
                restrictions_sheet = pd.read_excel(excel_file, sheet_name='WeekRestrictions')

                # Process each restriction
                for _, row in restrictions_sheet.iterrows():
                    course = row['Course']
                    week_type = row['Week Type']  # 'First', 'Second', 'Third', 'Fourth', 'Last'
                    restricted = row['Restricted']  # Boolean: True = cannot run, False = can run

                    if course not in self.week_restrictions:
                        self.week_restrictions[course] = {}

                    self.week_restrictions[course][week_type] = restricted

                print(f"Loaded {len(restrictions_sheet)} week restrictions for courses")
            except Exception as e:
                print(f"No week restrictions found or error loading them: {e}")
                self.week_restrictions = {}

            return True

        except Exception as e:
            st.error(f"Error loading data: {e}")
            return False

    def process_wide_fleximatrix(self, flexi_sheet):
        """Process a wide-format Fleximatrix where consultants are columns"""
        fleximatrix = {}
        course_champions = {}

        # Get list of consultants (all columns except CourseName, CategoryName, Language, and Champion)
        consultants = [col for col in flexi_sheet.columns
                       if col not in ['CourseName', 'CategoryName', 'Language', 'Champion']]

        # Process each row (course+language combination)
        for _, row in flexi_sheet.iterrows():
            course = row['CourseName']
            language = row['Language']
            champion = row['Champion']

            # Initialize empty list for this course+language
            fleximatrix[(course, language)] = []

            # Add the champion
            if champion and str(champion).strip():
                course_champions[(course, language)] = champion

            # Add all qualified consultants (marked with "U")
            for consultant in consultants:
                if row[consultant] == "U":
                    fleximatrix[(course, language)].append(consultant)

        return fleximatrix, course_champions


    def initialize_calendar(self, year, weekend_selection):
        """Initialize the weekly calendar based on year and weekend selection"""

        # Function to Generate Work Week Start Dates
        def get_week_start_dates(year, weekend_selection):
            first_day = datetime.date(year, 1, 1)
            if weekend_selection == "FS":  # Friday-Saturday
                week_start = 6  # Sunday
            else:  # "Saturday-Sunday"
                week_start = 0  # Monday

            weeks = []
            current_date = first_day

            # Find the first day that matches our week start
            while current_date.weekday() != week_start:
                current_date += datetime.timedelta(days=1)

            # Generate all week start dates for the year
            while current_date.year == year:
                weeks.append(current_date)
                current_date += datetime.timedelta(days=7)  # Move to next week

            return weeks

        # Generate Work Week Start Dates
        self.weekly_calendar = get_week_start_dates(year, weekend_selection)

        # Ensure correct mapping of weeks to months
        self.week_to_month_map = {}
        self.week_position_in_month = {}  # Track position of week within its month

        # First, map weeks to months
        for i, week_start in enumerate(self.weekly_calendar):
            month = week_start.month  # Extract the month from the start date
            self.week_to_month_map[i + 1] = month  # Map week number to month

        # Then, calculate position of each week within its month
        for week_num, month in self.week_to_month_map.items():
            # Get all weeks in this month
            month_weeks = [w for w, m in self.week_to_month_map.items() if m == month]
            month_weeks.sort()

            # Find position (1-indexed)
            position = month_weeks.index(week_num) + 1

            # Also mark if it's the last week of the month
            is_last = (position == len(month_weeks))

            # Store as a dictionary with position info
            self.week_position_in_month[week_num] = {
                'position': position,
                'total_weeks': len(month_weeks),
                'is_first': (position == 1),
                'is_second': (position == 2),
                'is_third': (position == 3),
                'is_fourth': (position == 4),
                'is_last': is_last
            }

        # Calculate working days for each week accounting for public holidays
        self.calculate_working_days()

        return True


    def calculate_working_days(self):
        """Calculate working days for each week accounting for public holidays"""
        self.weekly_working_days = {}

        for i, week_start in enumerate(self.weekly_calendar):
            week_num = i + 1
            working_days = 5  # Assume a full workweek

            # Check if this week is completely inside a long holiday
            fully_inside_long_holiday = False

            for _, row in self.public_holidays_data.iterrows():
                holiday_length = (row["End Date"] - row["Start Date"]).days + 1

                # If it's a long holiday (25+ days) and the week is fully inside it
                if holiday_length >= 25:
                    if week_start >= row["Start Date"] and (week_start + datetime.timedelta(days=4)) <= row["End Date"]:
                        fully_inside_long_holiday = True
                        break

            # If fully inside a long holiday, block the week completely
            if fully_inside_long_holiday:
                working_days = 0
            else:
                # Adjust only affected days for regular holidays
                for _, row in self.public_holidays_data.iterrows():
                    holiday_days_in_week = sum(
                        1 for d in range(5)  # Check only workdays
                        if row["Start Date"] <= (week_start + datetime.timedelta(days=d)) <= row["End Date"]
                        and (week_start + datetime.timedelta(days=d)).weekday() not in [5, 6]  # Exclude weekends
                    )
                    working_days -= holiday_days_in_week
                    working_days = max(0, working_days)  # Ensure no negative values

            self.weekly_working_days[week_num] = working_days


    def is_trainer_available(self, trainer_name, week_num):
        """Check if trainer is available during the given week (not on leave)"""
        week_start = self.weekly_calendar[week_num - 1]
        week_end = week_start + datetime.timedelta(days=4)  # Assuming 5-day courses

        # Check against all leave periods
        for _, leave in self.annual_leaves[self.annual_leaves["Name"] == trainer_name].iterrows():
            # If leave period overlaps with course week, trainer is unavailable
            if (leave["Start_Date"] <= week_end and leave["End_Date"] >= week_start):
                return False
        return True


    def get_trainer_priority(self, trainer_name):
        """Get priority level for a trainer based on their title"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        priority = self.priority_data.loc[self.priority_data["Title"] == title, "Priority"].iloc[0]
        return priority


    def is_freelancer(self, trainer_name):
        """Check if a trainer is a freelancer"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        return title == "Freelancer"

    def _create_objective_function(self, model, penalties, mode="regular", weights=None):
        """Create the objective function based on the specified mode and weights."""
        default_weights = {
            "regular": {
                "affinity": 2,
                "below_min_days": 50,  # Very high penalty for being below minimum
                "above_max_days": 10,   # Lower penalty for exceeding maximum
                "monthly": 5
            },
            "prioritize_all": {
                "unscheduled": 50,
                "below_min_days": 50,
                "above_max_days": 10,
                "spacing": 20,
                "monthly": 5,
                "affinity": 2
            }
        }

        # Use provided weights or defaults
        active_weights = weights or default_weights[mode]

        if mode == "regular":
            return model.Minimize(
                active_weights["affinity"] * sum(penalties["affinity"]) +
                active_weights["below_min_days"] * sum(penalties["below_min_days"]) +
                active_weights["above_max_days"] * sum(penalties["above_max_days"]) +
                active_weights["monthly"] * sum(penalties["monthly"])
            )
        elif mode == "prioritize_all":
            return model.Minimize(
                50 * sum(penalties["unscheduled"]) +
                50 * sum(penalties["below_min_days"]) +
                10 * sum(penalties["above_max_days"]) +
                20 * sum(penalties["spacing"]) +
                5 * sum(penalties["monthly"]) +
                2 * sum(penalties["affinity"])
            )
        else:
            raise ValueError(f"Unknown optimization mode: {mode}")

    def run_optimization(self, monthly_weight=5, below_min_weight=50, above_max_weight=10,
                         affinity_weight=2, solver_time_minutes=5, num_workers=8,
                         min_course_spacing=2, solution_strategy="BALANCED",
                         enforce_monthly_distribution=False, prioritize_all_courses=False):
        """Run the course scheduling optimization with better handling of long runs"""
        
        # Initialize model and variables
        model = cp_model.CpModel()
        schedule = {}
        trainer_assignments = {}
        max_weeks = len(self.weekly_calendar)

        # Create the schedule variables for each course and run
        print("Creating variables for courses and trainers...")
        
        # All penalty lists
        month_deviation_penalties = []
        affinity_penalties = []
        unscheduled_course_penalties = []  # For unscheduled courses
        workload_violation_penalties = []  # For workload violations
        spacing_violation_penalties = []   # For course spacing violations
        priority_penalties = []  # For priority-based penalties

        # After initializing variables but before creating schedule variables
        # Identify high-volume courses first
        # print("Identifying high-volume courses...")
        # high_volume_courses = {}  # Dict to store courses with >5 combined runs
        # course_runs_by_language = {}  # Dict to store run counts by language

        # # Group runs by course name and language
        # for _, row in self.course_run_data.iterrows():
        #     course = row["Course Name"]
        #     language = row["Language"]
        #     runs = row["Runs"]
            
        #     if course not in course_runs_by_language:
        #         course_runs_by_language[course] = {"Arabic": 0, "English": 0, "total": 0}
            
        #     course_runs_by_language[course][language] += runs
        #     course_runs_by_language[course]["total"] += runs

        # # Identify courses with more than 5 combined runs
        # for course, runs_info in course_runs_by_language.items():
        #     if runs_info["total"] >= 5:
        #         high_volume_courses[course] = runs_info
        #         print(f"High-volume course found: {course}")
        #         print(f"  Arabic runs: {runs_info['Arabic']}")
        #         print(f"  English runs: {runs_info['English']}")
        #         print(f"  Total runs: {runs_info['total']}")

        # # Store high-volume courses in the class for later use
        # self.high_volume_courses = high_volume_courses

        # Create the schedule variables for each course and run
        for _, row in self.course_run_data.iterrows():
            course, delivery_type, language, runs, duration = row["Course Name"], row["Delivery Type"], row["Language"], \
                row["Runs"], row["Duration"]

            for i in range(runs):
                # Create a variable for the start week that includes 0 (unscheduled)
                start_week = model.NewIntVar(0, max_weeks, f"start_week_{course}_{i}")
                schedule[(course, delivery_type, language, i)] = start_week

                # Create trainer assignment variable
                qualified_trainers = self.fleximatrix.get((course, language), [])
                if not qualified_trainers:
                    print(f"Warning: No qualified trainers for {course} ({language})")
                    continue

                trainer_var = model.NewIntVar(0, len(qualified_trainers) - 1, f"trainer_{course}_{i}")
                trainer_assignments[(course, delivery_type, language, i)] = trainer_var

                # Track if course is scheduled
                is_scheduled = model.NewBoolVar(f"{course}_{i}_is_scheduled")
                model.Add(start_week > 0).OnlyEnforceIf(is_scheduled)
                model.Add(start_week == 0).OnlyEnforceIf(is_scheduled.Not())
                unscheduled_course_penalties.append(is_scheduled.Not())

                # Only schedule in weeks with enough working days AND available trainers
                valid_weeks = []
                for w, days in self.weekly_working_days.items():
                    if days >= duration:
                        for trainer in qualified_trainers:
                            if self.is_trainer_available(trainer, w):
                                valid_weeks.append(w)
                                break

                # HARD CONSTRAINT: Only allow start_week to be in valid_weeks
                if valid_weeks:
                    model.AddAllowedAssignments([start_week], [[w] for w in valid_weeks] + [[0]])  # 0 means unscheduled
                else:
                    model.Add(start_week == 0)

        # Add special handling for high-volume courses
        if self.high_volume_courses:  # Use local variable instead of self.high_volume_courses
            high_volume_vars = self._add_high_volume_course_constraints(
                model, schedule, min_course_spacing, month_deviation_penalties
            )

        # CONSTRAINT 1: Monthly distribution
        print("Adding monthly distribution constraints...")
        total_f2f_runs = sum(self.course_run_data["Runs"])
        adjusted_f2f_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # For each month, track and enforce
        for month in range(1, 13):
            target_demand = adjusted_f2f_demand.get(month, 0)
            if target_demand == 0:
                continue

            courses_in_month = []
            month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]

            if not month_weeks:
                continue

            # For each course run, create a Boolean indicating if it's in this month
            for (course, delivery_type, language, i), week_var in schedule.items():
                is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")

                # Add constraints: is_in_month is True iff week_var is in month's weeks
                week_choices = []
                for week in month_weeks:
                    is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                    week_choices.append(is_in_this_week)

                model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())

                courses_in_month.append(is_in_month)

            # Enforce the target demand for this month
            if courses_in_month and target_demand > 0:
                if enforce_monthly_distribution:
                    model.Add(sum(courses_in_month) == target_demand)
                else:
                    month_deviation = model.NewIntVar(0, total_f2f_runs, f"month_{month}_deviation")
                    actual_courses = sum(courses_in_month)
                    
                    # Model absolute value: |actual - target|
                    is_over_target = model.NewBoolVar(f"month_{month}_over_target")
                    model.Add(actual_courses >= target_demand).OnlyEnforceIf(is_over_target)
                    model.Add(actual_courses < target_demand).OnlyEnforceIf(is_over_target.Not())

                    over_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_over_dev")
                    model.Add(over_dev == actual_courses - target_demand).OnlyEnforceIf(is_over_target)
                    model.Add(over_dev == 0).OnlyEnforceIf(is_over_target.Not())

                    under_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_under_dev")
                    model.Add(under_dev == target_demand - actual_courses).OnlyEnforceIf(is_over_target.Not())
                    model.Add(under_dev == 0).OnlyEnforceIf(is_over_target)

                    model.Add(month_deviation == over_dev + under_dev)
                    month_deviation_penalties.extend([month_deviation] * monthly_weight)

        # CONSTRAINT 2: Course spacing
        print("Adding course spacing constraints...")
        for course_name in set(self.course_run_data["Course Name"]):
            course_runs = []
            for (course, delivery_type, language, i), var in schedule.items():
                if course == course_name:
                    course_runs.append((i, var))

            course_runs.sort(key=lambda x: x[0])

            for i in range(len(course_runs) - 1):
                run_num1, var1 = course_runs[i]
                run_num2, var2 = course_runs[i + 1]
                model.Add(var2 >= var1 + min_course_spacing)

        # CONSTRAINT 3: Trainer workload
        print("Adding trainer workload constraints...")
        trainer_workload = {name: [] for name in self.consultant_data["Name"]}

        # Track course assignments for each trainer
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            qualified_trainers = self.fleximatrix.get((course, language), [])
            course_champion = self.course_champions.get((course, language))

            for t_idx, trainer in enumerate(qualified_trainers):
                is_assigned = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}")
                model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned)
                model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned.Not())

                # Accumulate workload
                trainer_workload[trainer].append((is_assigned, duration))

                # Add priority-based penalties
                if trainer == course_champion:
                    priority = 1
                else:
                    title = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Title"].iloc[0]
                    priority = self.priority_data.loc[self.priority_data["Title"] == title, "Priority"].iloc[0]

                priority_penalties.extend([is_assigned] * priority)

        # Calculate and constrain workload
        for trainer, workload_items in trainer_workload.items():
            if not workload_items:
                continue

            max_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Max_Days"].iloc[0]
            min_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Min_Days"].iloc[0]

            # Calculate total workload
            total_workload = model.NewIntVar(0, max_days * 2, f"total_workload_{trainer}")
            weighted_terms = []
            for is_assigned, duration in workload_items:
                term = model.NewIntVar(0, duration, f"term_{id(is_assigned)}_{duration}")
                model.Add(term == duration).OnlyEnforceIf(is_assigned)
                model.Add(term == 0).OnlyEnforceIf(is_assigned.Not())
                weighted_terms.append(term)
            model.Add(total_workload == sum(weighted_terms))

            # Add penalties for workload violations
            if min_days > 0:
                below_min = model.NewBoolVar(f"{trainer}_below_min")
                model.Add(total_workload < min_days).OnlyEnforceIf(below_min)
                model.Add(total_workload >= min_days).OnlyEnforceIf(below_min.Not())
                workload_violation_penalties.extend([below_min] * below_min_weight)

            above_max = model.NewBoolVar(f"{trainer}_above_max")
            model.Add(total_workload > max_days).OnlyEnforceIf(above_max)
            model.Add(total_workload <= max_days).OnlyEnforceIf(above_max.Not())
            workload_violation_penalties.extend([above_max] * above_max_weight)

        # CONSTRAINT: Trainer can only teach one course per week
        print("Adding constraint: trainer can only teach one course per week")
        for trainer in self.consultant_data["Name"]:
            for week in range(1, max_weeks + 1):
                assignments = []
                for (course, delivery_type, language, i), week_var in schedule.items():
                    if (course, delivery_type, language, i) not in trainer_assignments:
                        continue
                    qualified_trainers = self.fleximatrix.get((course, language), [])
                    if trainer not in qualified_trainers:
                        continue
                    t_idx = qualified_trainers.index(trainer)
                    trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                    is_this_week = model.NewBoolVar(f"{course}_{i}_{trainer}_in_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_this_week.Not())
                    is_this_trainer = model.NewBoolVar(f"{course}_{i}_{trainer}_assigned")
                    model.Add(trainer_var == t_idx).OnlyEnforceIf(is_this_trainer)
                    model.Add(trainer_var != t_idx).OnlyEnforceIf(is_this_trainer.Not())
                    is_assigned = model.NewBoolVar(f"{course}_{i}_{trainer}_assigned_in_week_{week}")
                    model.AddBoolAnd([is_this_week, is_this_trainer]).OnlyEnforceIf(is_assigned)
                    model.AddBoolOr([is_this_week.Not(), is_this_trainer.Not()]).OnlyEnforceIf(is_assigned.Not())
                    assignments.append(is_assigned)
                if assignments:
                    model.Add(sum(assignments) <= 1)

        # After the trainer workload constraints but before the objective function
        # CONSTRAINT 4: Course affinities
        print("Adding affinity constraints for course pairs...")

        # Add constraints for course affinities (soft constraints/penalties)
        for _, row in self.affinity_matrix_data.iterrows():
            c1, c2, gap_weeks = row["Course 1"], row["Course 2"], row["Gap Weeks"]

            c1_runs = []
            c2_runs = []

            for (course, _, _, i), var in schedule.items():
                if course == c1:
                    c1_runs.append((i, var))
                elif course == c2:
                    c2_runs.append((i, var))

            # Skip if either course doesn't have any runs
            if not c1_runs or not c2_runs:
                continue

            # Sort by run index
            c1_runs.sort(key=lambda x: x[0])
            c2_runs.sort(key=lambda x: x[0])

            # Only check first run of each course to reduce constraints
            run1, var1 = c1_runs[0]
            run2, var2 = c2_runs[0]

            # Create variables to check if either course is in Q4
            is_c1_q4 = model.NewBoolVar(f"{c1}_{run1}_in_last_term")
            is_c2_q4 = model.NewBoolVar(f"{c2}_{run2}_in_last_term")

            # Get last term weeks (Sep-Dec, months 9-12)
            last_term_weeks = [w for w, m in self.week_to_month_map.items() if m in [9, 10, 11, 12]]

            # Set up last term detection for course 1
            c1_last_term_choices = []
            for week in last_term_weeks:
                is_in_this_week = model.NewBoolVar(f"{c1}_{run1}_in_week_{week}")
                model.Add(var1 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var1 != week).OnlyEnforceIf(is_in_this_week.Not())
                c1_last_term_choices.append(is_in_this_week)

            model.AddBoolOr(c1_last_term_choices).OnlyEnforceIf(is_c1_q4)
            model.AddBoolAnd([choice.Not() for choice in c1_last_term_choices]).OnlyEnforceIf(is_c1_q4.Not())

            # Set up last term detection for course 2
            c2_last_term_choices = []
            for week in last_term_weeks:
                is_in_this_week = model.NewBoolVar(f"{c2}_{run2}_in_week_{week}")
                model.Add(var2 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var2 != week).OnlyEnforceIf(is_in_this_week.Not())
                c2_last_term_choices.append(is_in_this_week)

            model.AddBoolOr(c2_last_term_choices).OnlyEnforceIf(is_c2_q4)
            model.AddBoolAnd([choice.Not() for choice in c2_last_term_choices]).OnlyEnforceIf(is_c2_q4.Not())

            # Either course is in last term
            either_in_last_term = model.NewBoolVar(f"either_{c1}_{c2}_in_last_term")
            model.AddBoolOr([is_c1_q4, is_c2_q4]).OnlyEnforceIf(either_in_last_term)
            model.AddBoolAnd([is_c1_q4.Not(), is_c2_q4.Not()]).OnlyEnforceIf(either_in_last_term.Not())

            # Soft affinity constraint with reduced gap for last term
            too_close = model.NewBoolVar(f"affinity_too_close_{c1}_{c2}_{run1}_{run2}")

            # Regular gap weeks for non-last term
            far_enough_after_normal = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_normal")
            far_enough_before_normal = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_normal")

            # Reduced gap weeks for last term (Sep-Dec)
            reduced_gap = max(1, gap_weeks - 1)  # Reduce by 1 week, minimum 1 week gap
            far_enough_after_last_term = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_last_term")
            far_enough_before_last_term = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_last_term")

            # Normal gap constraints
            model.Add(var2 >= var1 + gap_weeks).OnlyEnforceIf([far_enough_after_normal, either_in_last_term.Not()])
            model.Add(var2 <= var1 - gap_weeks).OnlyEnforceIf([far_enough_before_normal, either_in_last_term.Not()])

            # Reduced gap constraints for last term
            model.Add(var2 >= var1 + reduced_gap).OnlyEnforceIf([far_enough_after_last_term, either_in_last_term])
            model.Add(var2 <= var1 - reduced_gap).OnlyEnforceIf([far_enough_before_last_term, either_in_last_term])

            # Combine normal and last term constraints
            model.AddBoolOr([far_enough_after_normal, far_enough_before_normal]).OnlyEnforceIf([too_close.Not(), either_in_last_term.Not()])
            model.AddBoolOr([far_enough_after_last_term, far_enough_before_last_term]).OnlyEnforceIf([too_close.Not(), either_in_last_term])

            # Add violation constraints
            model.Add(var2 < var1 + reduced_gap).OnlyEnforceIf([too_close, either_in_last_term])
            model.Add(var2 > var1 - reduced_gap).OnlyEnforceIf([too_close, either_in_last_term])
            model.Add(var2 < var1 + gap_weeks).OnlyEnforceIf([too_close, either_in_last_term.Not()])
            model.Add(var2 > var1 - gap_weeks).OnlyEnforceIf([too_close, either_in_last_term.Not()])

            # Add penalty for being too close
            affinity_penalties.extend([too_close] * affinity_weight)

            print(f"  Added affinity constraint: {c1} and {c2} should be {gap_weeks} weeks apart (reduced to {reduced_gap} in last term Sep-Dec)")

        # After trainer workload constraints but before the objective function
        # CONSTRAINT: Champion must teach at least one instance of their course
        print("Adding champion teaching requirements...")

        # Track champion teaching assignments
        champion_course_vars = {}  # Dict to track champion course assignments

        # Collect champion teaching assignments
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            champion = self.course_champions.get((course, language))
            if champion:
                key = (course, language, champion)
                if key not in champion_course_vars:
                    champion_course_vars[key] = []
                
                # Create a boolean variable for this course instance being taught by its champion
                is_champion_teaching = model.NewBoolVar(f"{course}_{language}_{i}_taught_by_champion")
                qualified_trainers = self.fleximatrix.get((course, language), [])
                if champion in qualified_trainers:
                    champion_idx = qualified_trainers.index(champion)
                    model.Add(trainer_var == champion_idx).OnlyEnforceIf(is_champion_teaching)
                    model.Add(trainer_var != champion_idx).OnlyEnforceIf(is_champion_teaching.Not())
                    champion_course_vars[key].append(is_champion_teaching)

        # Enforce champion teaching requirements
        print("Adding champion teaching constraints...")
        for (course, language, champion), teaching_vars in champion_course_vars.items():
            if teaching_vars:  # Only if there are instances of this course
                # Champion must teach at least one instance of their course
                model.Add(sum(teaching_vars) >= 1)
                print(f"Added constraint: Champion {champion} must teach at least one instance of {course} ({language})")

        # Create objective function
        print("Creating objective function...")
        if prioritize_all_courses:
            model.Minimize(
                50 * sum(unscheduled_course_penalties) +
                30 * sum(workload_violation_penalties) +
                20 * sum(spacing_violation_penalties) +
                monthly_weight * sum(month_deviation_penalties) +
                affinity_weight * sum(affinity_penalties)  # Now includes affinity penalties
            )
        else:
            model.Minimize(
                monthly_weight * sum(month_deviation_penalties) +
                below_min_weight * sum(workload_violation_penalties) +
                above_max_weight * sum(workload_violation_penalties) +
                affinity_weight * sum(affinity_penalties)  # Now includes affinity penalties
            )

        # Initialize solver with customized parameters for long runs
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = solver_time_minutes * 60
        solver.parameters.num_search_workers = num_workers
        solver.parameters.log_search_progress = True
        
        # Additional parameters for long runs
        if solver_time_minutes > 60:
            print("Configuring solver for long-running optimization...")
            solver.parameters.cp_model_presolve = False
            solver.parameters.linearization_level = 0
            solver.parameters.random_seed = 42
            solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
        
        # Create and use the timeout callback with all penalties
        solution_callback = TimeoutCallback(
            solver_time_minutes * 60,
            model,
            {
                "monthly": month_deviation_penalties,
                "below_min_days": workload_violation_penalties,
                "above_max_days": workload_violation_penalties,
                "affinity": affinity_penalties
            },
            {
                'monthly': monthly_weight,
                'below_min': below_min_weight,
                'above_max': above_max_weight,
                'affinity': affinity_weight
            }
        )

        # Solve the model
        print("Starting solver...")
        status = solver.Solve(model, solution_callback)

        # Process results and create schedule DataFrame
        schedule_results = []
        unscheduled_courses = []
        
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print(f"Solution found with status: {solver.StatusName(status)}")
            
            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    assigned_week = solver.Value(week_var)
                    if assigned_week > 0:  # Course was scheduled
                        start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")

                        # Get trainer assignment if it exists
                        if (course, delivery_type, language, i) in trainer_assignments:
                            trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                            trainer_idx = solver.Value(trainer_var)

                            # Check if this course has qualified trainers in the fleximatrix
                            if (course, language) in self.fleximatrix and len(self.fleximatrix[(course, language)]) > 0:
                                if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                                    trainer = self.fleximatrix[(course, language)][trainer_idx]
                                    is_champion = "✓" if self.course_champions.get(
                                        (course, language)) == trainer else " "
                                else:
                                    trainer = "Unknown (Index Error)"
                                    is_champion = " "
                            else:
                                trainer = "No Qualified Trainers"
                                is_champion = " "

                        schedule_results.append({
                            "Week": assigned_week,
                            "Start Date": start_date,
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1,
                            "Trainer": trainer,
                            "Champion": is_champion
                        })
                    else:  # Course wasn't scheduled
                        unscheduled_courses.append({
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1
                        })

                except Exception as e:
                    print(f"Error processing result for {course} (run {i + 1}): {e}")
                    unscheduled_courses.append({
                        "Course": course,
                        "Delivery Type": delivery_type,
                        "Language": language,
                        "Run": i + 1,
                        "Error": str(e)
                    })

            # Convert to DataFrame and sort
            schedule_df = pd.DataFrame(schedule_results)
            if not schedule_df.empty:
                schedule_df = schedule_df.sort_values(by=["Week", "Course"])
        else:
            print(f"No solution found. Status: {solver.StatusName(status)}")
            schedule_df = pd.DataFrame()  # Empty DataFrame if no solution found

        return status, schedule_df, solver, schedule, trainer_assignments, unscheduled_courses

    def run_incremental_optimization(self, solver_time_minutes=5):
        """Run optimization by incrementally adding constraints to diagnose issues"""
        print("Starting incremental constraint optimization...")

        # Track feasibility at each step
        diagnostics = []

        # Initialize model and variables
        model = cp_model.CpModel()
        schedule = {}
        trainer_assignments = {}
        max_weeks = len(self.weekly_calendar)

        # Create the schedule variables for each course and run
        print("Step 1: Basic scheduling variables")
        for _, row in self.course_run_data.iterrows():
            course, delivery_type, language, runs, duration = row["Course Name"], row["Delivery Type"], row["Language"], \
                row["Runs"], row["Duration"]

            for i in range(runs):
                # Create a variable for the start week
                start_week = model.NewIntVar(1, max_weeks, f"start_week_{course}_{i}")
                schedule[(course, delivery_type, language, i)] = start_week

                # Create trainer assignment variable
                qualified_trainers = self.fleximatrix.get((course, language), [])
                if not qualified_trainers:
                    print(f"Warning: No qualified trainers for {course} ({language})")
                    continue

                trainer_var = model.NewIntVar(0, len(qualified_trainers) - 1, f"trainer_{course}_{i}")
                trainer_assignments[(course, delivery_type, language, i)] = trainer_var

        # Check feasibility with just variables
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 30  # Short timeout for each step
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Basic variables only",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        # STEP 2: Add minimum spacing between runs of same course
        print("Step 2: Adding course spacing constraints")
        for course_name in set(self.course_run_data["Course Name"]):
            course_runs = []
            for (course, delivery_type, language, i), var in schedule.items():
                if course == course_name:
                    course_runs.append((i, var))

            # Sort by run index
            course_runs.sort(key=lambda x: x[0])

            for i in range(len(course_runs) - 1):
                run_num1, var1 = course_runs[i]
                run_num2, var2 = course_runs[i + 1]

                # Hard constraint for minimum spacing
                model.Add(var2 >= var1 + 2)  # 2-week minimum spacing

        # Check feasibility with spacing constraints
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Course spacing constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible with just course spacing constraints!")
            return "INFEASIBLE", diagnostics, None

        # STEP 3: Add trainer availability constraints
        print("Step 3: Adding trainer availability constraints")
        for (course, delivery_type, language, i), week_var in schedule.items():
            trainer_var = trainer_assignments.get((course, delivery_type, language, i))
            if trainer_var is None:
                continue

            qualified_trainers = self.fleximatrix.get((course, language), [])

            # For each week, set which trainers are available
            for week in range(1, max_weeks + 1):
                is_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                model.Add(week_var == week).OnlyEnforceIf(is_this_week)
                model.Add(week_var != week).OnlyEnforceIf(is_this_week.Not())

                # For each trainer, check availability
                for t_idx, trainer in enumerate(qualified_trainers):
                    if not self.is_trainer_available(trainer, week):
                        # If trainer is not available this week, cannot select this trainer
                        is_this_trainer = model.NewBoolVar(f"{course}_{i}_trainer_{t_idx}_week_{week}")
                        model.Add(trainer_var == t_idx).OnlyEnforceIf(is_this_trainer)
                        model.Add(trainer_var != t_idx).OnlyEnforceIf(is_this_trainer.Not())

                        # Cannot have both is_this_week and is_this_trainer be true
                        model.AddBoolOr([is_this_week.Not(), is_this_trainer.Not()])

        # Check feasibility with trainer availability
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Trainer availability constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding trainer availability constraints!")
            return "INFEASIBLE", diagnostics, None

        # STEP 4: Add week restrictions
        print("Step 4: Adding week restriction constraints")
        for course in self.week_restrictions:
            for (c, delivery_type, language, i), week_var in schedule.items():
                if c != course:
                    continue

                # For each week, check if it's restricted
                for week in range(1, max_weeks + 1):
                    week_info = self.week_position_in_month.get(week, {})
                    skip_week = False

                    if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                        skip_week = True
                    elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                        skip_week = True
                    elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                        skip_week = True
                    elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                        skip_week = True
                    elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                        skip_week = True

                    if skip_week:
                        model.Add(week_var != week)

        # Check feasibility with week restrictions
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Week restriction constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding week restrictions!")
            return "INFEASIBLE", diagnostics, None

        # STEP 5: Add monthly distribution constraints
        print("Step 5: Adding monthly distribution constraints")

        # Get total F2F runs
        total_f2f_runs = sum(self.course_run_data["Runs"])

        # Create adjusted monthly demand dictionary
        adjusted_f2f_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Adjust rounding errors
        total_allocated = sum(adjusted_f2f_demand.values())
        if total_allocated != total_f2f_runs:
            max_month = max(adjusted_f2f_demand.items(), key=lambda x: x[1])[0]
            adjusted_f2f_demand[max_month] += (total_f2f_runs - total_allocated)

        # For each month, track and enforce
        for month in range(1, 13):
            target_demand = adjusted_f2f_demand.get(month, 0)
            if target_demand == 0:
                continue

            courses_in_month = []
            # Get all weeks that belong to this month
            month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]

            if not month_weeks:
                continue

            # For each course run, create a Boolean indicating if it's in this month
            for (course, delivery_type, language, i), week_var in schedule.items():
                is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")

                # Add constraints: is_in_month is True iff week_var is in month's weeks
                week_choices = []
                for week in month_weeks:
                    is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                    model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                    model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                    week_choices.append(is_in_this_week)

                model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())

                courses_in_month.append(is_in_month)

            # Enforce the target demand for this month
            if courses_in_month and target_demand > 0:
                model.Add(sum(courses_in_month) == target_demand)

        # Check feasibility with monthly distribution
        status = solver.Solve(model)
        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Monthly distribution constraints",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        if not feasible:
            print("  WARNING: Model infeasible after adding monthly distribution!")
            return "INFEASIBLE", diagnostics, None

        # STEP 6: Final full-featured model with all constraints and objective
        print("Step 6: Building full model with objective function")

        # Add penalties for objective function
        affinity_penalties = []
        trainer_utilization_penalties = []
        unscheduled_course_penalties = []
        workload_violation_penalties = []
        spacing_violation_penalties = []
        month_deviation_penalties = []
        priority_penalties = []  # For priority-based penalties

        # Add affinity constraints (soft)
        for _, row in self.affinity_matrix_data.iterrows():
            c1, c2, gap_weeks = row["Course 1"], row["Course 2"], row["Gap Weeks"]

            c1_runs = []
            c2_runs = []

            for (course, _, _, i), var in schedule.items():
                if course == c1:
                    c1_runs.append((i, var))
                elif course == c2:
                    c2_runs.append((i, var))

            if not c1_runs or not c2_runs:
                continue

            # Sort by run index
            c1_runs.sort(key=lambda x: x[0])
            c2_runs.sort(key=lambda x: x[0])

            # Only check first run of each course
            run1, var1 = c1_runs[0]
            run2, var2 = c2_runs[0]

            # Create variables to check if either course is in Q4
            is_c1_q4 = model.NewBoolVar(f"{c1}_{run1}_in_last_term")
            is_c2_q4 = model.NewBoolVar(f"{c2}_{run2}_in_last_term")

            # Get last term weeks (Sep-Dec, months 9-12)
            last_term_weeks = [w for w, m in self.week_to_month_map.items() if m in [9, 10, 11, 12]]

            # Set up last term detection for course 1
            c1_last_term_choices = []
            for week in last_term_weeks:
                is_in_this_week = model.NewBoolVar(f"{c1}_{run1}_in_week_{week}")
                model.Add(var1 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var1 != week).OnlyEnforceIf(is_in_this_week.Not())
                c1_last_term_choices.append(is_in_this_week)

            model.AddBoolOr(c1_last_term_choices).OnlyEnforceIf(is_c1_q4)
            model.AddBoolAnd([choice.Not() for choice in c1_last_term_choices]).OnlyEnforceIf(is_c1_q4.Not())

            # Set up last term detection for course 2
            c2_last_term_choices = []
            for week in last_term_weeks:
                is_in_this_week = model.NewBoolVar(f"{c2}_{run2}_in_week_{week}")
                model.Add(var2 == week).OnlyEnforceIf(is_in_this_week)
                model.Add(var2 != week).OnlyEnforceIf(is_in_this_week.Not())
                c2_last_term_choices.append(is_in_this_week)

            model.AddBoolOr(c2_last_term_choices).OnlyEnforceIf(is_c2_q4)
            model.AddBoolAnd([choice.Not() for choice in c2_last_term_choices]).OnlyEnforceIf(is_c2_q4.Not())

            # Either course is in last term
            either_in_last_term = model.NewBoolVar(f"either_{c1}_{c2}_in_last_term")
            model.AddBoolOr([is_c1_q4, is_c2_q4]).OnlyEnforceIf(either_in_last_term)
            model.AddBoolAnd([is_c1_q4.Not(), is_c2_q4.Not()]).OnlyEnforceIf(either_in_last_term.Not())

            # Soft affinity constraint with reduced gap for last term
            too_close = model.NewBoolVar(f"affinity_too_close_{c1}_{c2}_{run1}_{run2}")

            # Regular gap weeks for non-last term
            far_enough_after_normal = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_normal")
            far_enough_before_normal = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_normal")

            # Reduced gap weeks for last term (Sep-Dec)
            reduced_gap = max(1, gap_weeks - 1)  # Reduce by 1 week, minimum 1 week gap
            far_enough_after_last_term = model.NewBoolVar(f"far_after_{c1}_{c2}_{run1}_{run2}_last_term")
            far_enough_before_last_term = model.NewBoolVar(f"far_before_{c1}_{c2}_{run1}_{run2}_last_term")

            # Normal gap constraints
            model.Add(var2 >= var1 + gap_weeks).OnlyEnforceIf([far_enough_after_normal, either_in_last_term.Not()])
            model.Add(var2 <= var1 - gap_weeks).OnlyEnforceIf([far_enough_before_normal, either_in_last_term.Not()])

            # Reduced gap constraints for last term
            model.Add(var2 >= var1 + reduced_gap).OnlyEnforceIf([far_enough_after_last_term, either_in_last_term])
            model.Add(var2 <= var1 - reduced_gap).OnlyEnforceIf([far_enough_before_last_term, either_in_last_term])

            # Combine normal and last term constraints
            model.AddBoolOr([far_enough_after_normal, far_enough_before_normal]).OnlyEnforceIf([too_close.Not(), either_in_last_term.Not()])
            model.AddBoolOr([far_enough_after_last_term, far_enough_before_last_term]).OnlyEnforceIf([too_close.Not(), either_in_last_term])

            # Add violation constraints
            model.Add(var2 < var1 + reduced_gap).OnlyEnforceIf([too_close, either_in_last_term])
            model.Add(var2 > var1 - reduced_gap).OnlyEnforceIf([too_close, either_in_last_term])
            model.Add(var2 < var1 + gap_weeks).OnlyEnforceIf([too_close, either_in_last_term.Not()])
            model.Add(var2 > var1 - gap_weeks).OnlyEnforceIf([too_close, either_in_last_term.Not()])

            affinity_penalties.append(too_close)
            print(f"  Added affinity constraint: {c1} and {c2} should be {gap_weeks} weeks apart (reduced to {reduced_gap} in last term Sep-Dec)")

        # Add trainer utilization (soft)
        trainer_workload = {name: [] for name in self.consultant_data["Name"]}

        # Track course assignments for each trainer
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            qualified_trainers = self.fleximatrix.get((course, language), [])
            course_champion = self.course_champions.get((course, language))

            for t_idx, trainer in enumerate(qualified_trainers):
                is_assigned = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}")
                model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned)
                model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned.Not())

                # Accumulate workload
                trainer_workload[trainer].append((is_assigned, duration))

                # Get trainer's priority - if they're the champion, use priority 1, otherwise use their title's priority
                if trainer == course_champion:
                    priority = 1
                else:
                    title = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Title"].iloc[0]
                    priority = self.priority_data.loc[self.priority_data["Title"] == title, "Priority"].iloc[0]

                # Add weighted penalty based on priority (multiply by priority to make higher numbers more expensive)
                for _ in range(priority):
                    priority_penalties.append(is_assigned)

        # Calculate and constrain workload
        for trainer, workload_items in trainer_workload.items():
            if not workload_items:
                continue

            max_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Max_Days"].iloc[0]

            # Calculate total workload
            total_workload = model.NewIntVar(0, max_days * 2, f"total_workload_{trainer}")  # Allow exceeding max but penalize
            weighted_sum = []
            for is_assigned, duration in workload_items:
                weighted_sum.append((is_assigned, duration))

            if weighted_sum:
                weighted_terms = []
                for is_assigned, duration in weighted_sum:
                    term = model.NewIntVar(0, duration, f"term_{id(is_assigned)}_{duration}")
                    model.Add(term == duration).OnlyEnforceIf(is_assigned)
                    model.Add(term == 0).OnlyEnforceIf(is_assigned.Not())
                    weighted_terms.append(term)
                model.Add(total_workload == sum(weighted_terms))
                
                # Convert hard workload constraints to soft constraints
                over_max = model.NewBoolVar(f"{trainer}_over_max")
                model.Add(total_workload > max_days).OnlyEnforceIf(over_max)
                model.Add(total_workload <= max_days).OnlyEnforceIf(over_max.Not())
                workload_violation_penalties.append(over_max)

                # Add minimum workload constraint using Min_Days from Excel
                min_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Min_Days"].iloc[0]
                if min_days > 0:
                    under_min = model.NewBoolVar(f"{trainer}_under_min")
                    model.Add(total_workload < min_days).OnlyEnforceIf(under_min)
                    model.Add(total_workload >= min_days).OnlyEnforceIf(under_min.Not())
                    workload_violation_penalties.append(under_min)

        # Collect all penalties in a dictionary
        penalties = {
            "affinity": affinity_penalties,
            "utilization": trainer_utilization_penalties,
            "unscheduled": unscheduled_course_penalties,
            "workload": workload_violation_penalties,
            "spacing": spacing_violation_penalties,
            "monthly": month_deviation_penalties,
            "priority": priority_penalties
        }

        # Create weights dictionary from parameters
        weights = {
            "regular": {
                "affinity": affinity_weight,
                "monthly": monthly_weight
            },
            "prioritize_all": {
                "unscheduled": 50,
                "workload": 30,
                "spacing": 20,
                "monthly": monthly_weight
            }
        }

        # Choose optimization mode
        mode = "prioritize_all" if prioritize_all_courses else "regular"

        # Create and add objective function
        self._create_objective_function(model, penalties, mode, weights[mode])

        # Solve with final model
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = solver_time_minutes * 60
        status = solver.Solve(model)

        feasible = status in [cp_model.OPTIMAL, cp_model.FEASIBLE]
        diagnostics.append({
            "step": "Full model with objective",
            "feasible": feasible,
            "status": solver.StatusName(status)
        })
        print(f"  Status: {solver.StatusName(status)}")

        # Create and return results if feasible
        if feasible:
            schedule_results = []
            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    assigned_week = solver.Value(week_var)
                    start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")

                    # Get trainer assignment if it exists
                    if (course, delivery_type, language, i) in trainer_assignments:
                        trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                        trainer_idx = solver.Value(trainer_var)

                        # Check if this course has qualified trainers in the fleximatrix
                        if (course, language) in self.fleximatrix and len(self.fleximatrix[(course, language)]) > 0:
                            if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                                trainer = self.fleximatrix[(course, language)][trainer_idx]
                                is_champion = "✓" if self.course_champions.get(
                                    (course, language)) == trainer else " "
                            else:
                                # Handle out of range index
                                trainer = "Unknown (Index Error)"
                                is_champion = " "
                        else:
                            # Handle missing course in fleximatrix
                            trainer = "No Qualified Trainers"
                            is_champion = " "

                        schedule_results.append({
                            "Week": assigned_week,
                            "Start Date": start_date,
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1,
                            "Trainer": trainer,
                            "Champion": is_champion
                        })

                except Exception as e:
                    print(f"Error processing result for {course} (run {i+1}): {e}")
                    # Add to unscheduled due to error
                    unscheduled_courses.append({
                        "Course": course,
                        "Delivery Type": delivery_type,
                        "Language": language,
                        "Run": i + 1,
                        "Error": str(e)
                    })

            # Convert to DataFrame
            schedule_df = pd.DataFrame(schedule_results)

            # Sort by week first, then by course name
            schedule_df = schedule_df.sort_values(by=["Week", "Course"])

            return "FEASIBLE", diagnostics, schedule_df, unscheduled_courses
        else:
            return "INFEASIBLE", diagnostics, None, []

    def generate_trainer_workload_report(self, schedule, trainer_assignments, solver):
        """Generates a report on trainer workload and min/max days compliance"""
        # Calculate days assigned to each trainer
        trainer_days = {name: 0 for name in self.consultant_data["Name"]}
        trainer_courses = {name: 0 for name in self.consultant_data["Name"]}
        champion_courses = {name: 0 for name in self.consultant_data["Name"]}

        for (course, delivery_type, language, i), week_var in schedule.items():
            try:
                duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]

                # Get trainer assignment
                if (course, delivery_type, language, i) not in trainer_assignments:
                    continue
                
                trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                trainer_idx = solver.Value(trainer_var)

                # Check if this course has an entry in the fleximatrix
                if (course, language) not in self.fleximatrix:
                    continue

                if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                    trainer = self.fleximatrix[(course, language)][trainer_idx]
                    trainer_days[trainer] += duration
                    trainer_courses[trainer] += 1

                    # Check if this is a champion course
                    if self.course_champions.get((course, language)) == trainer:
                        champion_courses[trainer] += 1
            except Exception as e:
                print(f"Error in workload report for {course} (run {i+1}): {e}")
                continue

        # Create trainer workload report dataframe
        workload_data = []

        for _, row in self.consultant_data.iterrows():
            name = row["Name"]
            title = row["Title"]
            min_days = row["Min_Days"]
            max_days = row["Max_Days"]
            assigned_days = trainer_days[name]
            
            # Calculate compliance with min/max days
            below_min = assigned_days < min_days if min_days > 0 else False
            above_max = assigned_days > max_days if max_days > 0 else False
            
            workload_data.append({
                "Name": name,
                "Title": title,
                "Min Days": min_days,
                "Max Days": max_days,
                "Assigned Days": assigned_days,
                "Below Min": "❌" if below_min else "✓",
                "Above Max": "❌" if above_max else "✓",
                "Courses": trainer_courses[name],
                "Champion Courses": champion_courses[name]
            })

        # Add total row
        total_days = sum(trainer_days.values())
        total_min = self.consultant_data["Min_Days"].sum()
        total_max = self.consultant_data["Max_Days"].sum()

        workload_data.append({
            "Name": "TOTAL",
            "Title": "",
            "Min Days": total_min,
            "Max Days": total_max,
            "Assigned Days": total_days,
            "Below Min": "",
            "Above Max": "",
            "Courses": sum(trainer_courses.values()),
            "Champion Courses": sum(champion_courses.values())
        })

        return pd.DataFrame(workload_data)

    def plot_weekly_course_bar_chart(self, schedule, solver):
        """Creates a bar chart showing number of courses per week."""
        try:
            max_weeks = len(self.weekly_calendar)

            # Print debugging information
            print(f"DEBUG: max_weeks = {max_weeks}")
            print(f"DEBUG: schedule length = {len(schedule)}")

            # Count courses per week
            weekly_counts = {week: 0 for week in range(1, max_weeks + 1)}

            for (course, _, _, _), week_var in schedule.items():
                try:
                    assigned_week = solver.Value(week_var)
                    if 1 <= assigned_week <= max_weeks:  # Add range check
                        weekly_counts[assigned_week] += 1
                    else:
                        print(f"DEBUG: Week value {assigned_week} is out of range (1-{max_weeks})")
                except Exception as e:
                    print(f"DEBUG: Error getting week value for {course}: {e}")
                    continue

            # Convert to dataframe for plotting
            df = pd.DataFrame({
                'Week': list(weekly_counts.keys()),
                'Courses': list(weekly_counts.values())
            })

            # Create a bar chart
            plt.figure(figsize=(12, 6))
            bars = plt.bar(df['Week'], df['Courses'], width=0.8)

            # Add week dates as second x-axis
            ax1 = plt.gca()
            ax2 = ax1.twiny()

            # Select a subset of week numbers for readability
            step = max(1, max_weeks // 12)  # Show about 12 date labels
            selected_weeks = list(range(1, max_weeks + 1, step))

            # Get date strings for selected weeks
            date_labels = []
            for w in selected_weeks:
                if 0 <= w - 1 < len(self.weekly_calendar):  # Add range check
                    date_labels.append(self.weekly_calendar[w - 1].strftime("%b %d"))
                else:
                    date_labels.append(f"Week {w}")

            # Set positions and labels
            ax2.set_xticks([w for w in selected_weeks])
            ax2.set_xticklabels(date_labels, rotation=45)

            # Formatting
            plt.title("Number of Courses per Week", fontsize=16)
            ax1.set_xlabel("Week Number", fontsize=14)
            ax1.set_ylabel("Number of Courses", fontsize=14)
            ax2.set_xlabel("Week Start Date", fontsize=14)

            # Set y limit to make small differences more visible
            max_courses = max(weekly_counts.values()) if weekly_counts else 0
            ax1.set_ylim(0, max(4, max_courses + 1))

            # Highlight weeks with no courses
            for week, count in weekly_counts.items():
                if count == 0:
                    plt.axvspan(week - 0.4, week + 0.4, color='lightgray', alpha=0.3)

            # Add grid for readability
            plt.grid(axis='y', linestyle='--', alpha=0.7)

            plt.tight_layout()

            # Return the figure
            return plt.gcf()
        except Exception as e:
            print(f"DEBUG: Error in plot_weekly_course_bar_chart: {e}")
            # Return a simple error figure
            plt.figure(figsize=(8, 6))
            plt.text(0.5, 0.5, f"Error generating chart:\n{str(e)}",
                     horizontalalignment='center', verticalalignment='center')
            return plt.gcf()

    def plot_trainer_workload_chart(self, schedule, trainer_assignments, solver):
        """Creates a bar chart showing trainer workload and utilization"""
        try:
            # Calculate days assigned to each trainer
            trainer_days = {name: 0 for name in self.consultant_data["Name"]}

            for (course, delivery_type, language, i), week_var in schedule.items():
                try:
                    # Get the course duration
                    duration_series = self.course_run_data.loc[
                        self.course_run_data["Course Name"] == course, "Duration"]
                    if len(duration_series) == 0:
                        print(f"DEBUG: Course {course} not found in course_run_data")
                        continue

                    duration = duration_series.iloc[0]

                    # Check if this course+run has a trainer assignment
                    if (course, delivery_type, language, i) not in trainer_assignments:
                        print(f"DEBUG: No trainer assignment for {course} run {i + 1}")
                        continue

                    # Get trainer assignment
                    trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                    trainer_idx = solver.Value(trainer_var)

                    # Check if the course+language has qualified trainers
                    if (course, language) not in self.fleximatrix:
                        print(f"DEBUG: No fleximatrix entry for {course} ({language})")
                        continue

                    qualified_trainers = self.fleximatrix[(course, language)]

                    # Check if trainer index is valid
                    if 0 <= trainer_idx < len(qualified_trainers):
                        trainer = qualified_trainers[trainer_idx]
                        trainer_days[trainer] += duration
                    else:
                        print(
                            f"DEBUG: Invalid trainer index {trainer_idx} for {course} (max {len(qualified_trainers) - 1})")
                except Exception as e:
                    print(f"DEBUG: Error processing course {course} run {i + 1}: {e}")
                    continue

            # Prepare data for plotting - only include trainers with assignments
            trainers = []
            assigned_days = []
            max_days_list = []
            colors = []

            for name, days in sorted(trainer_days.items(), key=lambda x: x[1], reverse=True):
                if days > 0:  # Only include trainers with assignments
                    trainers.append(name)
                    assigned_days.append(days)

                    # Get max days from consultant data
                    max_days_series = self.consultant_data.loc[self.consultant_data["Name"] == name, "Max_Days"]

                    if len(max_days_series) > 0:
                        max_days = max_days_series.iloc[0]
                        max_days_list.append(max_days)

                        # Get title for color
                        title_series = self.consultant_data.loc[self.consultant_data["Name"] == name, "Title"]
                        title = title_series.iloc[0] if len(title_series) > 0 else "Unknown"
                        colors.append("tab:orange" if title == "Freelancer" else "tab:blue")
                    else:
                        print(f"DEBUG: Trainer {name} not found in consultant_data")
                        max_days_list.append(100)  # Default value
                        colors.append("tab:gray")  # Default color

            # Check if we have data to plot
            if not trainers:
                plt.figure(figsize=(10, 6))
                plt.text(0.5, 0.5, "No trainer assignments to display",
                         horizontalalignment='center', verticalalignment='center', fontsize=14)
                return plt.gcf()

            # Create figure and axis
            fig, ax = plt.subplots(figsize=(12, 6))

            # Plot assigned days
            bars = ax.bar(trainers, assigned_days, label="Assigned Days", color=colors)

            # Plot max days as a line
            ax.plot(trainers, max_days_list, 'rx-', label="Maximum Days", linewidth=2)

            # Add percentage labels
            for i, (bar, max_days) in enumerate(zip(bars, max_days_list)):
                height = bar.get_height()
                percentage = (height / max_days * 100) if max_days > 0 else 0
                ax.text(bar.get_x() + bar.get_width() / 2., height + 5,
                        f"{percentage:.1f}%", ha='center', va='bottom', fontsize=9)

            # Improve readability
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()

            # Add labels and title
            plt.ylabel("Training Days")
            plt.title("Trainer Workload and Utilization")
            plt.legend()
            plt.grid(axis='y', linestyle='--', alpha=0.7)

            # Return the figure
            return plt.gcf()
        except Exception as e:
            print(f"DEBUG: Error in plot_trainer_workload_chart: {e}")
            # Return a simple error figure
            plt.figure(figsize=(8, 6))
            plt.text(0.5, 0.5, f"Error generating chart:\n{str(e)}",
                     horizontalalignment='center', verticalalignment='center')
            return plt.gcf()

    def generate_monthly_validation(self, schedule, solver):
        """Generates a monthly validation report for the schedule"""
        # Get adjusted F2F demand from the monthly demand dataframe
        adjusted_f2f_demand = {}

        # Create the dictionary using the same logic as in run_optimization
        total_f2f_runs = sum(self.course_run_data["Runs"])
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Count courses by month
        monthly_counts = {month: 0 for month in range(1, 13)}

        for (course, delivery_type, language, i), week_var in schedule.items():
            assigned_week = solver.Value(week_var)
            if assigned_week == 0:
                continue  # Skip unscheduled runs
            month = self.week_to_month_map[assigned_week]
            monthly_counts[month] += 1

        # Create validation dataframe
        validation_data = []

        total_target = 0
        total_actual = 0
        all_match = True

        for month in range(1, 13):
            target = adjusted_f2f_demand.get(month, 0)
            actual = monthly_counts[month]
            diff = actual - target

            status = "MATCH" if diff == 0 else "OVER" if diff > 0 else "UNDER"
            if diff != 0:
                all_match = False

            validation_data.append({
                "Month": str(month),  # Convert month to string
                "Target": target,
                "Actual": actual,
                "Difference": diff,
                "Status": status
            })

            total_target += target
            total_actual += actual

        # Add total row
        validation_data.append({
            "Month": "TOTAL",
            "Target": total_target,
            "Actual": total_actual,
            "Difference": total_actual - total_target,
            "Status": "MATCH" if all_match else "MISMATCH"
        })

        # Convert to DataFrame and ensure proper column types
        df = pd.DataFrame(validation_data)
        df['Target'] = df['Target'].astype(int)
        df['Actual'] = df['Actual'].astype(int)
        df['Difference'] = df['Difference'].astype(int)
        
        return df

    def generate_excel_report(self, schedule_df, monthly_validation_df, trainer_utilization_df):
        """Generate an Excel report with multiple sheets containing all results"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if schedule_df is not None:
                schedule_df.to_excel(writer, sheet_name='Schedule', index=False)
            else:
                # Create an empty DataFrame if None
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Schedule', index=False)

            if monthly_validation_df is not None:
                monthly_validation_df.to_excel(writer, sheet_name='Monthly Validation', index=False)
            else:
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Monthly Validation',
                                                                     index=False)

            if trainer_utilization_df is not None:
                trainer_utilization_df.to_excel(writer, sheet_name='Trainer Utilization', index=False)
            else:
                pd.DataFrame(columns=["No data available"]).to_excel(writer, sheet_name='Trainer Utilization',
                                                                     index=False)
            
            # Create the new Calendar View sheet
            if schedule_df is not None and not schedule_df.empty:
                # Get all trainers from the consultant data
                trainers = self.consultant_data["Name"].tolist()
                
                # Create a new dataframe with trainers as rows
                calendar_data = []
                for trainer in trainers:
                    calendar_data.append({"Trainer": trainer})
                
                calendar_df = pd.DataFrame(calendar_data)
                
                # Add the dataframe to the Excel but we'll manually format it later
                calendar_df.to_excel(writer, sheet_name='Calendar View', index=False)
                
                # Get the worksheet and workbook objects
                worksheet = writer.sheets['Calendar View']
                workbook = writer.book
                
                # Define fill colors
                light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Assigned
                dark_green_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")   # Champion
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")         # Vacation
                purple_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")      # Holiday
                
                # Create header row with week start dates
                for col_idx, week_start_date in enumerate(self.weekly_calendar, start=2):
                    # Format date as DD-MMM
                    formatted_date = week_start_date.strftime("%d-%b")
                    week_num = col_idx - 1  # Adjusted for Excel column indexing
                    
                    # Write the date to the header
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = formatted_date
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    
                    # Format column width
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 10
                
                # Create dictionary to track trainer assignments by week
                trainer_assignments_by_week = {}
                champion_assignments_by_week = {}
                course_assignments_by_week = {}  # New dictionary to track course names and languages
                
                # Process schedule data to get assignments
                for _, row in schedule_df.iterrows():
                    week = row['Week']
                    trainer = row['Trainer']
                    course = row['Course']  # Get course name
                    language = row['Language']  # Get language
                    is_champion = row['Champion'].strip() == "✓"
                    
                    # Add language suffix
                    course_with_lang = f"{course} ({language[0]})"  # Use first letter of language (E/A)
                    
                    if (trainer, week) not in trainer_assignments_by_week:
                        trainer_assignments_by_week[(trainer, week)] = 0
                        course_assignments_by_week[(trainer, week)] = course_with_lang  # Store course name with language
                    
                    trainer_assignments_by_week[(trainer, week)] += 1
                    
                    if is_champion:
                        if (trainer, week) not in champion_assignments_by_week:
                            champion_assignments_by_week[(trainer, week)] = 0
                        
                        champion_assignments_by_week[(trainer, week)] += 1
                
                # Increase column width for course names
                for col_idx in range(2, len(self.weekly_calendar) + 2):
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 20
                
                # Process all trainers and weeks
                for row_idx, trainer in enumerate(trainers, start=2):
                    for col_idx, week_start_date in enumerate(self.weekly_calendar, start=2):
                        week_num = col_idx - 1  # Adjusted for Excel column indexing
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Check for public holidays
                        has_public_holiday = False
                        for _, holiday in self.public_holidays_data.iterrows():
                            week_end_date = week_start_date + datetime.timedelta(days=4)  # 5-day work week
                            if (holiday["Start Date"] <= week_end_date and holiday["End Date"] >= week_start_date):
                                has_public_holiday = True
                                break
                        
                        # Check for trainer vacation
                        is_on_vacation = False
                        for _, leave in self.annual_leaves[self.annual_leaves["Name"] == trainer].iterrows():
                            week_end_date = week_start_date + datetime.timedelta(days=4)  # 5-day work week
                            if (leave["Start_Date"] <= week_end_date and leave["End_Date"] >= week_start_date):
                                is_on_vacation = True
                                break
                        
                        # Check for trainer assignment
                        is_assigned = (trainer, week_num) in trainer_assignments_by_week
                        is_champion_assignment = (trainer, week_num) in champion_assignments_by_week
                        
                        # Apply appropriate fill and set value
                        if is_on_vacation:
                            cell.fill = red_fill
                            cell.value = "Vacation"
                        elif has_public_holiday:
                            cell.fill = purple_fill
                            cell.value = "Holiday"
                        elif is_champion_assignment:
                            cell.fill = dark_green_fill
                            cell.value = course_assignments_by_week.get((trainer, week_num), "")
                            cell.font = Font(color="FFFFFF")  # White text for dark background
                        elif is_assigned:
                            cell.fill = light_green_fill
                            cell.value = course_assignments_by_week.get((trainer, week_num), "")
                        else:
                            cell.value = ""
                
                # Add a legend
                legend_row = len(trainers) + 4
                
                worksheet.cell(row=legend_row, column=1).value = "Legend:"
                worksheet.cell(row=legend_row, column=1).font = Font(bold=True)
                
                # Trainer assigned
                worksheet.cell(row=legend_row + 1, column=1).value = "Light Green:"
                worksheet.cell(row=legend_row + 1, column=2).value = "Trainer Assigned"
                worksheet.cell(row=legend_row + 1, column=2).fill = light_green_fill
                
                # Champion assigned
                worksheet.cell(row=legend_row + 2, column=1).value = "Dark Green:"
                worksheet.cell(row=legend_row + 2, column=2).value = "Champion Assigned"
                worksheet.cell(row=legend_row + 2, column=2).fill = dark_green_fill
                worksheet.cell(row=legend_row + 2, column=2).font = Font(color="FFFFFF")
                
                # Vacation
                worksheet.cell(row=legend_row + 3, column=1).value = "Red:"
                worksheet.cell(row=legend_row + 3, column=2).value = "Trainer on Vacation"
                worksheet.cell(row=legend_row + 3, column=2).fill = red_fill
                
                # Public holiday
                worksheet.cell(row=legend_row + 4, column=1).value = "Purple:"
                worksheet.cell(row=legend_row + 4, column=2).value = "Public Holiday"
                worksheet.cell(row=legend_row + 4, column=2).fill = purple_fill
                
                # Adjust column width for trainer names
                worksheet.column_dimensions['A'].width = 25
            else:
                # Create an empty calendar view if no schedule data
                pd.DataFrame(columns=["No schedule data available"]).to_excel(writer, sheet_name='Calendar View', index=False)
        
        # Return the output buffer contents
        output.seek(0)
        return output.getvalue()


    def analyze_constraints_visually(self):
        """Generate visual analysis of constraints to help identify sources of infeasibility"""
        import matplotlib.pyplot as plt
        import seaborn as sns
        import numpy as np
        from matplotlib.colors import LinearSegmentedColormap

        # 1. Analyze monthly distribution vs. available weeks
        months = range(1, 13)
        total_f2f_runs = sum(self.course_run_data["Runs"])

        # Calculate monthly demand
        monthly_demand = {}
        total_percentage = self.monthly_demand['Percentage'].sum()
        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            monthly_demand[month] = demand

        # Calculate available weeks per month
        monthly_weeks = {}
        for week, month in self.week_to_month_map.items():
            if month not in monthly_weeks:
                monthly_weeks[month] = 0
            monthly_weeks[month] += 1

        # Calculate available working days per month
        monthly_working_days = {}
        for week, month in self.week_to_month_map.items():
            if month not in monthly_working_days:
                monthly_working_days[month] = 0
            monthly_working_days[month] += self.weekly_working_days.get(week, 0)

        # Create dataframe for plotting
        monthly_data = []
        for month in range(1, 13):
            monthly_data.append({
                'Month': month,
                'Demand': monthly_demand.get(month, 0),
                'Weeks': monthly_weeks.get(month, 0),
                'Working Days': monthly_working_days.get(month, 0)
            })

        monthly_df = pd.DataFrame(monthly_data)

        # Calculate two metrics:
        # 1. Courses per week: How many courses need to be scheduled per available week
        # 2. Demand pressure: Percentage of working days required for courses
        monthly_df['Courses per Week'] = monthly_df['Demand'] / monthly_df['Weeks'].replace(0, np.nan)

        # Assuming each course takes 5 days on average
        avg_course_days = self.course_run_data["Duration"].mean()
        monthly_df['Demand Pressure (%)'] = (monthly_df['Demand'] * avg_course_days * 100) / monthly_df[
            'Working Days'].replace(0, np.nan)

        # Create figures
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))

        # Plot 1: Monthly Demand vs. Available Weeks
        ax1 = axes[0, 0]
        monthly_df.plot(x='Month', y=['Demand', 'Weeks'], kind='bar', ax=ax1)
        ax1.set_title('Monthly Course Demand vs. Available Weeks')
        ax1.set_xlabel('Month')
        ax1.set_ylabel('Count')
        ax1.axhline(y=monthly_df['Demand'].mean(), color='r', linestyle='--', label='Avg. Demand')
        ax1.legend()

        # Plot 2: Courses per Week
        ax2 = axes[0, 1]
        monthly_df.plot(x='Month', y='Courses per Week', kind='bar', ax=ax2, color='orange')
        ax2.set_title('Courses per Available Week')
        ax2.set_xlabel('Month')
        ax2.set_ylabel('Courses per Week')
        ax2.axhline(y=1, color='g', linestyle='--', label='1 Course/Week')
        ax2.axhline(y=2, color='y', linestyle='--', label='2 Courses/Week')
        ax2.axhline(y=3, color='r', linestyle='--', label='3 Courses/Week')
        ax2.legend()

        # Plot 3: Demand Pressure
        ax3 = axes[1, 0]
        bars = monthly_df.plot(x='Month', y='Demand Pressure (%)', kind='bar', ax=ax3, color='purple')
        ax3.set_title('Monthly Demand Pressure (% of Working Days Required)')
        ax3.set_xlabel('Month')
        ax3.set_ylabel('Percentage of Working Days')
        ax3.axhline(y=100, color='r', linestyle='--', label='100% (Full Capacity)')
        ax3.axhline(y=80, color='y', linestyle='--', label='80% (Threshold)')

        # Color code bars by severity
        for i, p in enumerate(bars.patches):
            pressure = monthly_df['Demand Pressure (%)'].iloc[i]
            if np.isnan(pressure):
                p.set_color('gray')
            elif pressure > 100:
                p.set_color('red')
            elif pressure > 80:
                p.set_color('orange')
            elif pressure > 60:
                p.set_color('yellow')
            else:
                p.set_color('green')

        ax3.legend()

        # Plot 4: Week Restrictions Analysis
        ax4 = axes[1, 1]

        # Count how many courses are restricted in each week position
        restrictions = {'First': 0, 'Second': 0, 'Third': 0, 'Fourth': 0, 'Last': 0}

        for course, course_restrictions in self.week_restrictions.items():
            for position, is_restricted in course_restrictions.items():
                if is_restricted:
                    restrictions[position] += 1

        # Calculate percentage of courses restricted
        total_courses = len(set(self.course_run_data["Course Name"]))
        restriction_percentages = {pos: (count * 100 / total_courses) if total_courses > 0 else 0
                                   for pos, count in restrictions.items()}

        # Create restriction dataframe
        restriction_df = pd.DataFrame({
            'Week Position': list(restriction_percentages.keys()),
            'Percentage Restricted': list(restriction_percentages.values())
        })

        # Plot restriction percentages
        restriction_df.plot(x='Week Position', y='Percentage Restricted', kind='bar', ax=ax4, color='teal')
        ax4.set_title('Percentage of Courses with Week Restrictions')
        ax4.set_xlabel('Week Position in Month')
        ax4.set_ylabel('Percentage of Courses')
        ax4.axhline(y=50, color='y', linestyle='--', label='50% Threshold')
        ax4.legend()

        # Adjust layout
        plt.tight_layout()

        # Create another figure for the heatmap of week restrictions
        plt.figure(figsize=(12, 10))

        # Create a matrix of week restrictions for heatmap
        courses = list(set(self.course_run_data["Course Name"]))
        weeks = list(range(1, len(self.weekly_calendar) + 1))

        # Initialize matrix with 1s (allowed)
        heatmap_data = np.ones((len(courses), len(weeks)))

        # Fill in restricted weeks with 0s
        # Modify this part too:
        # Fill in restricted weeks with 0s
        for i, course in enumerate(courses):
            for j, week in enumerate(weeks):
                # Make sure we don't exceed the valid week range
                week_num = week + 1
                if week_num <= len(self.weekly_calendar):  # Add this check
                    # Check if this week is restricted due to working days
                    duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[
                        0]
                    if self.weekly_working_days.get(week_num, 0) < duration:
                        heatmap_data[i, j] = 0

                    # Check week position restrictions
                    if course in self.week_restrictions:
                        week_info = self.week_position_in_month.get(week_num, {})

                        if week_info.get('is_first') and self.week_restrictions[course].get('First', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_second') and self.week_restrictions[course].get('Second', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_third') and self.week_restrictions[course].get('Third', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_fourth') and self.week_restrictions[course].get('Fourth', False):
                            heatmap_data[i, j] = 0
                        elif week_info.get('is_last') and self.week_restrictions[course].get('Last', False):
                            heatmap_data[i, j] = 0

        # Create heatmap
        course_week_heatmap = plt.figure(figsize=(18, 8))
        sns.heatmap(heatmap_data, cmap=['red', 'green'], cbar=False,
                    xticklabels=[f"W{w + 1}" for w in range(len(weeks))],
                    yticklabels=courses)
        plt.title('Course-Week Availability Matrix (Green = Available, Red = Restricted)')
        plt.xlabel('Week Number')
        plt.ylabel('Course')
        plt.tight_layout()

        # Analysis of trainer qualifications
        plt.figure(figsize=(10, 8))

        # Count how many qualified trainers per course/language
        course_lang_trainers = {}
        for (course, language), trainers in self.fleximatrix.items():
            course_lang_trainers[(course, language)] = len(trainers)

        # Convert to DataFrame
        trainers_df = pd.DataFrame([
            {'Course': course, 'Language': lang, 'Qualified Trainers': count}
            for (course, lang), count in course_lang_trainers.items()
        ])

        # Get runs per course/language
        runs_df = self.course_run_data.groupby(['Course Name', 'Language'])['Runs'].sum().reset_index()
        runs_df.columns = ['Course', 'Language', 'Runs']

        # Merge trainers and runs
        analysis_df = pd.merge(trainers_df, runs_df, on=['Course', 'Language'])

        # Add trainers per run ratio
        analysis_df['Trainers per Run'] = analysis_df['Qualified Trainers'] / analysis_df['Runs']

        # Sort by trainers per run (ascending)
        analysis_df = analysis_df.sort_values('Trainers per Run')

        # Plot trainers per run
        trainer_ratio_fig = plt.figure(figsize=(12, 6))
        trainer_ratio_plot = sns.barplot(x='Course', y='Trainers per Run', data=analysis_df)
        plt.title('Qualified Trainers per Course Run')
        plt.xticks(rotation=90)
        plt.axhline(y=1, color='r', linestyle='--', label='1:1 Ratio')
        plt.axhline(y=2, color='y', linestyle='--', label='2:1 Ratio')
        plt.tight_layout()

        # Trainer availability heatmap
        trainer_avail_fig = plt.figure(figsize=(18, 10))

        # Create a matrix of trainer availability
        trainers = list(self.consultant_data["Name"])

        # Initialize matrix with 1s (available)
        availability_data = np.ones((len(trainers), len(weeks)))

        # Fill in unavailable weeks with 0s
        for i, trainer in enumerate(trainers):
            for j, week in enumerate(weeks):
                # Make sure we don't exceed the length of weekly_calendar
                week_num = week + 1
                if week_num <= len(self.weekly_calendar):  # Add this check
                    if not self.is_trainer_available(trainer, week_num):
                        availability_data[i, j] = 0


        # Create heatmap
        sns.heatmap(availability_data, cmap=['red', 'green'], cbar=False,
                    xticklabels=[f"W{w + 1}" for w in range(len(weeks))],
                    yticklabels=trainers)
        plt.title('Trainer Availability Matrix (Green = Available, Red = Unavailable)')
        plt.xlabel('Week Number')
        plt.ylabel('Trainer')
        plt.tight_layout()

        # Calculate Critical Metrics
        total_runs = sum(self.course_run_data["Runs"])
        total_available_weeks = sum(w > 0 for w in self.weekly_working_days.values())
        total_available_days = sum(self.weekly_working_days.values())
        total_course_days = sum(self.course_run_data["Runs"] * self.course_run_data["Duration"])

        # Create a summary stats figure
        summary_fig = plt.figure(figsize=(10, 6))

        # Format as a table
        cell_text = [
            ["Total Course Runs", f"{total_runs}"],
            ["Available Weeks", f"{total_available_weeks}"],
            ["Available Working Days", f"{total_available_days}"],
            ["Required Course Days", f"{total_course_days}"],
            ["Overall Capacity Utilization",
             f"{(total_course_days / total_available_days * 100) if total_available_days > 0 else 0:.1f}%"],
            ["Runs per Available Week", f"{(total_runs / total_available_weeks) if total_available_weeks > 0 else 0:.2f}"]
        ]

        # Add color coding
        cell_colors = [['lightgray', 'lightgray'] for _ in range(len(cell_text))]

        # Highlight potential issues
        capacity_pct = (total_course_days / total_available_days * 100) if total_available_days > 0 else 0
        if capacity_pct > 90:
            cell_colors[4][1] = 'lightcoral'
        elif capacity_pct > 75:
            cell_colors[4][1] = 'lightyellow'
        else:
            cell_colors[4][1] = 'lightgreen'

        runs_per_week = (total_runs / total_available_weeks) if total_available_weeks > 0 else 0
        if runs_per_week > 1.5:
            cell_colors[5][1] = 'lightcoral'
        elif runs_per_week > 1:
            cell_colors[5][1] = 'lightyellow'
        else:
            cell_colors[5][1] = 'lightgreen'

        # Create table
        plt.axis('off')
        summary_table = plt.table(cellText=cell_text, cellColours=cell_colors,
                                  colLabels=["Metric", "Value"],
                                  loc='center', cellLoc='center')

        # Format table
        summary_table.auto_set_font_size(False)
        summary_table.set_fontsize(12)
        summary_table.scale(1, 2)

        plt.title('Overall Scheduling Feasibility Metrics', fontsize=16)
        plt.tight_layout()

        # Return all figures for Streamlit
        return {
            'monthly_analysis': fig,
            'course_week_heatmap': course_week_heatmap,
            'trainer_ratio_plot': trainer_ratio_fig,
            'trainer_avail_heatmap': trainer_avail_fig,
            'summary_metrics': summary_fig
        }

    # Updated function that doesn't use the unsupported 'pattern' property


    # Add this method to the CourseScheduler class
    def analyze_unscheduled_course(self, course, language, run):
        """Analyze why a specific course couldn't be scheduled"""
        analysis = {}

        # 1. Check trainer qualification
        qualified_trainers = self.fleximatrix.get((course, language), [])
        analysis["qualified_trainers"] = len(qualified_trainers)

        if qualified_trainers:
            # 2. Check course duration
            duration = self.course_run_data.loc[self.course_run_data["Course Name"] == course, "Duration"].iloc[0]
            analysis["duration"] = duration

            # 3. Check valid weeks (accounting for working days)
            valid_weeks = []
            for w, days in self.weekly_working_days.items():
                if days >= duration:
                    valid_weeks.append(w)
            analysis["weeks_with_enough_days"] = len(valid_weeks)

            # 4. Check week restrictions
            if course in self.week_restrictions:
                restricted_positions = []
                for position, is_restricted in self.week_restrictions[course].items():
                    if is_restricted:
                        restricted_positions.append(position)
                analysis["restricted_week_positions"] = restricted_positions

                # Calculate impact of restrictions
                restricted_weeks = 0
                for w in valid_weeks:
                    week_info = self.week_position_in_month.get(w, {})
                    if (week_info.get('is_first') and 'First' in restricted_positions) or \
                            (week_info.get('is_second') and 'Second' in restricted_positions) or \
                            (week_info.get('is_third') and 'Third' in restricted_positions) or \
                            (week_info.get('is_fourth') and 'Fourth' in restricted_positions) or \
                            (week_info.get('is_last') and 'Last' in restricted_positions):
                        restricted_weeks += 1

                analysis["weeks_restricted"] = restricted_weeks
                analysis["weeks_available_after_restrictions"] = len(valid_weeks) - restricted_weeks
            else:
                analysis["restricted_week_positions"] = []
                analysis["weeks_restricted"] = 0
                analysis["weeks_available_after_restrictions"] = len(valid_weeks)

            # 5. Check trainer availability in valid weeks
            available_weeks = {}
            for trainer in qualified_trainers:
                available_weeks[trainer] = []
                for w in valid_weeks:
                    # Skip weeks with restrictions
                    if course in self.week_restrictions:
                        week_info = self.week_position_in_month.get(w, {})
                        skip_week = False
                        for position in restricted_positions:
                            if position == 'First' and week_info.get('is_first'):
                                skip_week = True
                            elif position == 'Second' and week_info.get('is_second'):
                                skip_week = True
                            elif position == 'Third' and week_info.get('is_third'):
                                skip_week = True
                            elif position == 'Fourth' and week_info.get('is_fourth'):
                                skip_week = True
                            elif position == 'Last' and week_info.get('is_last'):
                                skip_week = True
                        if skip_week:
                            continue

                    if self.is_trainer_available(trainer, w):
                        available_weeks[trainer].append(w)

            # Count how many trainers are available for each week
            weeks_with_trainers = {}
            for w in valid_weeks:
                weeks_with_trainers[w] = sum(1 for trainer, weeks in available_weeks.items() if w in weeks)

            analysis["trainer_availability"] = available_weeks
            analysis["weeks_with_trainers"] = weeks_with_trainers
            analysis["weeks_with_any_trainer"] = sum(1 for count in weeks_with_trainers.values() if count > 0)

        return analysis

    # Add this method to the CourseScheduler class
    def _add_high_volume_course_constraints(self, model, schedule, min_course_spacing, month_deviation_penalties):
        """Add special constraints for courses with more than 5 combined runs."""
        # Commenting out high volume course constraints
        return None
        # print("\nAdding special handling for high-volume courses...")
        
        # # Calculate spacing for runs
        # runs_list = high_volume_vars[course_name]["alternating_runs"]
        # total_runs = len(runs_list)
        # total_weeks = len(self.weekly_calendar)
        
        # # Calculate maximum possible spacing while ensuring good distribution
        # quarter_sizes = []
        # for quarter, months in quarters.items():
        #     quarter_weeks = [week for week, month in self.week_to_month_map.items() if month in months]
        #     quarter_sizes.append(len(quarter_weeks))
        # min_quarter_size = min(quarter_sizes)
        
        # # Calculate ideal spacing based on total weeks and runs
        # ideal_spacing = (total_weeks - len(december_weeks)) // (total_runs - 1) if total_runs > 1 else 0
        
        # # Use ideal spacing but never go below min_course_spacing
        # spacing = max(ideal_spacing, min_course_spacing)
        
        # # For last term, reduce by 1 but not below min_course_spacing
        # reduced_spacing = max(min_course_spacing, spacing - 1)
        
        # print(f"  Dynamic spacing calculated: {spacing} weeks (reduced to {reduced_spacing} in last term)")

        # # Apply spacing constraints between consecutive runs
        # for i in range(len(runs_list) - 1):
        #     curr_lang, curr_run = runs_list[i]
        #     next_lang, next_run = runs_list[i + 1]
        #     curr_var = curr_run[2]  # Week variable
        #     next_var = next_run[2]  # Week variable
            
        #     # Check if current run is in last term (Sep-Dec)
        #     is_in_last_term = model.NewBoolVar(f"{course_name}_{curr_lang}_{i}_in_last_term")
        #     last_term_choices = []
        #     for week in last_term_weeks:
        #         is_in_week = model.NewBoolVar(f"{course_name}_{curr_lang}_{i}_in_week_{week}")
        #         model.Add(curr_var == week).OnlyEnforceIf(is_in_week)
        #         model.Add(curr_var != week).OnlyEnforceIf(is_in_week.Not())
        #         last_term_choices.append(is_in_week)
            
        #     model.AddBoolOr(last_term_choices).OnlyEnforceIf(is_in_last_term)
        #     model.AddBoolAnd([choice.Not() for choice in last_term_choices]).OnlyEnforceIf(is_in_last_term.Not())
            
        #     # Apply spacing based on whether in last term
        #     model.Add(curr_var >= next_var + spacing).OnlyEnforceIf(is_in_last_term.Not())
        #     model.Add(curr_var >= next_var + reduced_spacing).OnlyEnforceIf(is_in_last_term)
            
        #     print(f"  Added spacing constraint between {next_lang} and {curr_lang} runs")

        # print(f"  Finished processing {course_name}\n")



# Add after imports but before any classes
def create_excel_template():
    """Create an Excel template for data input"""
    # ... existing code ...

class TimeoutCallback(cp_model.CpSolverSolutionCallback):
    def __init__(self, time_limit, model, penalties, weights):
        cp_model.CpSolverSolutionCallback.__init__(self)
        self._time_limit = time_limit
        self._start_time = datetime.datetime.now()
        self._stop_search = False
        self._solution_count = 0
        self._model = model
        self._penalties = penalties
        self._weights = weights
        self._last_save_time = time.time()
        self._save_interval = 300  # Save every 5 minutes for long runs
        self._best_solution = None
        self._best_objective = float('inf')
        self._needs_save = False

    def on_solution_callback(self):
        try:
            self._solution_count += 1
            current_time = datetime.datetime.now()
            elapsed_seconds = (current_time - self._start_time).total_seconds()
            
            # Check if this is the best solution
            current_objective = self.ObjectiveValue()
            if current_objective < self._best_objective:
                self._best_objective = current_objective
                self._best_solution = {
                    'objective': current_objective,
                    'solution_time': elapsed_seconds
                }
            
            # Print basic solution information
            print(f"Solution {self._solution_count} found at {elapsed_seconds:.1f}s")
            print(f"Objective value: {current_objective:.0f}")
            
            # For long runs, check if we need to save intermediate results
            if elapsed_seconds > 3600:  # If run is longer than 1 hour
                if time.time() - self._last_save_time >= self._save_interval:
                    self._needs_save = True
                    self._last_save_time = time.time()
                    print("Flagging need to save intermediate results...")
            
            # Check if we should stop
            if elapsed_seconds >= self._time_limit:
                print("Time limit reached, stopping search...")
                self._stop_search = True
                
        except Exception as e:
            print(f"Error in solution callback: {str(e)}")
            import traceback
            print(traceback.format_exc())

    def needs_save(self):
        """Check if intermediate results should be saved"""
        return self._needs_save

    def clear_save_flag(self):
        """Clear the save flag after saving"""
        self._needs_save = False

    def stop_search(self):
        return self._stop_search

    def get_best_solution(self):
        return self._best_solution

# Create Streamlit application
def main():
    try:
        # App title with better styling
        st.title("📚 Course Scheduler App")
        st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            color: #1E88E5;
            margin-bottom: 1rem;
        }
        .section-header {
            background-color: #f0f2f6;
            padding: 10px;
            border-radius: 5px;
            margin: 15px 0;
        }
        .input-section {
            background-color: white;
            padding: 20px;
            border-radius: 10px;
            border: 1px solid #e0e0e0;
            margin: 10px 0;
        }
        .optimization-section {
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            border: 1px solid #e0e0e0;
            margin: 10px 0;
        }
        .success-box {
            background-color: #d4edda;
            color: #155724;
            padding: 10px;
            border-radius: 5px;
            margin: 10px 0;
        }
        .warning-box {
            background-color: #fff3cd;
            color: #856404;
            padding: 10px;
            border-radius: 5px;
            margin: 10px 0;
        }
        .error-box {
            background-color: #f8d7da;
            color: #721c24;
            padding: 10px;
            border-radius: 5px;
            margin: 10px 0;
        }
        .section-divider {
            margin: 20px 0;
            border-top: 1px solid #e0e0e0;
        }
        .full-width {
            width: 100%;
            padding: 0;
            margin: 0;
        }
        .log-container {
            background-color: #f8f9fa;
            border: 1px solid #e0e0e0;
            border-radius: 5px;
            padding: 15px;
            margin: 10px 0;
            height: 400px;
            overflow-y: auto;
            font-family: 'Courier New', monospace;
            white-space: pre-wrap;
            word-wrap: break-word;
            color: #333;
            font-size: 14px;
            line-height: 1.5;
        }
        .log-message {
            margin: 5px 0;
            padding: 5px;
            border-bottom: 1px solid #eee;
        }
        .log-message:last-child {
            border-bottom: none;
        }
        .log-timestamp {
            color: #666;
            font-size: 0.9em;
        }
        .log-content {
            margin-left: 10px;
        }
        </style>
        """, unsafe_allow_html=True)

        # Initialize session state
        if 'scheduler' not in st.session_state:
            st.session_state.scheduler = CourseScheduler()
            st.session_state.schedule_df = None
            st.session_state.validation_df = None
            st.session_state.utilization_df = None
            st.session_state.optimization_status = None
            st.session_state.unscheduled_courses = []
            st.session_state.optimization_running = False
            st.session_state.last_optimization_time = None

        # Create two main tabs: Input & Optimization, and Results
        tab1, tab2, tab3 = st.tabs([
            "📊 Setup & Optimization",
            "📈 Results",
            "🔍 Debug & Analysis"
        ])

        with tab1:
            # Create two columns for the main layout
            left_col, right_col = st.columns([1, 1])

            with left_col:
                st.markdown('<div class="section-header"><h3>📤 Data Input</h3></div>', unsafe_allow_html=True)
                with st.container():
                    st.markdown('<div class="input-section">', unsafe_allow_html=True)
                    
                    # Template generator
                    with st.expander("Download Template", expanded=False):
                        st.write("Use this template as a starting point:")
                        if st.button("Generate Template"):
                            template = create_excel_template()
                            st.download_button(
                                "Download Excel Template",
                                data=template,
                                file_name="course_scheduler_template.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                    # Data upload
                    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])
                    if uploaded_file and st.button("Load Data"):
                        with st.spinner("Loading data..."):
                            if st.session_state.scheduler.load_data_from_excel(uploaded_file):
                                st.success("✅ Data loaded successfully!")
                            else:
                                st.error("❌ Failed to load data")

                    # Calendar setup (only shown after data load)
                    if st.session_state.scheduler.course_run_data is not None:
                        st.markdown("### Calendar Setup")
                        year = st.number_input("Year", min_value=2020, max_value=2030, value=2025)
                        weekend_options = {"FS": "Friday-Saturday", "SS": "Saturday-Sunday"}
                        weekend = st.selectbox("Weekend Configuration", 
                                             options=list(weekend_options.keys()),
                                             format_func=lambda x: weekend_options[x])
                        
                        if st.button("Initialize Calendar"):
                            with st.spinner("Setting up calendar..."):
                                if st.session_state.scheduler.initialize_calendar(year, weekend):
                                    st.success(f"✅ Calendar initialized with {len(st.session_state.scheduler.weekly_calendar)} weeks")
                                else:
                                    st.error("❌ Calendar initialization failed")
                    st.markdown('</div>', unsafe_allow_html=True)

            with right_col:
                st.markdown('<div class="section-header"><h3>⚙️ Optimization Settings</h3></div>', unsafe_allow_html=True)
                with st.container():
                    st.markdown('<div class="optimization-section">', unsafe_allow_html=True)
                    
                    if st.session_state.scheduler.course_run_data is None:
                        st.info("Please load your data first")
                    else:
                        # Priority weights
                        st.markdown("#### Priority Weights")
                        monthly_weight = st.slider("Monthly Distribution", 1, 10, 5,
                                               help="Higher values enforce monthly targets more strictly")
                        below_min_weight = st.slider("Below Minimum Days", 30, 100, 50,
                                                 help="Higher values enforce minimum workload days")
                        above_max_weight = st.slider("Above Maximum Days", 1, 20, 10,
                                                 help="Higher values discourage exceeding maximum days")
                        affinity_weight = st.slider("Course Affinity", 1, 10, 2,
                                                help="Higher values enforce gaps between courses")

                        # Solver settings
                        st.markdown("#### Solver Configuration")
                        solution_strategy = st.selectbox(
                            "Solution Strategy",
                            options=["BALANCED", "MAXIMIZE_QUALITY", "FIND_FEASIBLE_FAST"],
                            help="BALANCED = Default, MAXIMIZE_QUALITY = Best solution but slower"
                        )
                        
                        solver_time = st.slider(
                            "Time Limit (minutes)",
                            min_value=1,
                            max_value=240,
                            value=5,
                            help="Maximum solver runtime"
                        )
                        
                        num_workers = st.slider(
                            "CPU Workers",
                            min_value=1,
                            max_value=16,
                            value=8,
                            help="More workers = faster but more CPU usage"
                        )

                        # Constraint settings
                        st.markdown("#### Constraints")
                        enforce_monthly = st.checkbox(
                            "Strict Monthly Distribution",
                            value=False,
                            help="Enforce monthly targets as hard constraints"
                        )
                        
                        prioritize_all = st.checkbox(
                            "Prioritize All Courses",
                            value=False,
                            help="Focus on scheduling all courses"
                        )

                        min_spacing = st.slider(
                            "Minimum Course Spacing",
                            min_value=1,
                            max_value=8,
                            value=2,
                            help="Minimum weeks between course runs"
                        )

                        # Run optimization button
                        if st.session_state.scheduler.weekly_calendar is not None:
                            if st.button("🚀 Run Optimization"):
                                try:
                                    # Store weights in session state
                                    st.session_state['monthly_weight'] = monthly_weight
                                    st.session_state['below_min_weight'] = below_min_weight
                                    st.session_state['above_max_weight'] = above_max_weight
                                    st.session_state['affinity_weight'] = affinity_weight

                                    # Create a container for the log
                                    log_container = st.empty()
                                    log_messages = []

                                    def update_log(message):
                                        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
                                        log_messages.append(f"[{timestamp}] {message}")
                                        # Keep only the last 100 messages
                                        if len(log_messages) > 100:
                                            log_messages.pop(0)
                                        # Create HTML for log messages
                                        log_html = '<div class="log-container">'
                                        for msg in reversed(log_messages):  # Reverse to show newest first
                                            log_html += f'<div class="log-message">{msg}</div>'
                                        log_html += '</div>'
                                        log_container.markdown(log_html, unsafe_allow_html=True)

                                    # Run optimization
                                    with st.spinner(f"Running optimization (maximum time: {solver_time} minutes)..."):
                                        update_log("Starting optimization...")
                                        update_log(f"Configuration: {solver_time} minutes, {num_workers} workers")
                                        
                                        status, schedule_df, solver, schedule, trainer_assignments, unscheduled = st.session_state.scheduler.run_optimization(
                                            monthly_weight=monthly_weight,
                                            below_min_weight=below_min_weight,
                                            above_max_weight=above_max_weight,
                                            affinity_weight=affinity_weight,
                                            solver_time_minutes=solver_time,
                                            num_workers=num_workers,
                                            min_course_spacing=min_spacing,
                                            solution_strategy=solution_strategy,
                                            enforce_monthly_distribution=enforce_monthly,
                                            prioritize_all_courses=prioritize_all
                                        )

                                        # Check if the optimization was successful
                                        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                            update_log("Optimization completed successfully!")
                                            update_log("Generating reports...")
                                            
                                            # Generate reports first
                                            validation_df = st.session_state.scheduler.generate_monthly_validation(schedule, solver)
                                            workload_df = st.session_state.scheduler.generate_trainer_workload_report(schedule, trainer_assignments, solver)
                                            
                                            # Store all results in session state
                                            st.session_state.schedule_df = schedule_df
                                            st.session_state.validation_df = validation_df
                                            st.session_state.workload_df = workload_df
                                            st.session_state.unscheduled_courses = unscheduled
                                            st.session_state.optimization_complete = True
                                            st.session_state.schedule = schedule
                                            st.session_state.solver = solver
                                            st.session_state.trainer_assignments = trainer_assignments
                                            
                                            # Save results for long runs
                                            if solver_time > 60:
                                                update_log("Saving final results...")
                                                save_intermediate_results(
                                                    schedule_df,
                                                    validation_df,
                                                    workload_df,
                                                    unscheduled,
                                                    solver.StatusName(status)
                                                )
                                            
                                            # Display success message and download button in a prominent container
                                            success_container = st.container()
                                            with success_container:
                                                if unscheduled:
                                                    update_log(f"Warning: {len(unscheduled)} courses could not be scheduled")
                                                    st.warning(f"Optimization completed with {len(unscheduled)} unscheduled courses. See Results tab for details.")
                                                else:
                                                    st.success("Optimization completed successfully! All courses were scheduled.")
                                                
                                                # Show download button in a colored box
                                                st.markdown("""
                                                <div style='padding: 1rem; background-color: #e6f3ff; border-radius: 0.5rem; margin: 1rem 0;'>
                                                    <h3 style='margin: 0; color: #0066cc;'>📥 Download Results</h3>
                                                    <p style='margin: 0.5rem 0;'>Your optimization results are ready! Click below to download.</p>
                                                </div>
                                                """, unsafe_allow_html=True)
                                                
                                                output = st.session_state.scheduler.generate_excel_report(
                                                    schedule_df,
                                                    validation_df,
                                                    workload_df
                                                )
                                                st.download_button(
                                                    label="⬇️ Download Excel Report",
                                                    data=output,
                                                    file_name="course_schedule_report.xlsx",
                                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                    key="download_excel_btn",
                                                    use_container_width=True
                                                )
                                                
                                                # Add a note about results tab
                                                st.info("👉 Switch to the Results tab to view detailed schedules and visualizations.")
                                            
                                            # Save results for long runs
                                            if solver_time > 60:
                                                update_log("Saving final results...")
                                                save_intermediate_results(
                                                    schedule_df,
                                                    validation_df,
                                                    workload_df,
                                                    unscheduled,
                                                    solver.StatusName(status)
                                                )
                                            
                                            # Display success message and download button in a prominent container
                                            success_container = st.container()
                                            with success_container:
                                                if unscheduled:
                                                    update_log(f"Warning: {len(unscheduled)} courses could not be scheduled")
                                                    st.warning(f"Optimization completed with {len(unscheduled)} unscheduled courses. See Results tab for details.")
                                                else:
                                                    st.success("Optimization completed successfully! All courses were scheduled.")
                                                
                                                # Show download button in a colored box
                                                st.markdown("""
                                                <div style='padding: 1rem; background-color: #e6f3ff; border-radius: 0.5rem; margin: 1rem 0;'>
                                                    <h3 style='margin: 0; color: #0066cc;'>📥 Download Results</h3>
                                                    <p style='margin: 0.5rem 0;'>Your optimization results are ready! Click below to download.</p>
                                                </div>
                                                """, unsafe_allow_html=True)
                                                
                                                output = st.session_state.scheduler.generate_excel_report(
                                                    schedule_df,
                                                    validation_df,
                                                    workload_df
                                                )
                                                st.download_button(
                                                    label="⬇️ Download Excel Report",
                                                    data=output,
                                                    file_name="course_schedule_report.xlsx",
                                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                    key="download_excel_btn",
                                                    use_container_width=True
                                                )
                                                
                                                # Add a note about results tab
                                                st.info("👉 Switch to the Results tab to view detailed schedules and visualizations.")
                                        else:
                                            update_log(f"Optimization failed with status: {solver.StatusName(status)}")
                                            st.error(f"Optimization failed with status: {solver.StatusName(status)}. Try adjusting your parameters or check the Debug tab.")
                                except Exception as e:
                                    import traceback
                                    st.error(f"Error during optimization: {repr(e)}")
                                    st.error(traceback.format_exc())
                                    # Defensive: show values for debugging
                                    st.error(f"status: {locals().get('status', 'N/A')}")
                                    st.error(f"schedule_df: {type(locals().get('schedule_df', None))}")
                                    st.error(f"validation_df: {type(locals().get('validation_df', None))}")
                                    st.error(f"workload_df: {type(locals().get('workload_df', None))}")
                                    if 'status' in locals() and status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                        logging.error(f"Optimization failed: {repr(e)}")
                                    else:
                                        logging.info(f"Optimization completed with status: {status if 'status' in locals() else 'UNKNOWN'}")
                            else:
                                logging.info(f"Optimization completed with status: {status if 'status' in locals() else 'UNKNOWN'}")
                    st.markdown('</div>', unsafe_allow_html=True)

        with tab2:
            if st.session_state.schedule_df is not None:
                result_tabs = st.tabs(["Schedule", "Monthly Validation", "Trainer Workload", "Visualizations"])
                
                with result_tabs[0]:
                    st.dataframe(st.session_state.schedule_df)
                
                with result_tabs[1]:
                    st.dataframe(st.session_state.validation_df)
                
                with result_tabs[2]:
                    st.dataframe(st.session_state.workload_df)
                
                with result_tabs[3]:
                    if st.button("Generate Visualizations"):
                        with st.spinner("Creating visualizations..."):
                            try:
                                # Get schedule and solver from session state
                                if 'schedule' in st.session_state and 'solver' in st.session_state and 'trainer_assignments' in st.session_state:
                                    fig1 = st.session_state.scheduler.plot_weekly_course_bar_chart(st.session_state.schedule, st.session_state.solver)
                                    fig2 = st.session_state.scheduler.plot_trainer_workload_chart(st.session_state.schedule, st.session_state.trainer_assignments, st.session_state.solver)
                                    
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.pyplot(fig1)
                                    with col2:
                                        st.pyplot(fig2)
                                else:
                                    st.error("Please run the optimization first to generate visualizations.")
                            except Exception as e:
                                st.error(f"Error generating visualizations: {str(e)}")
                
                # Add download section
                st.markdown("### Export Results")
                st.write("Download your complete optimization results:")
                
                if st.button("Generate Excel Report", key="generate_excel_results_btn"):
                    with st.spinner("Generating Excel report..."):
                        output = st.session_state.scheduler.generate_excel_report(
                            st.session_state.schedule_df,
                            st.session_state.validation_df,
                            st.session_state.workload_df
                        )
                        
                        st.download_button(
                            label="⬇️ Download Excel Report",
                            data=output,
                            file_name="course_schedule_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_results_btn",
                            use_container_width=True
                        )
            else:
                st.info("Run optimization to see results here")

        with tab3:
            if st.session_state.scheduler.course_run_data is not None:
                st.markdown("### Debug & Analysis Tools")
                
                if st.button("Analyze Constraints"):
                    with st.spinner("Analyzing constraints..."):
                        figures = st.session_state.scheduler.analyze_constraints_visually()
                        
                        st.pyplot(figures['summary_metrics'])
                        
                        with st.expander("Detailed Analysis"):
                            st.pyplot(figures['monthly_analysis'])
                            st.pyplot(figures['course_week_heatmap'])
                            st.pyplot(figures['trainer_ratio_plot'])
                            st.pyplot(figures['trainer_avail_heatmap'])
                
                if st.session_state.unscheduled_courses:
                    st.markdown("### Unscheduled Course Analysis")
                    course_options = [
                        f"{c['Course']} (Run {c['Run']}, {c['Language']})"
                        for c in st.session_state.unscheduled_courses
                    ]
                    selected = st.selectbox("Select course to analyze", course_options)
                    
                    if selected:
                        idx = course_options.index(selected)
                        course_info = st.session_state.unscheduled_courses[idx]
                        analysis = st.session_state.scheduler.analyze_unscheduled_course(
                            course_info['Course'],
                            course_info['Language'],
                            course_info['Run']
                        )
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("#### Course Details")
                            st.write(f"**Course:** {course_info['Course']}")
                            st.write(f"**Language:** {course_info['Language']}")
                            st.write(f"**Run:** {course_info['Run']}")
                            st.write(f"**Qualified Trainers:** {analysis['qualified_trainers']}")
                        
                        with col2:
                            st.markdown("#### Scheduling Analysis")
                            st.write(f"**Available Weeks:** {analysis['weeks_with_enough_days']}")
                            st.write(f"**Weeks with Trainers:** {analysis['weeks_with_any_trainer']}")
            else:
                st.info("Load data to use analysis tools")

    except Exception as e:
        st.error(f"Application error: {str(e)}")
        logging.error(f"Application error: {e}")

if __name__ == "__main__":
    main()

