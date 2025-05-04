import pandas as pd
import datetime
import streamlit as st
st.set_page_config(page_title="Course Scheduler", layout="wide")
import matplotlib.pyplot as plt
from ortools.sat.python import cp_model
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import threading
import time
import os
import numpy as np
from io import BytesIO

# Set up logging at the beginning of your app.py
logging.basicConfig(level=logging.ERROR)


# Import template generator
from template_generator import create_excel_template

class CourseScheduler:
    def __init__(self):
        self.course_run_data = None
        self.consultant_data = None
        self.priority_data = None
        self.annual_leaves = None
        self.affinity_matrix_data = None
        self.public_holidays_data = None
        self.weekly_calendar = None
        self.week_to_month_map = {}
        self.weekly_working_days = {}
        self.fleximatrix = {}  # Dictionary mapping (course, language) to list of qualified trainers
        self.course_champions = {}  # Dictionary to store champions for each course

    def _sample_courses_for_optimization(self, course_data, max_courses=500):
        """
        Sample courses for optimization when dealing with extremely large problems.
        Prioritizes courses based on various factors.
        
        Args:
            course_data: DataFrame containing course information
            max_courses: Maximum number of courses to include in optimization
            
        Returns:
            DataFrame containing sampled courses
        """
        # Create a copy to avoid modifying original
        return course_data.copy()

    def load_data_from_excel(self, uploaded_file):
        """Load all required data from an Excel file with multiple sheets"""
        try:
            print("\nDEBUG: Loading data from Excel...")
            
            # Read the Excel file with multiple sheets
            excel_file = pd.ExcelFile(uploaded_file)

            # Load each required sheet
            self.course_run_data = pd.read_excel(excel_file, sheet_name='CourseData')
            print(f"DEBUG: Loaded {len(self.course_run_data)} courses")
            
            self.consultant_data = pd.read_excel(excel_file, sheet_name='TrainerData')
            print(f"DEBUG: Loaded {len(self.consultant_data)} trainers")
            
            self.priority_data = pd.read_excel(excel_file, sheet_name='PriorityData')
            print(f"DEBUG: Loaded {len(self.priority_data)} priority levels")
            
            self.annual_leaves = pd.read_excel(excel_file, sheet_name='AnnualLeaves')
            print(f"DEBUG: Loaded {len(self.annual_leaves)} leave records")
            
            self.affinity_matrix_data = pd.read_excel(excel_file, sheet_name='AffinityMatrix')
            print(f"DEBUG: Loaded {len(self.affinity_matrix_data)} affinity rules")
            
            self.public_holidays_data = pd.read_excel(excel_file, sheet_name='PublicHolidays')
            print(f"DEBUG: Loaded {len(self.public_holidays_data)} public holidays")

            # Parse date columns
            date_columns = ['Start_Date', 'End_Date']
            for col in date_columns:
                if col in self.annual_leaves.columns:
                    self.annual_leaves[col] = pd.to_datetime(self.annual_leaves[col]).dt.date

            if 'Start Date' in self.public_holidays_data.columns:
                self.public_holidays_data['Start Date'] = pd.to_datetime(
                    self.public_holidays_data['Start Date']).dt.date

            if 'End Date' in self.public_holidays_data.columns:
                self.public_holidays_data['End Date'] = pd.to_datetime(
                    self.public_holidays_data['End Date']).dt.date

            # Load fleximatrix
            flexi_sheet = pd.read_excel(excel_file, sheet_name='Fleximatrix')
            print(f"DEBUG: Processing Fleximatrix with {len(flexi_sheet)} rows")

            # Process the Fleximatrix
            self.fleximatrix, self.course_champions = self.process_wide_fleximatrix(flexi_sheet)
            print(f"DEBUG: Created {len(self.fleximatrix)} fleximatrix entries")
            print(f"DEBUG: Identified {len(self.course_champions)} course champions")

            # Load monthly demand data
            self.monthly_demand = pd.read_excel(excel_file, sheet_name='MonthlyDemand')
            print(f"DEBUG: Loaded monthly demand for {len(self.monthly_demand)} months")

            # Load week restrictions
            try:
                restrictions_sheet = pd.read_excel(excel_file, sheet_name='WeekRestrictions')
                self.week_restrictions = {}
                
                for _, row in restrictions_sheet.iterrows():
                    course = row['Course']
                    week_type = row['Week Type']
                    restricted = row['Restricted']
                    
                    if course not in self.week_restrictions:
                        self.week_restrictions[course] = {}
                    
                    self.week_restrictions[course][week_type] = restricted
                
                print(f"DEBUG: Loaded {len(restrictions_sheet)} week restrictions")
            except Exception as e:
                print(f"DEBUG: No week restrictions found or error loading them: {e}")
                self.week_restrictions = {}

            print("DEBUG: All data loaded successfully")
            return True

        except Exception as e:
            print(f"ERROR loading data: {e}")
            return False

    def process_wide_fleximatrix(self, flexi_sheet):
        """Process a wide-format Fleximatrix where consultants are columns"""
        fleximatrix = {}
        course_champions = {}

        # Get list of consultants (all columns except CourseName, CategoryName, Language, and Champion)
        consultants = [col for col in flexi_sheet.columns
                       if col not in ['CourseName', 'CategoryName', 'Language', 'Champion']]

        # Process each row (course+language combination)
        for _, row in flexi_sheet.iterrows():
            course = row['CourseName']
            language = row['Language']
            champion = row['Champion']

            # Initialize empty list for this course+language
            fleximatrix[(course, language)] = []

            # Add the champion
            if champion and str(champion).strip():
                course_champions[(course, language)] = champion

            # Add all qualified consultants (marked with "U")
            for consultant in consultants:
                if row[consultant] == "U":
                    fleximatrix[(course, language)].append(consultant)

        return fleximatrix, course_champions


    def initialize_calendar(self, year, weekend_selection):
        """Initialize the weekly calendar based on year and weekend selection"""
        print("\nDEBUG: Initializing calendar...")
        print(f"Year: {year}")
        print(f"Weekend selection: {weekend_selection}")

        # Function to Generate Work Week Start Dates
        def get_week_start_dates(year, weekend_selection):
            first_day = datetime.date(year, 1, 1)
            if weekend_selection == "FS":  # Friday-Saturday
                week_start = 6  # Sunday
            else:  # "Saturday-Sunday"
                week_start = 0  # Monday

            weeks = []
            current_date = first_day

            # Find the first day that matches our week start
            while current_date.weekday() != week_start:
                current_date += datetime.timedelta(days=1)

            # Generate all week start dates for the year
            while current_date.year == year:
                weeks.append(current_date)
                current_date += datetime.timedelta(days=7)  # Move to next week

            print(f"DEBUG: Generated {len(weeks)} weeks")
            return weeks

        # Generate Work Week Start Dates
        self.weekly_calendar = get_week_start_dates(year, weekend_selection)

        # Ensure correct mapping of weeks to months
        self.week_to_month_map = {}
        self.week_position_in_month = {}  # Track position of week within its month

        # First, map weeks to months
        for i, week_start in enumerate(self.weekly_calendar):
            month = week_start.month  # Extract the month from the start date
            self.week_to_month_map[i + 1] = month  # Map week number to month

        # Then, calculate position of each week within its month
        for week_num, month in self.week_to_month_map.items():
            # Get all weeks in this month
            month_weeks = [w for w, m in self.week_to_month_map.items() if m == month]
            month_weeks.sort()

            # Find position (1-indexed)
            position = month_weeks.index(week_num) + 1

            # Also mark if it's the last week of the month
            is_last = (position == len(month_weeks))

            # Store as a dictionary with position info
            self.week_position_in_month[week_num] = {
                'position': position,
                'total_weeks': len(month_weeks),
                'is_first': (position == 1),
                'is_second': (position == 2),
                'is_third': (position == 3),
                'is_fourth': (position == 4),
                'is_last': is_last
            }

        # Calculate working days for each week
        self.calculate_working_days()

        print(f"DEBUG: Calendar initialized with {len(self.weekly_calendar)} weeks")
        print(f"DEBUG: Week to month mapping created for {len(self.week_to_month_map)} weeks")
        print(f"DEBUG: Week positions calculated for {len(self.week_position_in_month)} weeks")
        
        return True


    def calculate_working_days(self):
        """Calculate working days for each week accounting for public holidays"""
        self.weekly_working_days = {}

        for i, week_start in enumerate(self.weekly_calendar):
            week_num = i + 1
            working_days = 5  # Assume a full workweek

            # Check if this week is completely inside a long holiday
            fully_inside_long_holiday = False

            for _, row in self.public_holidays_data.iterrows():
                holiday_length = (row["End Date"] - row["Start Date"]).days + 1

                # If it's a long holiday (25+ days) and the week is fully inside it
                if holiday_length >= 25:
                    if week_start >= row["Start Date"] and (week_start + datetime.timedelta(days=4)) <= row["End Date"]:
                        fully_inside_long_holiday = True
                        break

            # If fully inside a long holiday, block the week completely
            if fully_inside_long_holiday:
                working_days = 0
            else:
                # Adjust only affected days for regular holidays
                for _, row in self.public_holidays_data.iterrows():
                    holiday_days_in_week = sum(
                        1 for d in range(5)  # Check only workdays
                        if row["Start Date"] <= (week_start + datetime.timedelta(days=d)) <= row["End Date"]
                        and (week_start + datetime.timedelta(days=d)).weekday() not in [5, 6]  # Exclude weekends
                    )
                    working_days -= holiday_days_in_week
                    working_days = max(0, working_days)  # Ensure no negative values

            self.weekly_working_days[week_num] = working_days


    def is_trainer_available(self, trainer_name, week_num):
        """Check if trainer is available during the given week (not on leave)"""
        week_start = self.weekly_calendar[week_num - 1]
        week_end = week_start + datetime.timedelta(days=4)  # Assuming 5-day courses

        # Check against all leave periods
        for _, leave in self.annual_leaves[self.annual_leaves["Name"] == trainer_name].iterrows():
            # If leave period overlaps with course week, trainer is unavailable
            if (leave["Start_Date"] <= week_end and leave["End_Date"] >= week_start):
                return False
        return True


    def get_trainer_priority(self, trainer_name):
        """Get priority level for a trainer based on their title"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        priority = self.priority_data.loc[self.priority_data["Title"] == title, "Priority"].iloc[0]
        return priority

    def get_qualified_trainers_by_priority(self, course, language):
        """Get qualified trainers sorted by priority for a given course and language"""
        qualified_trainers = self.fleximatrix.get((course, language), [])
        if not qualified_trainers:
            return []
        
        # Create a list of (trainer, priority) tuples
        trainer_priorities = [(trainer, self.get_trainer_priority(trainer)) for trainer in qualified_trainers]
        
        # Sort by priority (lower number = higher priority)
        trainer_priorities.sort(key=lambda x: x[1])
        
        # Group trainers by priority level
        priority_groups = {}
        for trainer, priority in trainer_priorities:
            if priority not in priority_groups:
                priority_groups[priority] = []
            priority_groups[priority].append(trainer)
            
        return priority_groups

    def find_available_trainer(self, course, language, week, duration):
        """Find the highest priority available trainer for a given course and week"""
        priority_groups = self.get_qualified_trainers_by_priority(course, language)
        
        # Check each priority level in order
        for priority in sorted(priority_groups.keys()):
            trainers = priority_groups[priority]
            
            # Check each trainer in this priority group
            for trainer in trainers:
                # Check if trainer is available for the entire duration
                available = True
                for day in range(duration):
                    check_week = week + (day // 5)  # Move to next week if days exceed 5
                    if not self.is_trainer_available(trainer, check_week):
                        available = False
                        break
                
                if available:
                    return trainer, priority
        
        return None, None

    def is_freelancer(self, trainer_name):
        """Check if a trainer is a freelancer"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        return title == "Freelancer"

    def needs_minimum_workload(self, trainer_name):
        """Check if trainer needs minimum workload and maximization (Group 1)"""
        title = self.consultant_data.loc[self.consultant_data["Name"] == trainer_name, "Title"].iloc[0]
        return title in ["Consultant", "Senior Consultant", "Partner"]

    def _create_objective_function(self, model, penalties, mode="regular", weights=None):
        """Create the objective function based on the specified mode and weights.
        
        Args:
            model: The CP-SAT model
            penalties: Dictionary containing all penalty variables
            mode: One of "regular", "prioritize_all", or "incremental"
            weights: Optional dictionary of weights to override defaults
        """
        default_weights = {
            "regular": {
                "affinity": 2,
                "champion": 4,
                "utilization": 3,
                "monthly": 5
            },
            "prioritize_all": {
                "unscheduled": 50,
                "workload": 30,
                "spacing": 20,
                "monthly": 5,
                "affinity": 2,
                "champion": 4,
                "utilization": 3
            }
        }

        # Use provided weights or defaults
        active_weights = weights or default_weights[mode]

        if mode == "regular":
            return model.Minimize(
                active_weights["affinity"] * sum(penalties["affinity"]) +
                active_weights["champion"] * sum(penalties["champion"]) +
                active_weights["utilization"] * sum(penalties["utilization"])
            )
        elif mode == "prioritize_all":
            return model.Minimize(
                active_weights["unscheduled"] * sum(penalties["unscheduled"]) +
                active_weights["workload"] * sum(penalties["workload"]) +
                active_weights["spacing"] * sum(penalties["spacing"]) +
                active_weights["monthly"] * sum(penalties["monthly"]) +
                active_weights["affinity"] * sum(penalties["affinity"]) +
                active_weights["champion"] * sum(penalties["champion"]) +
                active_weights["utilization"] * sum(penalties["utilization"])
            )
        elif mode == "incremental":
            return model.Minimize(
                2 * sum(penalties["affinity"]) +
                4 * sum(penalties["champion"]) +
                3 * sum(penalties["utilization"])
            )
        else:
            raise ValueError(f"Unknown optimization mode: {mode}")

    def run_optimization(self, monthly_weight=5,
                         affinity_weight=2,
                         solver_time_minutes=5,
                         num_workers=8, min_course_spacing=2,
                         solution_strategy="BALANCED",
                         enforce_monthly_distribution=False,
                         prioritize_all_courses=False,
                         progress_callback=None):
        
        try:
            print("\nDEBUG: Starting optimization process...")
            
            # Check if data is loaded
            if self.course_run_data is None or self.consultant_data is None or self.weekly_calendar is None:
                print("ERROR: Required data not initialized")
                return None
            
            # Initialize model
            print("DEBUG: Initializing CP-SAT model...")
            model = cp_model.CpModel()
            schedule = {}
            trainer_assignments = {}
            max_weeks = len(self.weekly_calendar)
            
            # Get total F2F runs for demand calculation
            total_f2f_runs = sum(self.course_run_data["Runs"])
            print(f"DEBUG: Total F2F runs to schedule: {total_f2f_runs}")

            # Create adjusted monthly demand dictionary
            print("DEBUG: Calculating monthly demand...")
            adjusted_f2f_demand = {}
            total_percentage = self.monthly_demand['Percentage'].sum()
            
            for _, row in self.monthly_demand.iterrows():
                month = row['Month']
                percentage = row['Percentage'] / total_percentage
                demand = round(percentage * total_f2f_runs)
                adjusted_f2f_demand[month] = demand
                print(f"DEBUG: Month {month} demand: {demand} courses")

            # Initialize penalty lists
            print("DEBUG: Initializing penalty lists...")
            month_deviation_penalties = []
            affinity_penalties = []
            below_minimum_penalties = []
            above_max_penalties = []
            workload_sums = []
            unscheduled_course_penalties = []
            workload_violation_penalties = []
            spacing_violation_penalties = []

            # Initialize trainer workload tracking
            print("\nDEBUG: Initializing trainer workload tracking...")
            trainer_workload = {name: [] for name in self.consultant_data["Name"]}
            print(f"DEBUG: Initialized workload tracking for {len(trainer_workload)} trainers")

            # Validate working days
            print("\nDEBUG: Validating working days...")
            total_working_weeks = sum(1 for w, days in self.weekly_working_days.items() if days > 0)
            print(f"DEBUG: Total working weeks available: {total_working_weeks}")
            print(f"DEBUG: Working days per week: {self.weekly_working_days}")

            # Create variables for each course run and accumulate trainer workload
            print("\nDEBUG: Creating course variables...")
            course_vars_created = 0
            courses_with_no_valid_weeks = []
            courses_with_no_trainers = []

            for _, row in self.course_run_data.iterrows():
                try:
                    course, delivery_type, language, runs, duration = row["Course Name"], row["Delivery Type"], row["Language"], \
                        row["Runs"], row["Duration"]
                    
                    # Check if there are any weeks with enough working days for this course
                    valid_weeks = [w for w, days in self.weekly_working_days.items() if days >= duration]
                    if not valid_weeks:
                        courses_with_no_valid_weeks.append((course, duration))
                        print(f"WARNING: No weeks have enough working days for {course} (needs {duration} days)")
                        continue
                    
                    qualified_trainers = self.fleximatrix.get((course, language), [])
                    if not qualified_trainers:
                        courses_with_no_trainers.append((course, language))
                        print(f"WARNING: No qualified trainers for {course} ({language})")
                        continue

                    print(f"DEBUG: Processing {course} ({language}) - {runs} runs, {len(qualified_trainers)} qualified trainers")
                    print(f"DEBUG: Course duration: {duration} days, Valid weeks: {len(valid_weeks)}")
                    
                    # Check trainer availability for this course
                    available_trainers_by_week = {}
                    for week in valid_weeks:
                        available_trainers = []
                        for trainer in qualified_trainers:
                            if self.is_trainer_available(trainer, week):
                                available_trainers.append(trainer)
                        available_trainers_by_week[week] = available_trainers
                        if not available_trainers:
                            print(f"WARNING: No trainers available in week {week} for {course}")
                    
                    weeks_with_trainers = sum(1 for trainers in available_trainers_by_week.values() if trainers)
                    print(f"DEBUG: Weeks with available trainers: {weeks_with_trainers} out of {len(valid_weeks)}")
                    
                    for i in range(runs):
                        # Allow week 0 for unscheduled courses
                        start_week = model.NewIntVar(0, max_weeks, f"start_week_{course}_{i}")
                        schedule[(course, delivery_type, language, i)] = start_week

                        # Track if course is scheduled
                        is_scheduled = model.NewBoolVar(f"{course}_{i}_is_scheduled")
                        model.Add(start_week > 0).OnlyEnforceIf(is_scheduled)
                        model.Add(start_week == 0).OnlyEnforceIf(is_scheduled.Not())
                        unscheduled_course_penalties.append(is_scheduled.Not())
                        
                        trainer_var = model.NewIntVar(0, len(qualified_trainers) - 1, f"trainer_{course}_{i}")
                        trainer_assignments[(course, delivery_type, language, i)] = trainer_var
                        
                        # Add workload tracking for each qualified trainer
                        for t_idx, trainer in enumerate(qualified_trainers):
                            is_assigned = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}")
                            model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned)
                            model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned.Not())
                            trainer_workload[trainer].append((is_assigned, duration))

                        # Convert valid weeks to soft constraints
                        for w in range(1, max_weeks + 1):
                            if w not in valid_weeks:
                                is_invalid_week = model.NewBoolVar(f"{course}_{i}_invalid_week_{w}")
                                model.Add(start_week == w).OnlyEnforceIf(is_invalid_week)
                                model.Add(start_week != w).OnlyEnforceIf(is_invalid_week.Not())
                                unscheduled_course_penalties.append(is_invalid_week)
                        
                        course_vars_created += 1

                except Exception as e:
                    print(f"ERROR processing course {course}: {e}")
                    continue

            print(f"\nDEBUG: Created variables for {course_vars_created} course runs")
            print(f"DEBUG: Courses with no valid weeks: {len(courses_with_no_valid_weeks)}")
            for course, duration in courses_with_no_valid_weeks:
                print(f"  - {course} (needs {duration} days)")
            print(f"DEBUG: Courses with no trainers: {len(courses_with_no_trainers)}")
            for course, language in courses_with_no_trainers:
                print(f"  - {course} ({language})")
            print("DEBUG: Trainer workload tracking initialized")

            if course_vars_created == 0:
                print("ERROR: No course variables were created!")
                return None

            # Add monthly distribution constraints
            print("\nDEBUG: Adding monthly distribution constraints...")
            monthly_constraints_added = 0
            for month in range(1, 13):
                target_demand = adjusted_f2f_demand.get(month, 0)
                courses_in_month = []
                month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]

                if not month_weeks:
                    print(f"WARNING: No weeks found for month {month}")
                    continue

                print(f"DEBUG: Month {month}: Target demand = {target_demand}, Available weeks = {len(month_weeks)}")

                for (course, delivery_type, language, i), week_var in schedule.items():
                    is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")
                    week_choices = []
                    for week in month_weeks:
                        is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                        model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                        model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                        week_choices.append(is_in_this_week)

                    model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                    model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())
                    courses_in_month.append(is_in_month)

                if courses_in_month and target_demand > 0:
                    if enforce_monthly_distribution:
                        model.Add(sum(courses_in_month) == target_demand)
                    else:
                        month_deviation = model.NewIntVar(0, total_f2f_runs, f"month_{month}_deviation")
                        actual_courses = sum(courses_in_month)
                        
                        is_over_target = model.NewBoolVar(f"month_{month}_over_target")
                        model.Add(actual_courses >= target_demand).OnlyEnforceIf(is_over_target)
                        model.Add(actual_courses < target_demand).OnlyEnforceIf(is_over_target.Not())
                        
                        over_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_over_dev")
                        model.Add(over_dev == actual_courses - target_demand).OnlyEnforceIf(is_over_target)
                        model.Add(over_dev == 0).OnlyEnforceIf(is_over_target.Not())
                        
                        under_dev = model.NewIntVar(0, total_f2f_runs, f"month_{month}_under_dev")
                        model.Add(under_dev == target_demand - actual_courses).OnlyEnforceIf(is_over_target.Not())
                        model.Add(under_dev == 0).OnlyEnforceIf(is_over_target)
                        
                        model.Add(month_deviation == over_dev + under_dev)
                        month_deviation_penalties.append(month_deviation)
                monthly_constraints_added += 1

            print(f"DEBUG: Added constraints for {monthly_constraints_added} months")

            # Add spacing constraints
            print("\nDEBUG: Adding course spacing constraints...")
            spacing_constraints_added = 0
            for course_name in set(self.course_run_data["Course Name"]):
                course_runs = []
                for (course, delivery_type, language, i), var in schedule.items():
                    if course == course_name:
                        course_runs.append((i, var))
                
                course_runs.sort(key=lambda x: x[0])
                
                if len(course_runs) > 1:
                    for i in range(len(course_runs) - 1):
                        run_num1, var1 = course_runs[i]
                        run_num2, var2 = course_runs[i + 1]
                        
                        spacing_violation = model.NewBoolVar(f"{course_name}_spacing_violation_{run_num1}_{run_num2}")
                        model.Add(var2 - var1 < min_course_spacing).OnlyEnforceIf(spacing_violation)
                        model.Add(var2 - var1 >= min_course_spacing).OnlyEnforceIf(spacing_violation.Not())
                        spacing_violation_penalties.append(spacing_violation)
                    
                    spacing_constraints_added += len(course_runs) - 1
                    print(f"DEBUG: Added {len(course_runs) - 1} spacing constraints for {course_name}")

            print(f"DEBUG: Added {spacing_constraints_added} total spacing constraints")

            # Add trainer workload constraints...
            print("\nDEBUG: Adding trainer workload constraints...")
            workload_constraints_added = 0
            
            # First, create a mapping of which courses each trainer might teach in each week
            trainer_week_courses = {}  # (trainer, week) -> list of course assignments
            for (course, delivery_type, language, i), week_var in schedule.items():
                qualified_trainers = self.fleximatrix.get((course, language), [])
                for t_idx, trainer in enumerate(qualified_trainers):
                    for week in range(1, max_weeks + 1):
                        is_in_week = model.NewBoolVar(f"{course}_{i}_{trainer}_in_week_{week}")
                        model.Add(week_var == week).OnlyEnforceIf(is_in_week)
                        model.Add(week_var != week).OnlyEnforceIf(is_in_week.Not())
                        
                        trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                        is_assigned_to_trainer = model.NewBoolVar(f"{course}_{i}_assigned_to_{trainer}_week_{week}")
                        model.Add(trainer_var == t_idx).OnlyEnforceIf(is_assigned_to_trainer)
                        model.Add(trainer_var != t_idx).OnlyEnforceIf(is_assigned_to_trainer.Not())
                        
                        # Course is assigned to this trainer in this week if both conditions are true
                        is_active = model.NewBoolVar(f"{course}_{i}_{trainer}_active_in_{week}")
                        model.AddBoolAnd([is_in_week, is_assigned_to_trainer]).OnlyEnforceIf(is_active)
                        model.AddBoolOr([is_in_week.Not(), is_assigned_to_trainer.Not()]).OnlyEnforceIf(is_active.Not())
                        
                        key = (trainer, week)
                        if key not in trainer_week_courses:
                            trainer_week_courses[key] = []
                        trainer_week_courses[key].append(is_active)
            
            # Now add constraints that each trainer can only teach one course per week
            print("DEBUG: Adding constraints to prevent multiple courses per trainer per week...")
            for (trainer, week), courses in trainer_week_courses.items():
                if len(courses) > 1:  # Only need constraint if trainer could teach multiple courses
                    model.Add(sum(courses) <= 1)  # Can only teach at most one course per week
            
            # Process workload constraints for each trainer
            for trainer, workload_items in trainer_workload.items():
                if not workload_items:
                    print(f"DEBUG: No workload items for trainer {trainer}")
                    continue

                max_days = self.consultant_data.loc[self.consultant_data["Name"] == trainer, "Max_Days"].iloc[0]
                print(f"DEBUG: Processing {trainer} - Max days: {max_days}, Workload items: {len(workload_items)}")
                
                # Calculate total workload
                total_workload = model.NewIntVar(0, max_days * 2, f"total_workload_{trainer}")  # Allow exceeding max
                weighted_terms = []
                
                for is_assigned, duration in workload_items:
                    term = model.NewIntVar(0, duration, f"term_{id(is_assigned)}_{duration}")
                    model.Add(term == duration).OnlyEnforceIf(is_assigned)
                    model.Add(term == 0).OnlyEnforceIf(is_assigned.Not())
                    weighted_terms.append(term)
                
                model.Add(total_workload == sum(weighted_terms))

                if self.needs_minimum_workload(trainer):  # Group 1 trainers
                    # Add penalty for being below 140 days (very high weight)
                    below_min = model.NewBoolVar(f"{trainer}_below_140")
                    model.Add(total_workload < 140).OnlyEnforceIf(below_min)
                    model.Add(total_workload >= 140).OnlyEnforceIf(below_min.Not())
                    below_minimum_penalties.append(below_min)

                    # Add small penalty for exceeding max days
                    above_max = model.NewBoolVar(f"{trainer}_above_max")
                    model.Add(total_workload > max_days).OnlyEnforceIf(above_max)
                    model.Add(total_workload <= max_days).OnlyEnforceIf(above_max.Not())
                    above_max_penalties.append(above_max)

                    # Add workload sum for maximization
                    workload_sums.append(total_workload)
                else:  # Group 2 trainers (DMD, MD, GM, Freelancer)
                    # Just enforce max days as a hard constraint
                    model.Add(total_workload <= max_days)
                
                workload_constraints_added += 1

            print(f"DEBUG: Added workload constraints for {workload_constraints_added} trainers")

            # Create unified objective function with goal programming approach
            print("DEBUG: Creating objective function...")
            model.Minimize(
                # Primary Goals (Highest Priority)
                1000 * sum(below_minimum_penalties) +    # Critical: Group 1 minimum days (140)
                
                # Secondary Goals
                (50 if prioritize_all_courses else 0) * sum(unscheduled_course_penalties) +  # Optional: Schedule all courses
                monthly_weight * sum(month_deviation_penalties) +  # Monthly distribution
                affinity_weight * sum(affinity_penalties) +       # Course affinity
                
                # Workload Goals
                2 * sum(above_max_penalties) -          # Small penalty for exceeding max days
                1 * sum(workload_sums) +                # Maximize Group 1 workload
                
                # Other Constraints (only if prioritizing all courses)
                (30 if prioritize_all_courses else 0) * sum(workload_violation_penalties) +
                (20 if prioritize_all_courses else 0) * sum(spacing_violation_penalties)
            )

            # Create solver
            print("DEBUG: Creating solver...")
            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = solver_time_minutes * 60
            solver.parameters.num_search_workers = num_workers
            
            if solution_strategy == "MAXIMIZE_QUALITY":
                solver.parameters.optimize_with_max_hs = True
            elif solution_strategy == "FIND_FEASIBLE_FAST":
                solver.parameters.search_branching = cp_model.FIXED_SEARCH
                solver.parameters.optimize_with_core = False
            
            print("DEBUG: Starting solver...")
            status = solver.Solve(model)
            print(f"DEBUG: Solver completed with status: {solver.StatusName(status)}")
            
            if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                print("DEBUG: Creating schedule DataFrame...")
                schedule_results = []
                unscheduled_courses = []
                
                # Process results
                print("DEBUG: Processing optimization results...")
                for (course, delivery_type, language, i), week_var in schedule.items():
                    try:
                        assigned_week = solver.Value(week_var)
                        if assigned_week > 0:
                            start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")

                            # Get trainer assignment
                            if (course, delivery_type, language, i) in trainer_assignments:
                                trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                                trainer_idx = solver.Value(trainer_var)

                                if (course, language) in self.fleximatrix and len(self.fleximatrix[(course, language)]) > 0:
                                    if 0 <= trainer_idx < len(self.fleximatrix[(course, language)]):
                                        trainer = self.fleximatrix[(course, language)][trainer_idx]
                                        is_champion = "✓" if self.course_champions.get(
                                            (course, language)) == trainer else " "
                                    else:
                                        trainer = "Unknown (Index Error)"
                                        is_champion = " "
                                else:
                                    trainer = "No Qualified Trainers"
                                    is_champion = " "

                                schedule_results.append({
                                    "Week": assigned_week,
                                    "Start Date": start_date,
                                    "Course": course,
                                    "Delivery Type": delivery_type,
                                    "Language": language,
                                    "Run": i + 1,
                                    "Trainer": trainer,
                                    "Champion": is_champion
                                })
                        else:
                            unscheduled_courses.append({
                                "Course": course,
                                "Delivery Type": delivery_type,
                                "Language": language,
                                "Run": i + 1
                            })
                    except Exception as e:
                        print(f"DEBUG: Error processing result for {course} run {i}: {e}")
                        unscheduled_courses.append({
                            "Course": course,
                            "Delivery Type": delivery_type,
                            "Language": language,
                            "Run": i + 1,
                            "Error": str(e)
                        })
                        continue
                
                # Create DataFrame
                print("DEBUG: Converting results to DataFrame...")
                schedule_df = pd.DataFrame(schedule_results)
                if not schedule_df.empty:
                    schedule_df = schedule_df.sort_values(by=["Week", "Course"])
                
                print(f"DEBUG: Optimization complete. Scheduled courses: {len(schedule_df) if not schedule_df.empty else 0}")
                return status, schedule_df, solver, schedule, trainer_assignments, unscheduled_courses
            else:
                print(f"DEBUG: Optimization failed with status: {solver.StatusName(status)}")
                return status, None, solver, schedule, trainer_assignments, []
                
        except Exception as e:
            print(f"DEBUG: Error in optimization process: {e}")
            import traceback
            print(f"DEBUG: Traceback:\n{traceback.format_exc()}")
            return None

    def generate_monthly_validation(self, schedule, solver):
        """Generates a monthly validation report for the schedule"""
        # Get adjusted F2F demand from the monthly demand dataframe
        adjusted_f2f_demand = {}

        # Create the dictionary using the same logic as in run_optimization
        total_f2f_runs = sum(self.course_run_data["Runs"])
        total_percentage = self.monthly_demand['Percentage'].sum()

        for _, row in self.monthly_demand.iterrows():
            month = row['Month']
            percentage = row['Percentage'] / total_percentage
            demand = round(percentage * total_f2f_runs)
            adjusted_f2f_demand[month] = demand

        # Count courses by month
        monthly_counts = {month: 0 for month in range(1, 13)}

        for (course, delivery_type, language, i), week_var in schedule.items():
            assigned_week = solver.Value(week_var)
            month = self.week_to_month_map[assigned_week]
            monthly_counts[month] += 1

        # Create validation dataframe
        validation_data = []

        total_target = 0
        total_actual = 0
        all_match = True

        for month in range(1, 13):
            target = adjusted_f2f_demand.get(month, 0)
            actual = monthly_counts[month]
            diff = actual - target

            status = "MATCH" if diff == 0 else "OVER" if diff > 0 else "UNDER"
            if diff != 0:
                all_match = False

            validation_data.append({
                "Month": str(month),  # Convert month to string
                "Target": target,
                "Actual": actual,
                "Difference": diff,
                "Status": status
            })

            total_target += target
            total_actual += actual

        # Add total row
        validation_data.append({
            "Month": "TOTAL",
            "Target": total_target,
            "Actual": total_actual,
            "Difference": total_actual - total_target,
            "Status": "MATCH" if all_match else "MISMATCH"
        })

        # Convert to DataFrame and ensure proper column types
        df = pd.DataFrame(validation_data)
        df['Target'] = df['Target'].astype(int)
        df['Actual'] = df['Actual'].astype(int)
        df['Difference'] = df['Difference'].astype(int)
        
        return df

    def run_incremental_optimization(self, solver_time_minutes=5):
        """Run optimization incrementally adding constraints to diagnose infeasibility"""
        diagnostics = []
        current_schedule = None
        
        # Base model with just course assignment variables
        model = cp_model.CpModel()
        schedule = {}
        trainer_assignments = {}
        max_weeks = len(self.weekly_calendar)
        
        # Create variables for each course run
        for _, row in self.course_run_data.iterrows():
            course, delivery_type, language, runs = row["Course Name"], row["Delivery Type"], row["Language"], row["Runs"]
            for i in range(runs):
                schedule[(course, delivery_type, language, i)] = model.NewIntVar(1, max_weeks, f"start_week_{course}_{i}")
        
        # Test 1: Just course assignments
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = solver_time_minutes * 60
        status = solver.Solve(model)
        diagnostics.append({
            "step": "Basic course assignments",
            "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
            "status": solver.StatusName(status)
        })
        
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            # Test 2: Add trainer assignments
            for (course, delivery_type, language, i), week_var in schedule.items():
                qualified_trainers = self.fleximatrix.get((course, language), [])
                if qualified_trainers:
                    trainer_assignments[(course, delivery_type, language, i)] = model.NewIntVar(
                        0, len(qualified_trainers) - 1, f"trainer_{course}_{i}")
            
            status = solver.Solve(model)
            diagnostics.append({
                "step": "Trainer assignments",
                "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
                "status": solver.StatusName(status)
            })
            
            if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                # Test 3: Add course spacing constraints
                for course_name in set(self.course_run_data["Course Name"]):
                    course_runs = []
                    for (course, delivery_type, language, i), var in schedule.items():
                        if course == course_name:
                            course_runs.append((i, var))
                    
                    course_runs.sort(key=lambda x: x[0])
                    if len(course_runs) > 1:
                        for i in range(len(course_runs) - 1):
                            run_num1, var1 = course_runs[i]
                            run_num2, var2 = course_runs[i + 1]
                            model.Add(var2 - var1 >= 2)  # Minimum 2 weeks spacing
                
                status = solver.Solve(model)
                diagnostics.append({
                    "step": "Course spacing constraints",
                    "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
                    "status": solver.StatusName(status)
                })
                
                if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                    # Test 4: Add trainer availability constraints
                    for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
                        week_var = schedule[(course, delivery_type, language, i)]
                        qualified_trainers = self.fleximatrix.get((course, language), [])
                        
                        for week in range(1, max_weeks + 1):
                            for t_idx, trainer in enumerate(qualified_trainers):
                                if not self.is_trainer_available(trainer, week):
                                    # If trainer is not available this week, they can't be assigned
                                    unavailable = model.NewBoolVar(f"{trainer}_{course}_{i}_week_{week}")
                                    model.Add(week_var == week).OnlyEnforceIf(unavailable)
                                    model.Add(trainer_var != t_idx).OnlyEnforceIf(unavailable)
                    
                    status = solver.Solve(model)
                    diagnostics.append({
                        "step": "Trainer availability constraints",
                        "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
                        "status": solver.StatusName(status)
                    })
                    
                    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                        # Test 5: Add week restrictions
                        for (course, delivery_type, language, i), week_var in schedule.items():
                            if course in self.week_restrictions:
                                restrictions = self.week_restrictions[course]
                                for week in range(1, max_weeks + 1):
                                    week_info = self.week_position_in_month.get(week, {})
                                    
                                    if ((week_info.get('is_first') and restrictions.get('First', False)) or
                                            (week_info.get('is_second') and restrictions.get('Second', False)) or
                                            (week_info.get('is_third') and restrictions.get('Third', False)) or
                                            (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or
                                            (week_info.get('is_last') and restrictions.get('Last', False))):
                                        model.Add(week_var != week)
                        
                        status = solver.Solve(model)
                        diagnostics.append({
                            "step": "Week restriction constraints",
                            "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
                            "status": solver.StatusName(status)
                        })
                        
                        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                            # Test 6: Add monthly distribution
                            total_f2f_runs = sum(self.course_run_data["Runs"])
                            adjusted_f2f_demand = {}
                            total_percentage = self.monthly_demand['Percentage'].sum()
                            
                            for _, row in self.monthly_demand.iterrows():
                                month = row['Month']
                                percentage = row['Percentage'] / total_percentage
                                demand = round(percentage * total_f2f_runs)
                                adjusted_f2f_demand[month] = demand
                            
                            for month in range(1, 13):
                                target_demand = adjusted_f2f_demand.get(month, 0)
                                courses_in_month = []
                                month_weeks = [week for week, m in self.week_to_month_map.items() if m == month]
                                
                                for (course, delivery_type, language, i), week_var in schedule.items():
                                    is_in_month = model.NewBoolVar(f"{course}_{i}_in_month_{month}")
                                    week_choices = []
                                    for week in month_weeks:
                                        is_in_this_week = model.NewBoolVar(f"{course}_{i}_in_week_{week}")
                                        model.Add(week_var == week).OnlyEnforceIf(is_in_this_week)
                                        model.Add(week_var != week).OnlyEnforceIf(is_in_this_week.Not())
                                        week_choices.append(is_in_this_week)
                                    
                                    model.AddBoolOr(week_choices).OnlyEnforceIf(is_in_month)
                                    model.AddBoolAnd([choice.Not() for choice in week_choices]).OnlyEnforceIf(is_in_month.Not())
                                    courses_in_month.append(is_in_month)
                                
                                if courses_in_month and target_demand > 0:
                                    model.Add(sum(courses_in_month) == target_demand)
                            
                            status = solver.Solve(model)
                            diagnostics.append({
                                "step": "Monthly distribution constraints",
                                "feasible": status in [cp_model.OPTIMAL, cp_model.FEASIBLE],
                                "status": solver.StatusName(status)
                            })
                            
                            if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                # Create schedule DataFrame from solution
                                schedule_results = []
                                for (course, delivery_type, language, i), week_var in schedule.items():
                                    assigned_week = solver.Value(week_var)
                                    start_date = self.weekly_calendar[assigned_week - 1].strftime("%Y-%m-%d")
                                    
                                    if (course, delivery_type, language, i) in trainer_assignments:
                                        trainer_var = trainer_assignments[(course, delivery_type, language, i)]
                                        trainer_idx = solver.Value(trainer_var)
                                        qualified_trainers = self.fleximatrix.get((course, language), [])
                                        trainer = qualified_trainers[trainer_idx] if qualified_trainers else "No trainer"
                                    else:
                                        trainer = "No trainer"
                                    
                                    schedule_results.append({
                                        "Week": assigned_week,
                                        "Start Date": start_date,
                                        "Course": course,
                                        "Delivery Type": delivery_type,
                                        "Language": language,
                                        "Run": i + 1,
                                        "Trainer": trainer
                                    })
                                
                                current_schedule = pd.DataFrame(schedule_results)
                                current_schedule = current_schedule.sort_values(by=["Week", "Course"])
        
        return diagnostics, current_schedule

    def generate_trainer_utilization_report(self, schedule, trainer_assignments, solver):
        """Generate a report showing trainer utilization"""
        trainer_workload = {}
        
        # Initialize workload dictionary for all trainers
        for trainer in self.consultant_data["Name"]:
            trainer_workload[trainer] = {
                "Assigned Days": 0,
                "Max Days": self.consultant_data.loc[
                    self.consultant_data["Name"] == trainer, "Max_Days"
                ].iloc[0],
                "Courses Assigned": 0,
                "Is Group 1": self.needs_minimum_workload(trainer)
            }
        
        # Calculate workload for each trainer
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            trainer_idx = solver.Value(trainer_var)
            qualified_trainers = self.fleximatrix.get((course, language), [])
            
            if qualified_trainers and 0 <= trainer_idx < len(qualified_trainers):
                trainer = qualified_trainers[trainer_idx]
                duration = self.course_run_data.loc[
                    (self.course_run_data["Course Name"] == course) &
                    (self.course_run_data["Language"] == language),
                    "Duration"
                ].iloc[0]
                
                trainer_workload[trainer]["Assigned Days"] += duration
                trainer_workload[trainer]["Courses Assigned"] += 1
        
        # Create DataFrame
        utilization_data = []
        for trainer, data in trainer_workload.items():
            utilization = (data["Assigned Days"] / data["Max Days"]) * 100 if data["Max Days"] > 0 else 0
            status = "UNDER" if data["Is Group 1"] and data["Assigned Days"] < 140 else (
                "OVER" if data["Assigned Days"] > data["Max Days"] else "OK"
            )
            
            utilization_data.append({
                "Trainer": trainer,
                "Assigned Days": data["Assigned Days"],
                "Max Days": data["Max Days"],
                "Courses Assigned": data["Courses Assigned"],
                "Utilization %": round(utilization, 1),
                "Group 1": "Yes" if data["Is Group 1"] else "No",
                "Status": status
            })
        
        df = pd.DataFrame(utilization_data)
        df = df.sort_values(by=["Group 1", "Utilization %"], ascending=[False, False])
        return df

    def plot_weekly_course_bar_chart(self, schedule, solver):
        """Generate a bar chart showing courses per week"""
        # Count courses per week
        weekly_counts = {}
        for (course, delivery_type, language, i), week_var in schedule.items():
            week = solver.Value(week_var)
            if week not in weekly_counts:
                weekly_counts[week] = 0
            weekly_counts[week] += 1
        
        # Create lists for plotting
        weeks = sorted(weekly_counts.keys())
        counts = [weekly_counts[w] for w in weeks]
        
        # Create figure and axis
        fig, ax = plt.subplots(figsize=(15, 6))
        
        # Plot bars
        bars = ax.bar(weeks, counts)
        
        # Customize the plot
        ax.set_xlabel('Week Number')
        ax.set_ylabel('Number of Courses')
        ax.set_title('Course Distribution by Week')
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(height)}',
                   ha='center', va='bottom')
        
        # Set the x-axis to show all week numbers
        ax.set_xticks(weeks)
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45)
        
        # Adjust layout to prevent label cutoff
        plt.tight_layout()
        
        return fig

    def plot_trainer_workload_chart(self, schedule, trainer_assignments, solver):
        """Generate a bar chart showing trainer workload distribution"""
        # Calculate workload for each trainer
        trainer_workload = {}
        
        for trainer in self.consultant_data["Name"]:
            max_days = self.consultant_data.loc[
                self.consultant_data["Name"] == trainer, "Max_Days"
            ].iloc[0]
            trainer_workload[trainer] = {
                "assigned": 0,
                "max": max_days,
                "is_group_1": self.needs_minimum_workload(trainer)
            }
        
        # Calculate assigned days
        for (course, delivery_type, language, i), trainer_var in trainer_assignments.items():
            trainer_idx = solver.Value(trainer_var)
            qualified_trainers = self.fleximatrix.get((course, language), [])
            
            if qualified_trainers and 0 <= trainer_idx < len(qualified_trainers):
                trainer = qualified_trainers[trainer_idx]
                duration = self.course_run_data.loc[
                    (self.course_run_data["Course Name"] == course) &
                    (self.course_run_data["Language"] == language),
                    "Duration"
                ].iloc[0]
                
                trainer_workload[trainer]["assigned"] += duration
        
        # Prepare data for plotting
        trainers = []
        assigned_days = []
        max_days = []
        colors = []
        
        for trainer, data in trainer_workload.items():
            if data["assigned"] > 0:  # Only include trainers with assignments
                trainers.append(trainer)
                assigned_days.append(data["assigned"])
                max_days.append(data["max"])
                
                # Color coding based on utilization and group
                utilization = (data["assigned"] / data["max"]) * 100
                if data["is_group_1"]:
                    if data["assigned"] < 140:
                        colors.append('red')  # Under minimum for Group 1
                    elif data["assigned"] > data["max"]:
                        colors.append('orange')  # Over maximum
                    else:
                        colors.append('green')  # Good range
                else:
                    if data["assigned"] > data["max"]:
                        colors.append('orange')  # Over maximum
                    else:
                        colors.append('blue')  # Good range for Group 2
        
        # Create figure and axis
        fig, ax = plt.subplots(figsize=(15, 6))
        
        # Plot bars
        x = range(len(trainers))
        bars = ax.bar(x, assigned_days, color=colors)
        
        # Plot max days as horizontal lines
        for i, max_day in enumerate(max_days):
            ax.hlines(y=max_day, xmin=i-0.4, xmax=i+0.4, colors='red', linestyles='dashed')
        
        # Plot minimum requirement line for Group 1 trainers
        ax.axhline(y=140, color='green', linestyle='--', alpha=0.5, label='Group 1 Minimum (140 days)')
        
        # Customize the plot
        ax.set_xlabel('Trainers')
        ax.set_ylabel('Days')
        ax.set_title('Trainer Workload Distribution')
        
        # Set trainer names as x-axis labels
        ax.set_xticks(x)
        ax.set_xticklabels(trainers, rotation=45, ha='right')
        
        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(height)}',
                   ha='center', va='bottom')
        
        # Add legend
        ax.legend()
        
        # Adjust layout to prevent label cutoff
        plt.tight_layout()
        
        return fig

    def generate_excel_report(self, schedule_df, monthly_validation_df, trainer_utilization_df):
        """Generate a detailed Excel report with multiple sheets"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the main schedule
            schedule_df.to_excel(writer, sheet_name='Schedule', index=False)
            
            # Write the monthly validation
            monthly_validation_df.to_excel(writer, sheet_name='Monthly Validation', index=False)
            
            # Write the trainer utilization
            trainer_utilization_df.to_excel(writer, sheet_name='Trainer Utilization', index=False)
            
            # Get the workbook to apply formatting
            workbook = writer.book
            
            # Format Schedule sheet
            ws = workbook['Schedule']
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Format Monthly Validation sheet
            ws = workbook['Monthly Validation']
            for row in ws.iter_rows(min_row=2):  # Skip header
                status = row[-1].value  # Status is the last column
                if status == "UNDER":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif status == "OVER":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # Format Trainer Utilization sheet
            ws = workbook['Trainer Utilization']
            for row in ws.iter_rows(min_row=2):  # Skip header
                status = row[-1].value  # Status is the last column
                if status == "UNDER":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif status == "OVER":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            # Save the workbook
            writer.save()
        
        # Get the value and reset the pointer
        output.seek(0)
        return output.getvalue()

    def analyze_constraints_visually(self):
        """Generate visual analysis of constraints and potential bottlenecks"""
        figures = {}
        
        # 1. Summary metrics visualization
        fig, ax = plt.subplots(figsize=(10, 6))
        metrics = {
            'Total Courses': len(self.course_run_data),
            'Total Runs': self.course_run_data['Runs'].sum(),
            'Total Trainers': len(self.consultant_data),
            'Working Weeks': sum(1 for days in self.weekly_working_days.values() if days > 0)
        }
        
        ax.bar(metrics.keys(), metrics.values())
        ax.set_title('Summary Metrics')
        plt.xticks(rotation=45)
        plt.tight_layout()
        figures['summary_metrics'] = fig
        
        # 2. Monthly demand analysis
        fig, ax = plt.subplots(figsize=(10, 6))
        monthly_demand = self.monthly_demand.copy()
        ax.bar(monthly_demand['Month'], monthly_demand['Percentage'])
        ax.set_title('Monthly Demand Distribution')
        ax.set_xlabel('Month')
        ax.set_ylabel('Percentage')
        plt.tight_layout()
        figures['monthly_analysis'] = fig
        
        # 3. Course-Week availability heatmap
        # Create a matrix of which weeks each course can be scheduled in
        courses = self.course_run_data['Course Name'].unique()
        weeks = range(1, len(self.weekly_calendar) + 1)
        availability_matrix = np.zeros((len(courses), len(weeks)))
        
        for i, course in enumerate(courses):
            duration = self.course_run_data.loc[
                self.course_run_data['Course Name'] == course, 'Duration'
            ].iloc[0]
            
            for week in weeks:
                # Check working days
                if self.weekly_working_days.get(week, 0) >= duration:
                    # Check week restrictions
                    restricted = False
                    if course in self.week_restrictions:
                        restrictions = self.week_restrictions[course]
                        week_info = self.week_position_in_month.get(week, {})
                        if ((week_info.get('is_first') and restrictions.get('First', False)) or
                                (week_info.get('is_second') and restrictions.get('Second', False)) or
                                (week_info.get('is_third') and restrictions.get('Third', False)) or
                                (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or
                                (week_info.get('is_last') and restrictions.get('Last', False))):
                            restricted = True
                    
                    if not restricted:
                        availability_matrix[i, week-1] = 1
        
        fig, ax = plt.subplots(figsize=(15, 10))
        im = ax.imshow(availability_matrix, aspect='auto', cmap='RdYlGn')
        ax.set_title('Course-Week Availability')
        ax.set_xlabel('Week')
        ax.set_ylabel('Course')
        ax.set_yticks(range(len(courses)))
        ax.set_yticklabels(courses)
        plt.colorbar(im)
        plt.tight_layout()
        figures['course_week_heatmap'] = fig
        
        # 4. Trainer analysis
        # Calculate ratio of courses they're qualified for
        trainer_ratios = []
        for trainer in self.consultant_data['Name']:
            qualified_count = sum(1 for quals in self.fleximatrix.values() if trainer in quals)
            ratio = qualified_count / len(self.fleximatrix)
            trainer_ratios.append({'Trainer': trainer, 'Qualification Ratio': ratio})
        
        trainer_ratios_df = pd.DataFrame(trainer_ratios)
        trainer_ratios_df = trainer_ratios_df.sort_values('Qualification Ratio', ascending=False)
        
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.bar(trainer_ratios_df['Trainer'], trainer_ratios_df['Qualification Ratio'])
        ax.set_title('Trainer Qualification Ratios')
        ax.set_xlabel('Trainer')
        ax.set_ylabel('Ratio of Courses Qualified For')
        plt.xticks(rotation=45)
        plt.tight_layout()
        figures['trainer_ratio_plot'] = fig
        
        # 5. Trainer availability heatmap
        availability_matrix = np.zeros((len(self.consultant_data), len(weeks)))
        trainers = self.consultant_data['Name'].tolist()
        
        for i, trainer in enumerate(trainers):
            for week in weeks:
                if self.is_trainer_available(trainer, week):
                    availability_matrix[i, week-1] = 1
        
        fig, ax = plt.subplots(figsize=(15, 10))
        im = ax.imshow(availability_matrix, aspect='auto', cmap='RdYlGn')
        ax.set_title('Trainer Availability by Week')
        ax.set_xlabel('Week')
        ax.set_ylabel('Trainer')
        ax.set_yticks(range(len(trainers)))
        ax.set_yticklabels(trainers)
        plt.colorbar(im)
        plt.tight_layout()
        figures['trainer_avail_heatmap'] = fig
        
        return figures

    def analyze_unscheduled_course(self, course, language, run):
        """Analyze why a specific course run couldn't be scheduled"""
        analysis = {}
        
        # Get course details
        course_info = self.course_run_data[
            (self.course_run_data['Course Name'] == course) &
            (self.course_run_data['Language'] == language)
        ].iloc[0]
        
        duration = course_info['Duration']
        analysis['duration'] = duration
        
        # Check qualified trainers
        qualified_trainers = self.fleximatrix.get((course, language), [])
        analysis['qualified_trainers'] = len(qualified_trainers)
        
        # Check weeks with enough working days
        valid_weeks = [w for w, days in self.weekly_working_days.items() if days >= duration]
        analysis['weeks_with_enough_days'] = len(valid_weeks)
        
        # Check week restrictions
        if course in self.week_restrictions:
            restrictions = self.week_restrictions[course]
            restricted_positions = [pos for pos, restricted in restrictions.items() if restricted]
            analysis['restricted_week_positions'] = restricted_positions
            
            # Count weeks eliminated by restrictions
            restricted_weeks = 0
            for week in valid_weeks:
                week_info = self.week_position_in_month.get(week, {})
                if ((week_info.get('is_first') and restrictions.get('First', False)) or
                        (week_info.get('is_second') and restrictions.get('Second', False)) or
                        (week_info.get('is_third') and restrictions.get('Third', False)) or
                        (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or
                        (week_info.get('is_last') and restrictions.get('Last', False))):
                    restricted_weeks += 1
            analysis['weeks_restricted'] = restricted_weeks
        
        # Check trainer availability for valid weeks
        weeks_with_trainers = {}
        for week in valid_weeks:
            available_trainers = []
            for trainer in qualified_trainers:
                if self.is_trainer_available(trainer, week):
                    available_trainers.append(trainer)
            weeks_with_trainers[week] = len(available_trainers)
        
        analysis['weeks_with_trainers'] = weeks_with_trainers
        analysis['weeks_with_any_trainer'] = sum(1 for count in weeks_with_trainers.values() if count > 0)
        
        return analysis


# Create Streamlit application
def main():
    try:
        # App title with better styling
        st.title("📚 Course Scheduler App")
        st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            color: #1E88E5;
            margin-bottom: 1rem;
        }
        .section-header {
            background-color: #f0f2f6;
            padding: 10px;
            border-radius: 5px;
            margin-top: 15px;
            margin-bottom: 15px;
        }
        .success-box {
            background-color: #d4edda;
            color: #155724;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        .warning-box {
            background-color: #fff3cd;
            color: #856404;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        .error-box {
            background-color: #f8d7da;
            color: #721c24;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        }
        </style>
        """, unsafe_allow_html=True)

        st.markdown("Optimize your course schedules and trainer assignments efficiently")

        # Initialize session state for storing the scheduler object
        if 'scheduler' not in st.session_state:
            st.session_state.scheduler = CourseScheduler()

        if 'schedule_df' not in st.session_state:
            st.session_state.schedule_df = None

        if 'validation_df' not in st.session_state:
            st.session_state.validation_df = None

        if 'utilization_df' not in st.session_state:
            st.session_state.utilization_df = None

        if 'optimization_status' not in st.session_state:
            st.session_state.optimization_status = None

        if 'unscheduled_courses' not in st.session_state:
            st.session_state.unscheduled_courses = []

        # Create tabs for a more organized workflow
        tab1, tab2, tab3, tab4 = st.tabs([
            "📤 Data Input & Setup",
            "⚙️ Optimization Settings",
            "📊 Results",
            "🔍 Debug & Troubleshoot"
        ])

        with tab1:
            st.markdown('<div class="section-header"><h2>1. Upload & Setup Data</h2></div>', unsafe_allow_html=True)

            col1, col2 = st.columns([1, 1])

            with col1:
                # Template generator section
                with st.expander("Not sure about the format? Download our template first", expanded=False):
                    st.write("Use this template as a starting point for your data:")
                    if st.button("Generate Template", key="generate_template_button"):
                        template = create_excel_template()
                        st.download_button(
                            "Download Excel Template",
                            data=template,
                            file_name="course_scheduler_template.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_template_btn"
                        )

                # Data upload section
                uploaded_file = st.file_uploader("Upload Excel file with course and trainer data", type=["xlsx"])

                if uploaded_file is not None:
                    if st.button("Load Data", key="load_data_btn"):
                        with st.spinner("Loading data from Excel..."):
                            success = st.session_state.scheduler.load_data_from_excel(uploaded_file)
                            if success:
                                st.markdown('<div class="success-box">✅ Data loaded successfully!</div>',
                                            unsafe_allow_html=True)
                            else:
                                st.markdown(
                                    '<div class="error-box">❌ Failed to load data. Please check your Excel file format.</div>',
                                    unsafe_allow_html=True)

            with col2:
                # Data format guide
                st.markdown("### Required Excel Sheets")
                with st.expander("View format requirements", expanded=False):
                    st.markdown("""
                    Your Excel file should contain these sheets:
    
                    1. **CourseData**: Course information
                       - Columns: Course Name, Delivery Type, Language, Runs, Duration
    
                    2. **TrainerData**: Trainer information
                       - Columns: Name, Title, Max_Days
    
                    3. **PriorityData**: Trainer priority levels
                       - Columns: Title, Priority
    
                    4. **AnnualLeaves**: Trainer leave periods
                       - Columns: Name, Start_Date, End_Date
    
                    5. **AffinityMatrix**: Course timing relationships
                       - Columns: Course 1, Course 2, Gap Weeks
    
                    6. **PublicHolidays**: Holiday periods
                       - Columns: Start Date, End Date
    
                    7. **Fleximatrix**: Trainer qualifications
                       - "U" marks qualified trainers
    
                    8. **WeekRestrictions**: Course timing constraints
                       - Columns: Course, Week Type, Restricted, Notes
    
                    9. **MonthlyDemand**: Distribution targets
                       - Columns: Month, Percentage
                    """)

                # Calendar setup (only shown after data is loaded)
                if st.session_state.scheduler.course_run_data is not None:
                    st.markdown("### Calendar Setup")
                    year = st.number_input("Select scheduling year", min_value=2020, max_value=2030, value=2025)
                    weekend_options = {"FS": "Friday-Saturday", "SS": "Saturday-Sunday"}
                    weekend_selection = st.selectbox("Select weekend configuration",
                                                     options=list(weekend_options.keys()),
                                                     format_func=lambda x: weekend_options[x])

                    if st.button("Initialize Calendar", key="init_calendar_btn"):
                        with st.spinner("Generating calendar..."):
                            success = st.session_state.scheduler.initialize_calendar(year, weekend_selection)
                            if success:
                                st.markdown(
                                    f'<div class="success-box">✅ Calendar created with {len(st.session_state.scheduler.weekly_calendar)} weeks</div>',
                                    unsafe_allow_html=True)
                            else:
                                st.markdown('<div class="error-box">❌ Failed to create calendar</div>',
                                            unsafe_allow_html=True)

        # In the "Optimization Settings" tab
        with tab2:
            if st.session_state.scheduler.course_run_data is None:
                st.info("Please load your data in the 'Data Input & Setup' tab first")
            else:
                st.markdown('<div class="section-header"><h2>2. Optimization Parameters</h2></div>',
                            unsafe_allow_html=True)

                # Create a container with custom styling for the sliders
                st.markdown("""
                <style>
                .slider-container {
                    background-color: #f8f9fa;
                    border-radius: 10px;
                    padding: 20px;
                    border: 1px solid #e9ecef;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
                    margin-bottom: 20px;
                }
                .slider-header {
                    font-size: 1.2rem;
                    font-weight: 600;
                    margin-bottom: 15px;
                    color: #1E88E5;
                    border-bottom: 1px solid #e9ecef;
                    padding-bottom: 10px;
                }
                </style>
                """, unsafe_allow_html=True)

                # Create a two-column layout with sliders in the left column
                col1, col2 = st.columns([1, 1])

                with col1:
                    # Priority Weight Sliders in a nice frame
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Distribution Weights</div>', unsafe_allow_html=True)
                    monthly_weight = st.slider("Monthly Distribution Priority", 1, 10, 5,
                                               help="Higher values enforce monthly targets more strictly")
                    affinity_weight = st.slider("Course Affinity Priority", 1, 10, 2,
                                               help="Higher values enforce gaps between related courses")
                    st.markdown('</div>', unsafe_allow_html=True)

                    # Remove utilization weight as it's now handled by goal programming

                    # Constraint Management sliders in a nice frame
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Constraint Settings</div>', unsafe_allow_html=True)
                    min_course_spacing = st.slider(
                        "Minimum Weeks Between Course Runs",
                        min_value=1,
                        max_value=8,
                        value=2,
                        help="Lower values allow courses to be scheduled closer together"
                    )
                    # Remove utilization target as it's now handled differently
                    st.markdown('</div>', unsafe_allow_html=True)

                with col2:
                    # Checkboxes and other options
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Optimization Strategy</div>', unsafe_allow_html=True)
                    enforce_monthly = st.checkbox(
                        "Enforce Monthly Distribution as Hard Constraint",
                        value=False,
                        help="Uncheck to allow flexibility in monthly distribution (recommended for feasibility)"
                    )

                    prioritize_all_courses = st.checkbox(
                        "Prioritize scheduling all courses",
                        value=False,
                        help="When enabled, the optimizer will prioritize scheduling all courses, potentially at the expense of other goals"
                    )

                    solution_strategy = st.selectbox(
                        "Solution Strategy",
                        options=["BALANCED", "MAXIMIZE_QUALITY", "FIND_FEASIBLE_FAST"],
                        index=0,
                        help="BALANCED = Default, MAXIMIZE_QUALITY = Best solution but slower, FIND_FEASIBLE_FAST = Any valid solution quickly"
                    )
                    solver_time = st.slider(
                        "Solver Time Limit (minutes)",
                        min_value=1,
                        max_value=60,
                        value=5,
                        help="Maximum time the optimizer will run before returning best solution found"
                    )
                    num_workers = st.slider(
                        "Number of Search Workers",
                        min_value=1,
                        max_value=16,
                        value=8,
                        help="More workers can find solutions faster but use more CPU"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                    # Add new information box about trainer workload goals
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Trainer Workload Goals</div>', unsafe_allow_html=True)
                    st.info("""
                    **Group 1** (Consultants, Senior Consultants, Partners):
                    - Minimum 140 days (strictly enforced)
                    - Maximizes workload while respecting max days
                    - Small penalty for exceeding max days
                    
                    **Group 2** (DMD, MD, GM, Freelancer):
                    - Only enforces maximum days limit
                    - No minimum requirements
                    - No workload maximization
                    """)
                    st.markdown('</div>', unsafe_allow_html=True)

                    # Troubleshooting options
                    st.markdown('<div class="slider-container">', unsafe_allow_html=True)
                    st.markdown('<div class="slider-header">Troubleshooting</div>', unsafe_allow_html=True)
                    if st.button("Disable All Week Restrictions", key="disable_restrictions_btn"):
                        st.session_state.scheduler.week_restrictions = {}
                        st.markdown(
                            '<div class="success-box">✅ All week restrictions have been disabled for this optimization run</div>',
                            unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                # Run the optimization button at the bottom
                if st.session_state.scheduler.weekly_calendar is not None:
                    st.markdown("### Run Optimization")
                    # This is just the modified section for the "Run Optimization" button
                    # in the Optimization Settings tab (tab2)

                    # Find this section in app.py around line 1800-1850
                    if st.button("Optimize Schedule", key="optimize_schedule_btn"):
                        try:
                            # Create containers for progress bar and output
                            progress_container = st.container()
                            output_container = st.container()
                            
                            # Initialize progress bar and status text
                            progress_bar = progress_container.progress(0)
                            
                            with output_container:
                                st.markdown("### Optimization Progress")
                                log_area = st.empty()
                                log_messages = ["Initializing optimization..."]
                            
                            # Function to update progress and output - now keeps a history of messages
                            def update_progress(percent, message):
                                if percent is not None:
                                    progress_bar.progress(percent)
                                if message and message.strip():
                                    log_messages.append(message)
                                    # Display the last 15 messages (or all if fewer) to keep the display manageable
                                    displayed_msgs = log_messages[-15:]
                                    log_area.markdown('\n\n'.join(displayed_msgs))
                            
                            with st.spinner(f"Running optimization (maximum time: {solver_time} minutes)..."):
                                try:
                                    result = st.session_state.scheduler.run_optimization(
                                        monthly_weight=monthly_weight,
                                        affinity_weight=affinity_weight,
                                        solver_time_minutes=solver_time,
                                        num_workers=num_workers,
                                        min_course_spacing=min_course_spacing,
                                        solution_strategy=solution_strategy,
                                        enforce_monthly_distribution=enforce_monthly,
                                        prioritize_all_courses=prioritize_all_courses,
                                        progress_callback=update_progress
                                    )
                                    
                                    if result is None or len(result) != 6:
                                        st.error("Optimization failed: Invalid return value from optimizer")
                                        return
                                                
                                    status, schedule_df, solver, schedule, trainer_assignments, unscheduled_courses = result
                                    
                                    # Store the optimization status in session state
                                    st.session_state.optimization_status = status
                                    st.session_state.unscheduled_courses = unscheduled_courses if unscheduled_courses else []

                                    # Check if the optimization was successful
                                    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                        # Store results in session state
                                        st.session_state.schedule_df = schedule_df

                                        # Generate the validation and utilization reports
                                        st.session_state.validation_df = st.session_state.scheduler.generate_monthly_validation(
                                            schedule, solver)
                                        st.session_state.utilization_df = st.session_state.scheduler.generate_trainer_utilization_report(
                                            schedule, trainer_assignments, solver)

                                        # Display success message with unscheduled course info
                                        if st.session_state.unscheduled_courses:
                                            st.warning(
                                                f"Optimization completed with {len(st.session_state.unscheduled_courses)} unscheduled courses. See Results tab for details.")
                                        else:
                                            st.success("Optimization completed successfully! All courses were scheduled.")
                                    else:
                                        st.error(
                                            f"Optimization failed with status: {solver.StatusName(status)}. Try adjusting your parameters or check the Debug tab.")
                                except Exception as e:
                                    st.error(f"Error during optimization: {str(e)}")
                                    print(f"DEBUG: Optimization error details: {e}")
                                    import traceback
                                    print(f"DEBUG: Traceback:\n{traceback.format_exc()}")
                        except Exception as e:
                            # Only log as error if the optimization status is not FEASIBLE or OPTIMAL
                            if not hasattr(st.session_state, 'optimization_status') or st.session_state.optimization_status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                st.error(f"An error occurred: {e}")
                                logging.error(f"Optimization error: {str(e)}")

        with tab3:
            if st.session_state.schedule_df is not None:
                st.markdown('<div class="section-header"><h2>4. Optimization Results</h2></div>', unsafe_allow_html=True)

                # Display tabs for different views
                result_tab1, result_tab2, result_tab3, result_tab4 = st.tabs(
                    ["📅 Schedule", "📊 Monthly Validation", "👥 Trainer Utilization", "📈 Visualizations"])

                with result_tab1:
                    st.subheader("Course Schedule")
                    st.dataframe(st.session_state.schedule_df, use_container_width=True)

                with result_tab2:
                    st.subheader("Monthly Validation")
                    st.dataframe(st.session_state.validation_df, use_container_width=True)

                with result_tab3:
                    st.subheader("Trainer Utilization")
                    st.dataframe(st.session_state.utilization_df, use_container_width=True)

                with result_tab4:
                    st.subheader("Visualizations")

                    # We need to rerun the plotting functions
                    # Add this code to the Results tab (tab3) section of your app.py
                    # Replace the existing visualization button code with this modified version

                    # Replace the visualization button code in the Results tab section with this version:

                    # Update the Generate Visualizations button to include unscheduled_courses in the returned values
                    # Find this in the result_tab4 (Visualizations) tab

                    if st.button("Generate Visualizations", key="generate_viz_btn"):
                        with st.spinner("Generating visualizations..."):
                            try:
                                # Only try to generate new optimization if we don't have a completed one
                                if not hasattr(st.session_state, 'schedule_df') or st.session_state.schedule_df is None:
                                    st.warning(
                                        "No optimization results found. Running a quick optimization to generate visualizations...")
                                    # Run a quick new optimization just to get solver state and schedule objects
                                    status, schedule_df, solver, schedule, trainer_assignments, _ = st.session_state.scheduler.run_optimization(
                                        monthly_weight=5, champion_weight=4,
                                        utilization_weight=3, affinity_weight=2,
                                        utilization_target=70, solver_time_minutes=1,
                                        num_workers=8, min_course_spacing=2,
                                        solution_strategy="FIND_FEASIBLE_FAST",
                                        enforce_monthly_distribution=False,
                                        prioritize_all_courses=False,
                                        progress_callback=None
                                    )

                                    if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                                        st.error(
                                            f"Could not find a feasible solution for visualization. Status: {solver.StatusName(status)}")
                                        st.info(
                                            "Try running a full optimization with the 'Optimize Schedule' button first.")
                                        return
                                else:
                                    # Run a very short optimization to get the solver state and variables
                                    # but use the existing solution as a starting point
                                    st.info("Using existing optimization results for visualization...")
                                    status, _, solver, schedule, trainer_assignments, _ = st.session_state.scheduler.run_optimization(
                                        monthly_weight=5, champion_weight=4,
                                        utilization_weight=3, affinity_weight=2,
                                        utilization_target=70, solver_time_minutes=0.1,  # Very short time
                                        num_workers=8, min_course_spacing=2,
                                        solution_strategy="FIND_FEASIBLE_FAST",
                                        enforce_monthly_distribution=False,
                                        prioritize_all_courses=False,
                                        progress_callback=None
                                    )

                                # Rest of the visualization code stays the same...

                                # Safety check that schedule and solver are valid
                                if schedule and solver and len(schedule) > 0:
                                    # Generate and display charts
                                    col1, col2 = st.columns(2)

                                    with col1:
                                        st.subheader("Weekly Course Distribution")
                                        try:
                                            fig1 = st.session_state.scheduler.plot_weekly_course_bar_chart(schedule,
                                                                                                           solver)
                                            st.pyplot(fig1)
                                        except Exception as e:
                                            st.error(f"Error generating weekly chart: {str(e)}")
                                            st.info("The visualization couldn't be generated due to data issues.")

                                    with col2:
                                        st.subheader("Trainer Workload")
                                        try:
                                            fig2 = st.session_state.scheduler.plot_trainer_workload_chart(
                                                schedule, trainer_assignments, solver)
                                            st.pyplot(fig2)
                                        except Exception as e:
                                            st.error(f"Error generating trainer chart: {str(e)}")
                                            st.info("The visualization couldn't be generated due to data issues.")
                                else:
                                    st.error("Couldn't generate valid solver state for visualization.")
                                    st.info(
                                        "Please try running a complete optimization first with the 'Optimize Schedule' button.")

                            except Exception as e:
                                st.error(f"Error generating visualizations: {str(e)}")
                                st.info(
                                    "Try running a complete optimization first with the 'Optimize Schedule' button.")
                        # Add this to your app.py file in the Results tab section,
                        # under the result_tab4 (Visualizations) tab:

                # Export Results
                st.markdown("### Export Results")
                st.write("Download your complete optimization results:")

                try:
                    if st.button("Generate Excel Report", key="generate_excel_btn"):
                        with st.spinner("Generating Excel report..."):
                            # Check if we have all required data
                            if (st.session_state.schedule_df is None or 
                                st.session_state.validation_df is None or 
                                st.session_state.utilization_df is None):
                                st.error("Missing required data. Please run optimization first.")
                                return

                            # Generate the report
                            excel_data = st.session_state.scheduler.generate_excel_report(
                                st.session_state.schedule_df,
                                st.session_state.validation_df,
                                st.session_state.utilization_df
                            )

                            # Create download button
                            st.download_button(
                                label="⬇️ Download Excel Report",
                                data=excel_data,
                                file_name="course_schedule_report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_excel_btn",
                                help="Click to download the complete schedule report"
                            )
                            st.success("Excel report generated successfully! Click the download button above to save it.")
                except Exception as e:
                    st.error(f"Error generating Excel report: {str(e)}")
                    print(f"DEBUG: Excel generation error: {e}")
            else:
                if st.session_state.optimization_status == cp_model.INFEASIBLE:
                    st.error(
                        "The last optimization run was INFEASIBLE. Please go to the 'Debug & Analyze' tab to diagnose and fix constraints before trying again.")
                elif st.session_state.optimization_status is not None:
                    st.error(
                        f"The last optimization run failed with status: {st.session_state.optimization_status}. Please adjust your parameters and try again.")
                else:
                    st.info("Run the optimization in the 'Optimization Settings' tab to see results here.")

                # Add this code to the Results tab section (tab3)
                # After loading and displaying the schedule dataframe

                # Instead, always access unscheduled_courses through session state:
                if 'unscheduled_courses' in st.session_state and st.session_state.unscheduled_courses:
                    st.warning(f"{len(st.session_state.unscheduled_courses)} courses could not be scheduled")
                    with st.expander("View Unscheduled Courses"):
                        unscheduled_df = pd.DataFrame(st.session_state.unscheduled_courses)
                        st.dataframe(unscheduled_df, use_container_width=True)


                # In the results tab, after displaying the schedule dataframe:
                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    st.warning(f"{len(st.session_state.unscheduled_courses)} courses could not be scheduled")
                    with st.expander("View Unscheduled Courses"):
                        unscheduled_df = pd.DataFrame(st.session_state.unscheduled_courses)
                        st.dataframe(unscheduled_df, use_container_width=True)

                        st.info("""
                        **Why couldn't these courses be scheduled?**
                        - Insufficient qualified trainers
                        - Trainer availability constraints
                        - Week restrictions limiting valid weeks
                        - Spacing requirements between course runs

                        Try enabling "Prioritize scheduling all courses" in the Optimization Settings tab or relaxing some constraints.
                        """)


        with tab4:
            if st.session_state.scheduler.course_run_data is None:
                st.info("Please load your data in the 'Data Input & Setup' tab first")
            elif st.session_state.scheduler.weekly_calendar is None:
                st.info("Please initialize the calendar before debugging")
            else:
                st.markdown('<div class="section-header"><h2>3. Debug & Analysis Tools</h2></div>', unsafe_allow_html=True)
                st.markdown("Use these tools to diagnose issues and optimize your schedule:")

                # Create two columns for side-by-side analysis tools
                debug_col1, debug_col2 = st.columns(2)

                with debug_col1:
                    st.markdown("### Incremental Constraint Diagnosis")
                    st.markdown("Find exactly which constraint is causing infeasibility:")

                    if st.button("Diagnose with Incremental Constraints", key="incremental_constraints_btn"):
                        with st.spinner("Running incremental constraint analysis..."):
                            status, diagnostics, schedule_df = st.session_state.scheduler.run_incremental_optimization()

                            # Display diagnostics in a table
                            st.subheader("Constraint Diagnosis Results")
                            diagnosis_df = pd.DataFrame(diagnostics)
                            st.dataframe(diagnosis_df, use_container_width=True)

                            # Interpretation
                            st.subheader("Interpretation")

                            # Find the first step that became infeasible
                            infeasible_step = None
                            for step in diagnostics:
                                if not step["feasible"]:
                                    infeasible_step = step["step"]
                                    break

                            if infeasible_step:
                                st.markdown(
                                    f'<div class="error-box">The model becomes infeasible when adding: {infeasible_step}</div>',
                                    unsafe_allow_html=True)

                                if infeasible_step == "Course spacing constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Try reducing the minimum spacing between course runs using the slider in the Optimization Settings tab.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Trainer availability constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> There might be insufficient trainer availability. Check your trainer leave periods or add more qualified trainers.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Week restriction constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Week restrictions are too limiting. Try disabling some week restrictions using the button in Advanced Settings.</div>',
                                        unsafe_allow_html=True)
                                elif infeasible_step == "Monthly distribution constraints":
                                    st.markdown(
                                        '<div class="warning-box"><b>Solution:</b> Monthly distribution requirements cannot be satisfied. Uncheck "Enforce Monthly Distribution as Hard Constraint" option.</div>',
                                        unsafe_allow_html=True)
                            else:
                                st.markdown(
                                    '<div class="success-box">✅ The model is feasible with all constraints! If you\'re still having issues with the full optimization, it might be due to complex interactions between constraints.</div>',
                                    unsafe_allow_html=True)

                            # If we got a feasible schedule, show it
                            if schedule_df is not None:
                                st.subheader("Feasible Schedule Found")
                                st.dataframe(schedule_df, use_container_width=True)

                with debug_col2:
                    st.markdown("### Visual Constraint Analysis")
                    st.markdown("Visualize constraints and potential bottlenecks:")

                    if st.button("Visual Constraint Analysis", key="visual_analysis_btn"):
                        with st.spinner("Generating visual analysis of constraints..."):
                            figures = st.session_state.scheduler.analyze_constraints_visually()

                            st.subheader("Summary Metrics")
                            st.pyplot(figures['summary_metrics'])

                            with st.expander("Monthly Demand Analysis", expanded=False):
                                st.pyplot(figures['monthly_analysis'])

                            with st.expander("Course-Week Availability", expanded=False):
                                st.pyplot(figures['course_week_heatmap'])

                            with st.expander("Trainer Analysis", expanded=False):
                                st.pyplot(figures['trainer_ratio_plot'])
                                st.pyplot(figures['trainer_avail_heatmap'])

                            st.info(
                                "Red areas in the heatmaps indicate restrictions or unavailability that could be causing infeasibility.")

            # Replace the function definition:
            def display_scheduling_problem_analysis():
                st.subheader("Analyze Unscheduled Course")

                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    # Create dropdown to select unscheduled course
                    course_options = [f"{c['Course']} (Run {c['Run']}, {c['Language']})" for c in
                                      st.session_state.unscheduled_courses]
                    selected_course = st.selectbox("Select course to analyze", course_options)

                    if selected_course:
                        # Extract course details from selection
                        idx = course_options.index(selected_course)
                        course_info = st.session_state.unscheduled_courses[idx]

                        # Run analysis
                        analysis = st.session_state.scheduler.analyze_unscheduled_course(
                            course_info['Course'], course_info['Language'], course_info['Run'])

                        # Display results
                        col1, col2 = st.columns(2)

                        with col1:
                            st.markdown("### Course Information")
                            st.markdown(f"**Course**: {course_info['Course']}")
                            st.markdown(f"**Language**: {course_info['Language']}")
                            st.markdown(f"**Run**: {course_info['Run']}")
                            st.markdown(f"**Qualified Trainers**: {analysis['qualified_trainers']}")
                            st.markdown(f"**Course Duration**: {analysis['duration']} days")

                            if analysis['qualified_trainers'] == 0:
                                st.error("❌ No qualified trainers for this course")
                                st.markdown("**Solution**: Add more qualified trainers in the Fleximatrix")

                        with col2:
                            st.markdown("### Week Availability")
                            st.markdown(f"**Weeks with enough workdays**: {analysis['weeks_with_enough_days']}")

                            if 'restricted_week_positions' in analysis:
                                restrictions = analysis['restricted_week_positions']
                                if restrictions:
                                    st.markdown(f"**Restricted week positions**: {', '.join(restrictions)}")
                                    st.markdown(f"**Weeks eliminated by restrictions**: {analysis['weeks_restricted']}")

                            st.markdown(f"**Weeks with any trainer available**: {analysis['weeks_with_any_trainer']}")

                            if analysis['weeks_with_any_trainer'] == 0:
                                st.error("❌ No weeks with both enough workdays and available trainers")
                                st.markdown("**Solution**: Reduce leave overlaps or add more qualified trainers")

                        # Draw a timeline of week availability
                        if analysis['qualified_trainers'] > 0:
                            st.markdown("### Week-by-Week Availability")

                            # Create dataframe for visualization
                            weeks = sorted(analysis.get('weeks_with_trainers', {}).keys())
                            if weeks:
                                week_data = []

                                for week in weeks:
                                    trainers_available = analysis['weeks_with_trainers'].get(week, 0)
                                    month = st.session_state.scheduler.week_to_month_map.get(week, 0)
                                    week_info = st.session_state.scheduler.week_position_in_month.get(week, {})

                                    # Check if this week is restricted
                                    is_restricted = False
                                    if course_info['Course'] in st.session_state.scheduler.week_restrictions:
                                        restrictions = st.session_state.scheduler.week_restrictions[
                                            course_info['Course']]
                                        if (week_info.get('is_first') and restrictions.get('First', False)) or \
                                                (week_info.get('is_second') and restrictions.get('Second', False)) or \
                                                (week_info.get('is_third') and restrictions.get('Third', False)) or \
                                                (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or \
                                                (week_info.get('is_last') and restrictions.get('Last', False)):
                                            is_restricted = True

                                    working_days = st.session_state.scheduler.weekly_working_days.get(week, 0)

                                    week_data.append({
                                        'Week': week,
                                        'Month': month,
                                        'Trainers Available': trainers_available,
                                        'Working Days': working_days,
                                        'Is Restricted': is_restricted
                                    })

                                week_df = pd.DataFrame(week_data)

                                # Color columns based on conditions
                                def color_trainers(val):
                                    if val == 0:
                                        color = 'red'
                                    elif val < 3:
                                        color = 'orange'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_days(val, duration=analysis['duration']):
                                    if val < duration:
                                        color = 'red'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_restricted(val):
                                    return 'background-color: red' if val else 'background-color: white'

                                # Apply styling
                                styled_df = week_df.style.applymap(color_trainers, subset=['Trainers Available']) \
                                    .applymap(lambda x: color_days(x, analysis['duration']), subset=['Working Days']) \
                                    .applymap(color_restricted, subset=['Is Restricted'])

                                st.dataframe(styled_df)

                                # Identify the problem
                                problems = []
                                if analysis['weeks_with_enough_days'] == 0:
                                    problems.append("❌ No weeks have enough working days for this course duration")
                                if analysis.get('weeks_restricted') == analysis['weeks_with_enough_days']:
                                    problems.append("❌ All potential weeks are restricted by week position constraints")
                                if analysis['weeks_with_any_trainer'] == 0:
                                    problems.append("❌ No trainers are available in valid weeks")

                                st.markdown("### Summary")

                                if problems:
                                    for problem in problems:
                                        st.error(problem)
                                else:
                                    st.warning(
                                        "⚠️ This course may be unscheduled due to complex constraint interactions")

                                st.markdown("### Recommended Actions")
                                actions = []

                                if analysis['qualified_trainers'] < 2:
                                    actions.append("✅ Add more qualified trainers for this course")
                                if analysis['duration'] > 3 and analysis['weeks_with_enough_days'] < 10:
                                    actions.append(
                                        "✅ Review public holidays that may be limiting available working days")
                                if analysis.get('restricted_week_positions'):
                                    actions.append(
                                        "✅ Relax week position restrictions (use 'Disable All Week Restrictions' button)")
                                if analysis['weeks_with_any_trainer'] < 5:
                                    actions.append("✅ Review trainer leave patterns to reduce overlapping leaves")

                                # Always suggest prioritizing courses
                                actions.append(
                                    "✅ Enable 'Prioritize scheduling all courses' option in Optimization Settings")

                                for action in actions:
                                    st.markdown(action)
                else:
                    st.info(
                        "No unscheduled courses to analyze. Run an optimization first to identify scheduling problems.")

            # And then modify how it's used in tab4:
            # Instead of:
            # with st.expander("Course Scheduling Problem Analysis", expanded=False):
            #     display_scheduling_problem_analysis()

            # Replace with inline code:
            with st.expander("Course Scheduling Problem Analysis", expanded=False):
                st.subheader("Analyze Unscheduled Course")

                if hasattr(st.session_state, 'unscheduled_courses') and st.session_state.unscheduled_courses:
                    # Create dropdown to select unscheduled course
                    course_options = [f"{c['Course']} (Run {c['Run']}, {c['Language']})" for c in
                                      st.session_state.unscheduled_courses]
                    selected_course = st.selectbox("Select course to analyze", course_options)

                    if selected_course:
                        # Extract course details from selection
                        idx = course_options.index(selected_course)
                        course_info = st.session_state.unscheduled_courses[idx]

                        # Run analysis
                        analysis = st.session_state.scheduler.analyze_unscheduled_course(
                            course_info['Course'], course_info['Language'], course_info['Run'])

                        # Display results
                        col1, col2 = st.columns(2)

                        with col1:
                            st.markdown("### Course Information")
                            st.markdown(f"**Course**: {course_info['Course']}")
                            st.markdown(f"**Language**: {course_info['Language']}")
                            st.markdown(f"**Run**: {course_info['Run']}")
                            st.markdown(f"**Qualified Trainers**: {analysis['qualified_trainers']}")
                            st.markdown(f"**Course Duration**: {analysis['duration']} days")

                            if analysis['qualified_trainers'] == 0:
                                st.error("❌ No qualified trainers for this course")
                                st.markdown("**Solution**: Add more qualified trainers in the Fleximatrix")

                        with col2:
                            st.markdown("### Week Availability")
                            st.markdown(f"**Weeks with enough workdays**: {analysis['weeks_with_enough_days']}")

                            if 'restricted_week_positions' in analysis:
                                restrictions = analysis['restricted_week_positions']
                                if restrictions:
                                    st.markdown(f"**Restricted week positions**: {', '.join(restrictions)}")
                                    st.markdown(f"**Weeks eliminated by restrictions**: {analysis['weeks_restricted']}")

                            st.markdown(f"**Weeks with any trainer available**: {analysis['weeks_with_any_trainer']}")

                            if analysis['weeks_with_any_trainer'] == 0:
                                st.error("❌ No weeks with both enough workdays and available trainers")
                                st.markdown("**Solution**: Reduce leave overlaps or add more qualified trainers")

                        # Draw a timeline of week availability
                        if analysis['qualified_trainers'] > 0:
                            st.markdown("### Week-by-Week Availability")

                            # Create dataframe for visualization
                            weeks = sorted(analysis.get('weeks_with_trainers', {}).keys())
                            if weeks:
                                week_data = []

                                for week in weeks:
                                    trainers_available = analysis['weeks_with_trainers'].get(week, 0)
                                    month = st.session_state.scheduler.week_to_month_map.get(week, 0)
                                    week_info = st.session_state.scheduler.week_position_in_month.get(week, {})

                                    # Check if this week is restricted
                                    is_restricted = False
                                    if course_info['Course'] in st.session_state.scheduler.week_restrictions:
                                        restrictions = st.session_state.scheduler.week_restrictions[
                                            course_info['Course']]
                                        if (week_info.get('is_first') and restrictions.get('First', False)) or \
                                                (week_info.get('is_second') and restrictions.get('Second', False)) or \
                                                (week_info.get('is_third') and restrictions.get('Third', False)) or \
                                                (week_info.get('is_fourth') and restrictions.get('Fourth', False)) or \
                                                (week_info.get('is_last') and restrictions.get('Last', False)):
                                            is_restricted = True

                                    working_days = st.session_state.scheduler.weekly_working_days.get(week, 0)

                                    week_data.append({
                                        'Week': week,
                                        'Month': month,
                                        'Trainers Available': trainers_available,
                                        'Working Days': working_days,
                                        'Is Restricted': is_restricted
                                    })

                                week_df = pd.DataFrame(week_data)

                                # Color columns based on conditions
                                def color_trainers(val):
                                    if val == 0:
                                        color = 'red'
                                    elif val < 3:
                                        color = 'orange'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_days(val, duration=analysis['duration']):
                                    if val < duration:
                                        color = 'red'
                                    else:
                                        color = 'green'
                                    return f'background-color: {color}'

                                def color_restricted(val):
                                    return 'background-color: red' if val else 'background-color: white'

                                # Apply styling
                                styled_df = week_df.style.applymap(color_trainers, subset=['Trainers Available']) \
                                    .applymap(lambda x: color_days(x, analysis['duration']), subset=['Working Days']) \
                                    .applymap(color_restricted, subset=['Is Restricted'])

                                st.dataframe(styled_df)

                                # Identify the problem
                                problems = []
                                if analysis['weeks_with_enough_days'] == 0:
                                    problems.append("❌ No weeks have enough working days for this course duration")
                                if analysis.get('weeks_restricted') == analysis['weeks_with_enough_days']:
                                    problems.append("❌ All potential weeks are restricted by week position constraints")
                                if analysis['weeks_with_any_trainer'] == 0:
                                    problems.append("❌ No trainers are available in valid weeks")

                                st.markdown("### Summary")

                                if problems:
                                    for problem in problems:
                                        st.error(problem)
                                else:
                                    st.warning(
                                        "⚠️ This course may be unscheduled due to complex constraint interactions")

                                st.markdown("### Recommended Actions")
                                actions = []

                                if analysis['qualified_trainers'] < 2:
                                    actions.append("✅ Add more qualified trainers for this course")
                                if analysis['duration'] > 3 and analysis['weeks_with_enough_days'] < 10:
                                    actions.append(
                                        "✅ Review public holidays that may be limiting available working days")
                                if analysis.get('restricted_week_positions'):
                                    actions.append(
                                        "✅ Relax week position restrictions (use 'Disable All Week Restrictions' button)")
                                if analysis['weeks_with_any_trainer'] < 5:
                                    actions.append("✅ Review trainer leave patterns to reduce overlapping leaves")

                                # Always suggest prioritizing courses
                                actions.append(
                                    "✅ Enable 'Prioritize scheduling all courses' option in Optimization Settings")

                                for action in actions:
                                    st.markdown(action)
                else:
                    st.info(
                        "No unscheduled courses to analyze. Run an optimization first to identify scheduling problems.")
    except Exception as e:
        logging.error(f"Application error: {e}")
        st.error(f"An error occurred: {e}")


if __name__ == "__main__":
    # Add this before your main() call
    st.write("App is running")
    main()

